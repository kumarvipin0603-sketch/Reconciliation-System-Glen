"""
Standalone Amazon Reconciliation Telegram Bot.

Setup:
    1. Paste your Telegram bot token in BOT_TOKEN below.
    2. Install dependencies:
       pip install pandas openpyxl python-telegram-bot
    3. Run:
       python Amazon_Reconciliation_Telegram_Bot.py

The bot accepts one .xlsx workbook and returns a formatted reconciliation report.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from telegram import Update
    from telegram.ext import (
        Application,
        CommandHandler,
        ContextTypes,
        MessageHandler,
        filters,
    )
except ImportError:
    # Telegram is not required by the Streamlit website.
    Update = object
    Application = CommandHandler = ContextTypes = MessageHandler = None
    filters = None

# ==========================================================
# PASTE YOUR TELEGRAM BOT TOKEN BETWEEN THE QUOTES BELOW
# ==========================================================
BOT_TOKEN = "DISABLED_IN_WEB_APP"

# Example only:
# BOT_TOKEN = "123456789:AAxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"

OUTPUT_FOLDER = Path("Output")
MAX_FILE_SIZE_MB = 200

LOGGER = logging.getLogger("reconciliation")

REQUIRED_COLUMNS = {
    "Orders": {
        "Amazon Order Id",
        "Shipment Date",
        "Shipped Quantity",
        "Item Price",
        "Branch Code",
    },
    "Reverse": {
        "Return Type",
        "Return Created Date",
        "Order ID",
        "Courier & Customer Return Qty",
    },
    "Payments": {
        "Order ID",
        "Type",
        "Rece Amount",
        "Transaction Status",
    },
    "Sales": {
        "Invoice No",
        "Invoice Date",
        "Po Number",
        "Quantity",
        "Item Price",
        "Branch Code",
        "Sale/Return",
    },
    "Replacement_Map": {
        "Replacement Order Id",
        "Original Order Id",
        # Workbook uses Quantity; code normalises it to Repl Quantity.
        "Quantity",
    },
    "Reimbursement": {
        "Order Id",
        "Reimbursement Amount",
        "Reason",
    },
}

AMOUNT_TOLERANCE = 1.00
QTY_TOLERANCE = 0.000001


def clean_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Trim headers and normalize current FY source aliases."""
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(col)).strip() for col in df.columns]

    # Current Amazon source may call AA "Payment Received".
    # The proven reconciliation engine internally uses "Rece Amount".
    if "Payment Received" in df.columns and "Rece Amount" not in df.columns:
        df = df.rename(columns={"Payment Received": "Rece Amount"})

    return df


def clean_id(series: pd.Series) -> pd.Series:
    """Convert mixed Excel IDs to stable strings and remove trailing .0."""
    cleaned = series.fillna("").astype(str).str.strip()
    return cleaned.str.replace(r"\.0$", "", regex=True)


def numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def unique_join(values: Iterable[object]) -> str:
    items: set[str] = set()
    for value in values:
        if pd.isna(value):
            continue
        for part in str(value).split(","):
            part = part.strip()
            if part and part.lower() != "nan":
                items.add(part)
    return ", ".join(sorted(items))


def validate_excel(file_path: str | Path) -> None:
    """Raise a clear ValueError when a sheet or column is missing."""
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    with pd.ExcelFile(file_path) as xl:
        available_sheets = set(xl.sheet_names)
        missing_sheets = set(REQUIRED_COLUMNS) - available_sheets
        if missing_sheets:
            raise ValueError(
                "Missing sheet(s): " + ", ".join(sorted(missing_sheets))
            )

        problems: list[str] = []
        for sheet, expected in REQUIRED_COLUMNS.items():
            sample = clean_headers(pd.read_excel(xl, sheet_name=sheet, nrows=0))
            missing = expected - set(sample.columns)
            if missing:
                problems.append(f"{sheet}: {', '.join(sorted(missing))}")

        if problems:
            raise ValueError("Missing required column(s):\n- " + "\n- ".join(problems))


def build_order_map(
    order_ids: Iterable[str], replacement_summary: pd.DataFrame
) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {order_id: [order_id] for order_id in order_ids}
    for row in replacement_summary.itertuples(index=False):
        original = row._asdict()["Original_Order_Id"]
        replacements = [
            value.strip()
            for value in str(row._asdict()["Replacement_Order_Id"]).split(",")
            if value.strip()
        ]
        mapping.setdefault(original, [original])
        mapping[original].extend(
            value for value in replacements if value not in mapping[original]
        )
    return mapping


def get_remarks(row: pd.Series) -> str:
    """Return one clear business remark using the reviewed accounting hierarchy.

    Priority:
    1. A full sales return/credit note closes the billed quantity.
    2. Refund/negative settlement without a full CN remains CN Pending.
    3. Positive receipt of 50% or less is Short Payment Received, except
       unresolved replacement cases where the CN is still pending.
    4. Positive receipt above 50% of Net Sale Item Price is Reconciled.
    5. Operational return without a posted sales return remains CN Pending
       only when the payment threshold has not been crossed.
    """
    payment = round(float(row["Rece Amount"]), 2)
    reimbursement = round(float(row["Reimbursement Amount"]), 2)
    billed_amount = abs(round(float(row["Net Sale Item Price"]), 2))

    sale_qty = float(row["Sale Qty"])
    sales_return_qty = float(row["Return Qty"])
    net_order = float(row["Net Order Qty"])
    net_billed = float(row["Net Billed Qty"])
    courier_return_qty = float(row["Courier & Customer Return Qty"])

    replacement_order_id = str(row.get("Replacement Order Id", "") or "").strip()

    has_sale = sale_qty > QTY_TOLERANCE
    has_sales_return = sales_return_qty > QTY_TOLERANCE
    has_full_sales_return = (
        has_sale
        and has_sales_return
        and sales_return_qty + QTY_TOLERANCE >= sale_qty
    )
    has_partial_sales_return = (
        has_sale
        and has_sales_return
        and sales_return_qty + QTY_TOLERANCE < sale_qty
    )
    has_operational_return = courier_return_qty > QTY_TOLERANCE
    has_replacement = bool(replacement_order_id)

    payment_received = payment > AMOUNT_TOLERANCE
    payment_refunded = payment < -AMOUNT_TOLERANCE
    has_payment_activity = payment_received or payment_refunded

    positive_payment_ratio = (
        max(payment, 0.0) / billed_amount
        if billed_amount > AMOUNT_TOLERANCE
        else 0.0
    )

    # Highest priority: Sale Qty and Return Qty are fully matched.
    # The credit note has already been processed, so settlement reversals or
    # reimbursement entries must not change the order back to CN Pending.
    if has_full_sales_return:
        return "Reconciled"

    # Multi-quantity billed order with a major negative settlement and no CN.
    # Keep both issues visible in one remark.
    if (
        has_sale
        and not has_sales_return
        and payment_refunded
        and sale_qty >= 4 - QTY_TOLERANCE
        and billed_amount > AMOUNT_TOLERANCE
        and abs(payment) / billed_amount > 0.50
    ):
        return "CN Pending & Extra Payment Deducted"

    # A refund/reversal exists but the credit note has not been fully processed.
    if has_sale and payment_refunded:
        return "CN Pending"

    # Reviewed replacement cases: where no CN is posted against the replacement
    # billing, a receipt of 50% or less still represents CN Pending.
    if (
        has_sale
        and has_replacement
        and not has_sales_return
        and payment_received
        and positive_payment_ratio <= 0.50
    ):
        return "CN Pending"

    # Positive payment up to 50% against an open billed quantity is a short
    # payment case. This includes partial-CN orders where an outstanding billed
    # quantity remains.
    if (
        has_sale
        and not has_full_sales_return
        and payment_received
        and positive_payment_ratio <= 0.50
    ):
        return "Short Payment Received"

    # Approved payment-closure rule: once positive payment exceeds 50% of
    # Net Sale Item Price, the order is treated as financially reconciled.
    # This check must run before courier-return, partial-CN, replacement, or
    # reimbursement checks; otherwise valid paid orders are wrongly marked
    # CN Pending. Refund/negative-payment cases were already handled above.
    if (
        has_sale
        and payment_received
        and billed_amount > AMOUNT_TOLERANCE
        and positive_payment_ratio > 0.50
    ):
        return "Reconciled"

    # Physical/courier return exists but no complete credit note is visible
    # and payment has not crossed the approved 50% reconciliation threshold.
    if has_sale and has_operational_return and not has_full_sales_return:
        return "CN Pending"

    # A partial sales return still leaves an open billed quantity where the
    # payment threshold above has not been met.
    if has_partial_sales_return:
        return "CN Pending"

    # Positive accounting activity without an open return/CN issue.
    if payment_received:
        remarks = ["Reconciled"]
        if reimbursement > AMOUNT_TOLERANCE:
            remarks.append("Reimbursement Received")
        return "; ".join(remarks)

    # Sale exists but neither payment nor CN/refund exists yet.
    if has_sale and not has_sales_return and not has_payment_activity:
        if has_operational_return or reimbursement > AMOUNT_TOLERANCE:
            return "CN Pending"
        return "Payment Pending"

    # CN exists without a matching sale/payment entry.
    if has_sales_return and not has_sale and not has_payment_activity:
        remarks = ["Credit Note Processed - Billing Entry Review"]
        if reimbursement > AMOUNT_TOLERANCE:
            remarks.append("Reimbursement Received")
        return "; ".join(remarks)

    # No billing, CN, payment, or refund activity on an open order remains
    # Billing & Payment Pending.  Age alone must never convert a zero-payment
    # order to Reconciled.
    if (
        not has_sale
        and not has_sales_return
        and not has_payment_activity
        and net_order > QTY_TOLERANCE
    ):
        return "Billing & Payment Pending"

    if abs(net_order) <= QTY_TOLERANCE and abs(net_billed) <= QTY_TOLERANCE:
        remarks = ["Reconciled"]
        if reimbursement > AMOUNT_TOLERANCE:
            remarks.append("Reimbursement Received")
        return "; ".join(remarks)

    if reimbursement > AMOUNT_TOLERANCE:
        return "Reconciled; Reimbursement Received"

    return "Reconciled"


def process_reconciliation(
    file_path: str | Path,
    output_file: str | Path | None = None,
) -> Path:
    file_path = Path(file_path)
    output_path = (
        Path(output_file)
        if output_file
        else file_path.with_name("Amazon_Reconciliation.xlsx")
    )

    validate_excel(file_path)
    LOGGER.info("Reading workbook: %s", file_path)

    with pd.ExcelFile(file_path) as xl:
        orders = clean_headers(pd.read_excel(xl, sheet_name="Orders"))
        reverse = clean_headers(pd.read_excel(xl, sheet_name="Reverse"))
        payments = clean_headers(pd.read_excel(xl, sheet_name="Payments"))
        sales = clean_headers(pd.read_excel(xl, sheet_name="Sales"))
        replacement = clean_headers(pd.read_excel(xl, sheet_name="Replacement_Map"))
        reimbursement = clean_headers(pd.read_excel(xl, sheet_name="Reimbursement"))

    # ---------- Normalisation ----------
    orders["Amazon Order Id"] = clean_id(orders["Amazon Order Id"])
    reverse["Order ID"] = clean_id(reverse["Order ID"])
    payments["Order ID"] = clean_id(payments["Order ID"])
    sales["Po Number"] = clean_id(sales["Po Number"])
    replacement["Original Order Id"] = clean_id(replacement["Original Order Id"])
    replacement["Replacement Order Id"] = clean_id(
        replacement["Replacement Order Id"]
    )
    reimbursement["Order Id"] = clean_id(reimbursement["Order Id"])

    replacement = replacement.rename(columns={"Quantity": "Repl Quantity"})

    orders["Shipped Quantity"] = numeric(orders["Shipped Quantity"])
    orders["Item Price"] = numeric(orders["Item Price"])
    reverse["Courier & Customer Return Qty"] = numeric(
        reverse["Courier & Customer Return Qty"]
    ).abs()
    payments["Rece Amount"] = numeric(payments["Rece Amount"])
    sales["Quantity"] = numeric(sales["Quantity"])
    sales["Item Price"] = numeric(sales["Item Price"])
    replacement["Repl Quantity"] = numeric(replacement["Repl Quantity"]).abs()
    reimbursement["Reimbursement Amount"] = numeric(
        reimbursement["Reimbursement Amount"]
    )

    orders["Shipment Date"] = pd.to_datetime(
        orders["Shipment Date"], errors="coerce", utc=True
    ).dt.tz_localize(None)
    reverse["Return Created Date"] = pd.to_datetime(
        reverse["Return Created Date"], errors="coerce"
    )
    sales["Invoice Date"] = pd.to_datetime(sales["Invoice Date"], errors="coerce")

    # ---------- Orders ----------
    orders = orders[orders["Amazon Order Id"] != ""].copy()
    orders_summary = (
        orders.groupby("Amazon Order Id", as_index=False)
        .agg(
            {
                "Shipment Date": "min",
                "Branch Code": unique_join,
                "Shipped Quantity": "sum",
                "Item Price": "sum",
            }
        )
        .rename(columns={"Branch Code": "Order Branch Code"})
    )

    # ---------- Replacement mapping ----------
    replacement = replacement[
        (replacement["Original Order Id"] != "")
        & (replacement["Replacement Order Id"] != "")
    ].copy()

    # Replacement order IDs must never appear as separate primary rows in
    # column A. Their activity is rolled up to the original order instead.
    replacement_ids = set(replacement["Replacement Order Id"].tolist())
    orders_summary = orders_summary[
        ~orders_summary["Amazon Order Id"].isin(replacement_ids)
    ].copy()

    replacement_summary = (
        replacement.groupby("Original Order Id", as_index=False)
        .agg(
            {
                "Replacement Order Id": unique_join,
                "Repl Quantity": "sum",
            }
        )
    )

    # itertuples sanitises spaces to underscores; rename explicitly for stability.
    replacement_for_map = replacement_summary.rename(
        columns={
            "Original Order Id": "Original_Order_Id",
            "Replacement Order Id": "Replacement_Order_Id",
        }
    )
    order_map = build_order_map(
        orders_summary["Amazon Order Id"].tolist(), replacement_for_map
    )

    reco = orders_summary.merge(
        replacement_summary,
        how="left",
        left_on="Amazon Order Id",
        right_on="Original Order Id",
    ).drop(columns=["Original Order Id"], errors="ignore")
    reco["Replacement Order Id"] = reco["Replacement Order Id"].fillna("")
    reco["Repl Quantity"] = numeric(reco["Repl Quantity"])

    # ---------- Reverse / returns ----------
    reverse = reverse[reverse["Order ID"] != ""].copy()
    reverse_summary = (
        reverse.groupby("Order ID", as_index=False)
        .agg(
            {
                "Courier & Customer Return Qty": "sum",
                "Return Type": unique_join,
                "Return Created Date": "max",
            }
        )
        .set_index("Order ID")
        .to_dict("index")
    )

    # ---------- Sales ----------
    sales = sales[sales["Po Number"] != ""].copy()
    sales["Sale/Return"] = sales["Sale/Return"].fillna("").astype(str).str.strip().str.upper()
    sales_summary = (
        sales.groupby(["Po Number", "Sale/Return"], as_index=False)
        .agg({"Quantity": "sum", "Item Price": "sum"})
    )

    sales_dict: dict[str, dict[str, float]] = {}
    for row in sales_summary.itertuples(index=False):
        order_id = str(row[0])
        transaction_type = str(row[1])
        quantity = float(row[2])
        amount = float(row[3])
        bucket = sales_dict.setdefault(
            order_id, {"SALE_QTY": 0.0, "RETURN_QTY": 0.0, "NET_AMOUNT": 0.0}
        )
        if transaction_type == "SALE":
            bucket["SALE_QTY"] += abs(quantity)
        elif transaction_type == "RETURN":
            bucket["RETURN_QTY"] += abs(quantity)
        bucket["NET_AMOUNT"] += amount  # return values are already negative

    # ---------- Payments ----------
    # Current Amazon source:
    # D  = Order ID
    # AA = Payment Received / Rece Amount
    # AB = Transaction Status
    payments = payments[payments["Order ID"] != ""].copy()
    payments["Transaction Status"] = (
        payments["Transaction Status"].fillna("").astype(str).str.strip().str.upper()
    )

    released = payments[
        payments["Transaction Status"].str.contains("RELEASED", na=False)
    ].copy()
    deferred = payments[
        payments["Transaction Status"].str.contains("DEFERRED", na=False)
    ].copy()

    payment_dict = released.groupby("Order ID")["Rece Amount"].sum().to_dict()
    deferred_dict = deferred.groupby("Order ID")["Rece Amount"].sum().to_dict()

    # ---------- Reimbursements ----------
    reimbursement = reimbursement[reimbursement["Order Id"] != ""].copy()
    reimbursement["Reason"] = reimbursement["Reason"].fillna("").astype(str).str.strip()
    reimbursement_dict = (
        reimbursement.groupby("Order Id", as_index=False)
        .agg({"Reimbursement Amount": "sum", "Reason": unique_join})
        .set_index("Order Id")
        .to_dict("index")
    )

    # ---------- Combine original and replacement order activity ----------
    result_rows: list[dict[str, object]] = []
    for row in reco.to_dict("records"):
        original = row["Amazon Order Id"]
        related_orders = order_map.get(original, [original])

        total_return_qty = 0.0
        return_types: list[str] = []
        latest_return_date = pd.NaT
        sale_qty = 0.0
        return_qty = 0.0
        net_sale_amount = 0.0
        received_amount = 0.0
        deferred_amount = 0.0
        reimbursement_amount = 0.0
        reimbursement_reasons: list[str] = []

        for order_id in related_orders:
            return_data = reverse_summary.get(order_id)
            if return_data:
                total_return_qty += float(
                    return_data["Courier & Customer Return Qty"]
                )
                return_types.append(str(return_data["Return Type"]))
                current_date = return_data["Return Created Date"]
                if pd.notna(current_date) and (
                    pd.isna(latest_return_date) or current_date > latest_return_date
                ):
                    latest_return_date = current_date

            sale_data = sales_dict.get(order_id)
            if sale_data:
                sale_qty += sale_data["SALE_QTY"]
                return_qty += sale_data["RETURN_QTY"]
                net_sale_amount += sale_data["NET_AMOUNT"]

            # Roll both RELEASED and DEFERRED across original + replacement IDs.
            received_amount += float(payment_dict.get(order_id, 0.0))
            deferred_amount += float(deferred_dict.get(order_id, 0.0))

            reimb_data = reimbursement_dict.get(order_id)
            if reimb_data:
                reimbursement_amount += float(reimb_data["Reimbursement Amount"])
                reimbursement_reasons.append(str(reimb_data["Reason"]))

        row.update(
            {
                "Courier & Customer Return Qty": total_return_qty,
                "Return Type": unique_join(return_types),
                "Return Created Date": latest_return_date,
                "Sale Qty": sale_qty,
                "Return Qty": return_qty,
                "Net Sale Item Price": net_sale_amount,
                "Rece Amount": received_amount,
                "Deferred Amount": deferred_amount,
                "Reimbursement Amount": reimbursement_amount,
                "Reason": unique_join(reimbursement_reasons),
            }
        )
        result_rows.append(row)

    final_df = pd.DataFrame(result_rows)

    # The source workbook carries item value only against the original order.
    # Apply the original order's average unit value to replacement quantity.
    # This creates a replacement value and a combined net item value without
    # placing replacement IDs in column A.
    shipped_qty_safe = final_df["Shipped Quantity"].replace(0, pd.NA)
    final_df["Original Unit Item Price"] = (
        final_df["Item Price"] / shipped_qty_safe
    ).fillna(0.0)
    final_df["Replacement Item Price"] = (
        final_df["Original Unit Item Price"] * final_df["Repl Quantity"]
    )
    final_df["Net Item Price"] = (
        final_df["Item Price"] + final_df["Replacement Item Price"]
    )

    final_df["Net Order Qty"] = (
        final_df["Shipped Quantity"]
        + final_df["Repl Quantity"]
        - final_df["Courier & Customer Return Qty"]
    )
    final_df["Net Billed Qty"] = final_df["Sale Qty"] - final_df["Return Qty"]
    final_df["Qty Diff"] = final_df["Net Order Qty"] - final_df["Net Billed Qty"]
    final_df["Item Price Diff"] = (
        final_df["Net Item Price"] - final_df["Net Sale Item Price"]
    )
    final_df["Remarks"] = final_df.apply(get_remarks, axis=1)

    output_columns = [
        "Amazon Order Id",
        "Replacement Order Id",
        "Shipment Date",
        "Order Branch Code",
        "Shipped Quantity",
        "Repl Quantity",
        "Item Price",
        "Original Unit Item Price",
        "Replacement Item Price",
        "Net Item Price",
        "Courier & Customer Return Qty",
        "Return Type",
        "Return Created Date",
        "Net Order Qty",
        "Sale Qty",
        "Return Qty",
        "Net Billed Qty",
        "Net Sale Item Price",
        "Rece Amount",
        "Deferred Amount",
        "Reimbursement Amount",
        "Reason",
        "Qty Diff",
        "Item Price Diff",
        "Remarks",
    ]
    final_df = final_df[output_columns].copy()

    for column in [
        "Item Price",
        "Original Unit Item Price",
        "Replacement Item Price",
        "Net Item Price",
        "Net Sale Item Price",
        "Rece Amount",
        "Deferred Amount",
        "Reimbursement Amount",
        "Item Price Diff",
    ]:
        final_df[column] = numeric(final_df[column]).round(2)

    summary = (
        final_df.groupby("Remarks", dropna=False)
        .size()
        .reset_index(name="Order Count")
        .sort_values("Order Count", ascending=False)
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Reconciliation", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

    format_output(output_path)
    LOGGER.info("Output created: %s", output_path)
    return output_path


def format_output(output_path: Path) -> None:
    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 38)

    reco_ws = wb["Reconciliation"]
    header_map = {cell.value: cell.column for cell in reco_ws[1]}
    for name in ["Shipment Date", "Return Created Date"]:
        column = header_map.get(name)
        if column:
            for cell in reco_ws.iter_cols(
                min_col=column, max_col=column, min_row=2
            ):
                for item in cell:
                    item.number_format = "dd-mmm-yyyy"

    for name in [
        "Item Price",
        "Original Unit Item Price",
        "Replacement Item Price",
        "Net Item Price",
        "Net Sale Item Price",
        "Rece Amount",
        "Reimbursement Amount",
        "Item Price Diff",
    ]:
        column = header_map.get(name)
        if column:
            for cell in reco_ws.iter_cols(
                min_col=column, max_col=column, min_row=2
            ):
                for item in cell:
                    item.number_format = '#,##0.00;[Red]-#,##0.00'

    wb.save(output_path)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show the upload instructions."""
    if update.message:
        await update.message.reply_text(
            "Amazon Reconciliation Bot is ready.\n\n"
            "Upload one Excel (.xlsx) workbook containing these sheets:\n"
            "Orders, Reverse, Payments, Sales, Replacement_Map and Reimbursement."
        )


async def receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Download, validate, process and return the uploaded workbook."""
    if not update.message or not update.message.document:
        return

    document = update.message.document
    filename = document.file_name or "uploaded_workbook.xlsx"

    if not filename.lower().endswith(".xlsx"):
        await update.message.reply_text("Please upload an Excel file in .xlsx format.")
        return

    if document.file_size and document.file_size > MAX_FILE_SIZE_MB * 1024 * 1024:
        await update.message.reply_text(
            f"The file is larger than the allowed {MAX_FILE_SIZE_MB} MB limit."
        )
        return

    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename).name)
    unique_prefix = f"{update.message.chat_id}_{document.file_unique_id}"
    input_path = OUTPUT_FOLDER / f"{unique_prefix}_{safe_name}"
    output_path = OUTPUT_FOLDER / f"{unique_prefix}_Amazon_Reconciliation.xlsx"

    await update.message.reply_text("File received. Validating and processing...")

    try:
        telegram_file = await context.bot.get_file(document.file_id)
        await telegram_file.download_to_drive(custom_path=str(input_path))

        result_path = process_reconciliation(input_path, output_path)

        with result_path.open("rb") as report:
            await update.message.reply_document(
                document=report,
                filename="Amazon_Reconciliation.xlsx",
                caption="Reconciliation completed successfully.",
            )

    except (FileNotFoundError, ValueError) as exc:
        LOGGER.warning("Workbook validation failed: %s", exc)
        await update.message.reply_text(f"Validation error:\n{exc}")
    except Exception as exc:
        LOGGER.exception("Unexpected reconciliation error")
        await update.message.reply_text(
            "Processing failed due to an unexpected error.\n"
            f"Technical detail: {type(exc).__name__}: {exc}"
        )
    finally:
        for path in (input_path, output_path):
            try:
                if path.exists():
                    path.unlink()
            except OSError:
                LOGGER.warning("Could not delete temporary file: %s", path)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Log uncaught Telegram errors."""
    LOGGER.exception("Telegram update failed", exc_info=context.error)


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )

    if not BOT_TOKEN or BOT_TOKEN == "PASTE_YOUR_TELEGRAM_BOT_TOKEN_HERE":
        raise SystemExit(
            "Bot token is missing. Open this file and replace "
            "PASTE_YOUR_TELEGRAM_BOT_TOKEN_HERE with the token from @BotFather."
        )

    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    application = Application.builder().token(BOT_TOKEN.strip()).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.Document.ALL, receive_file))
    application.add_error_handler(error_handler)

    print("Amazon Reconciliation Bot started. Press Ctrl+C to stop.")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
