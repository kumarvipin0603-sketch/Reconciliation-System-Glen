"""
Flipkart Reconciliation Telegram Bot
====================================

Update the bot token on this line:

    BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"

Install dependencies:
    pip install pandas openpyxl python-telegram-bot==21.*

Run Telegram bot:
    python flipkart_reconciliation_bot_full.py

Run from command line:
    python flipkart_reconciliation_bot_full.py "Flipkart FY 26-27.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import tempfile
from pathlib import Path
from typing import Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from telegram import Update
    from telegram.ext import (
        Application,
        CommandHandler,
        ContextTypes,
        MessageHandler,
        filters,
    )
except ImportError:
    # Telegram is not required by the Streamlit website.
    Update = object
    Application = CommandHandler = ContextTypes = MessageHandler = None
    filters = None


# ==========================================================
# UPDATE YOUR TELEGRAM BOT TOKEN HERE
# ==========================================================

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "DISABLED_IN_WEB_APP")

# Example:
# BOT_TOKEN = "7845123456:AAxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"


# ==========================================================
# SETTINGS
# ==========================================================

OUTPUT_FILE_NAME = "Flipkart_Reconciliation_With_Remarks.xlsx"
REMARK_LOGIC_VERSION = "2026-07-22-CANCELLATION-SALE-AS-DELIVERED"

AMOUNT_TOLERANCE = 1.00
QUANTITY_TOLERANCE = 0.0001
MIN_PAYMENT_RECEIPT_RATIO = 0.50  # 80% of Net ERP Billing

REMARK_PRIORITY = [
    "Extra Billing",
    "CN & Payment Pending",
    "CN Pending",
    "Billing & Payment Pending",
    "Billing Pending",
    "Payment Pending",
    "Short Payment Received",
    "Order Details Review",
    "Reconciled",
]


# A PO with Flipkart Return Qty > 0 and Net ERP Qty = 0
# is treated as Reconciled. Status order does not matter; for example,
# "DELIVERED, RETURNED" and "RETURNED, DELIVERED" are equivalent.
ZERO_NET_ERP_RECONCILED_STATUSES = {
    frozenset({"RETURNED"}),
    frozenset({"CANCELLED"}),
    frozenset({"DELIVERED", "RETURNED"}),
    frozenset({"CANCELLED", "RETURNED"}),
    frozenset({"RETURN_REQUESTED"}),
    frozenset({"CANCELLED", "DELIVERED"}),
    frozenset({"REJECTED"}),
    frozenset({"REJECTED", "RETURNED"}),
    frozenset({"CANCELLED", "RETURN_REQUESTED"}),
    frozenset({"CANCELLED", "DELIVERED", "RETURNED"}),
    frozenset({"DELIVERED", "REJECTED", "RETURNED"}),
    frozenset({"CANCELLED", "DELIVERED", "REJECTED"}),
    frozenset({"RETURNED", "RETURN_REQUESTED"}),
    frozenset({"DELIVERED", "RETURNED", "RETURN_REQUESTED"}),
    frozenset({"DELIVERED", "RETURN_REQUESTED"}),
    frozenset({"READY_TO_SHIP", "RETURN_REQUESTED"}),
    frozenset({"APPROVAL_HOLD", "RETURN_REQUESTED"}),
    frozenset({"APPROVED", "RETURN_REQUESTED"}),
    frozenset({"APPROVED", "CANCELLED"}),
    frozenset({"READY_TO_SHIP", "RETURNED"}),
    frozenset({"APPROVAL_HOLD"}),
}


def normalize_order_status_set(value: object) -> frozenset[str]:
    """Return a status combination as an order-independent normalized set."""
    if value is None or pd.isna(value):
        return frozenset()

    statuses = {
        re.sub(r"\s+", "_", part.strip().upper())
        for part in str(value).split(",")
        if part.strip()
    }
    return frozenset(statuses)


def normalize_status_parts(value: object) -> tuple[str, ...]:
    """Normalize comma-separated status values for mapping lookup."""
    if value is None or pd.isna(value):
        return tuple()
    return tuple(sorted({
        re.sub(r"\s+", " ", part.strip()).upper()
        for part in str(value).split(",")
        if part.strip()
    }))


STATUS_CONSIDERATION_MAP = {
    (('RETURN_REQUESTED',), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED',), (), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('DELIVERED', 'RETURNED'), (), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'INIT')): 'Delivered',
    (('DELIVERED',), ('RETURN', 'RETURN CANCELLATION'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('RETURN_REQUESTED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('DELIVERED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('DELIVERED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('RETURN_REQUESTED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'DELIVERED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('RETURNED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('DELIVERED', 'RETURNED', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'COMPLETED', 'INIT')): 'Delivered',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED', 'INIT')): 'Delivered',
    (('RETURN_REQUESTED',), ('CANCELLATION', 'RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVAL_HOLD', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Delivered',
    (('READY_TO_SHIP',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Delivered',
    (('READY_TO_SHIP',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURN_REQUESTED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('READY_TO_SHIP',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('READY_TO_SHIP',), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('APPROVED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('APPROVED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('APPROVED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('APPROVED', 'READY_TO_SHIP'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('APPROVED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVED',), ('SALE',), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Delivered',
    (('READY_TO_SHIP', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURN_REQUESTED',), ('SALE',), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('DELIVERED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURN_REQUESTED',), ('SALE',), ('COURIER_RETURN',), ('INIT',)): 'Delivered',
    (('RETURNED',), (), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURN_REQUESTED'), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURN_REQUESTED',), (), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED', 'INIT')): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURN_REQUESTED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Delivered',
    (('RETURNED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned',
    (('RETURNED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('DELIVERED',), ('SALE',), (), ()): 'Delivered',
    (('DELIVERED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Delivered',
    (('APPROVED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'INIT')): 'Delivered',
    (('READY_TO_SHIP', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('READY_TO_SHIP', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('INIT',)): 'Returned',
    (('RETURN_REQUESTED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('APPROVED',), ('SALE',), ('COURIER_RETURN',), ('INIT',)): 'Delivered',
    (('APPROVED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Returned',
    (('READY_TO_SHIP', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'DELIVERED'), (), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('CANCELLED', 'DELIVERED'), ('RETURN', 'RETURN CANCELLATION'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('CANCELLED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'DELIVERED', 'RETURNED'), ('CANCELLATION', 'RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Returned',
    (('CANCELLED', 'DELIVERED', 'REJECTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED',)): 'Delivered',
    (('CANCELLED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('APPROVED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURN_REQUESTED'), ('SALE',), ('COURIER_RETURN',), ('INIT',)): 'Delivered',
    (('CANCELLED', 'RETURN_REQUESTED'), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'DELIVERED'), ('RETURN CANCELLATION',), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('CANCELLED', 'DELIVERED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVAL_HOLD', 'RETURN_REQUESTED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('APPROVED', 'RETURN_REQUESTED'), ('SALE',), ('CUSTOMER_RETURN',), ('INIT',)): 'Delivered',
    (('CANCELLED', 'RETURN_REQUESTED'), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('APPROVED', 'RETURN_REQUESTED'), ('SALE',), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED',), (), (), ()): 'Cancelled',
    (('REJECTED',), (), (), ()): 'Cancelled',
    (('CANCELLED',), ('CANCELLATION', 'SALE'), (), ()): 'Cancelled',
    (('APPROVED',), (), (), ()): 'Delivered',
    (('APPROVED',), ('SALE',), (), ()): 'Delivered',
    (('READY_TO_SHIP',), ('SALE',), (), ()): 'Delivered',
    (('READY_TO_SHIP',), (), (), ()): 'Delivered',
    (('DELIVERED',), (), (), ()): 'Delivered',
    ((), ('RETURN', 'RETURN CANCELLATION'), (), ()): 'Delivered',
    (('APPROVED',), ('CANCELLATION', 'SALE'), (), ()): 'Returned',
    (('APPROVAL_HOLD',), (), (), ()): 'Delivered',
    (('RETURN_REQUESTED',), (), (), ()): 'Delivered',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVAL_HOLD', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'READY_TO_SHIP'), ('SALE',), (), ()): 'Delivered',
    (('APPROVAL_HOLD', 'RETURN_REQUESTED'), ('SALE',), (), ()): 'Delivered',
    (('RETURN_REQUESTED',), ('SALE',), (), ()): 'Delivered',
    (('READY_TO_SHIP',), ('CANCELLATION', 'SALE'), (), ()): 'Delivered',
    (('DELIVERED',), ('CANCELLATION', 'SALE'), (), ()): 'Delivered',
    (('RETURNED',), (), (), ()): 'Returned',
    (('DELIVERED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'RETURNED'), (), (), ()): 'Delivered',
    (('READY_TO_SHIP', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURNED'), (), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURNED'), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), (), ()): 'Delivered',
    ((), ('RETURN',), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'COMPLETED')): 'Returned & Cancelled',
    (('RETURNED',), (), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), (), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), (), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED',)): 'Returned',
    (('RETURNED',), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), (), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('RETURN',), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN',), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN',), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('RETURN',), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN',), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('REJECTED', 'RETURNED'), ('RETURN',), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'INIT')): 'Returned',
    (('DELIVERED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('DELIVERED', 'REJECTED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED', 'INIT')): 'Returned',
    (('REJECTED', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('RETURNED',), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('RETURNED',), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('DELIVERED', 'REJECTED'), ('SALE',), (), ()): 'Delivered',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'READY_TO_SHIP'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED', 'INIT')): 'Returned',
    (('CANCELLED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVED',), ('SALE',), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVED', 'CANCELLED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned & Cancelled',
    (('CANCELLED', 'RETURNED'), (), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned & Cancelled',
    (('CANCELLED', 'RETURNED'), ('RETURN',), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('DELIVERED', 'RETURNED'), ('SALE',), (), ()): 'Delivered',
    (('RETURNED',), ('RETURN',), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('RETURN',), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned & Cancelled',
    (('RETURN_REQUESTED',), ('RETURN',), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('CANCELLATION', 'RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Returned',
    (('RETURNED',), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('INIT',)): 'Returned',
    (('CANCELLED', 'RETURNED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED', 'INIT')): 'Returned',
    (('RETURNED',), ('CANCELLATION', 'RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('RETURNED',), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('REJECTED', 'RETURNED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('COMPLETED',)): 'Returned',
    (('APPROVED', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('CUSTOMER_RETURN',), ('INIT',)): 'Returned',
    (('READY_TO_SHIP', 'RETURNED'), ('RETURN', 'SALE'), ('COURIER_RETURN',), ('COMPLETED', 'INIT')): 'Returned',
    (('APPROVAL_HOLD', 'RETURN_REQUESTED'), ('RETURN', 'SALE'), ('COURIER_RETURN', 'CUSTOMER_RETURN'), ('COMPLETED', 'INIT')): 'Returned',
    (('RETURN_REQUESTED',), ('CANCELLATION', 'RETURN', 'SALE'), ('COURIER_RETURN',), ('INIT',)): 'Returned',
    (('CANCELLED', 'DELIVERED'), ('RETURN', 'RETURN CANCELLATION', 'SALE'), ('CUSTOMER_RETURN',), ('CANCELLED', 'COMPLETED')): 'Delivered',
}


def get_considered_order_status(row: pd.Series) -> str:
    """Return the business status from the attached status-mapping table."""
    key = (
        normalize_status_parts(row.get("Order Status", "")),
        normalize_status_parts(row.get("Sales Event Sub Type", "")),
        normalize_status_parts(row.get("Return Type", "")),
        normalize_status_parts(row.get("Return Status", "")),
    )
    mapped = STATUS_CONSIDERATION_MAP.get(key)
    if mapped:
        return mapped

    # Conservative fallback for combinations not present in the supplied map.
    raw_statuses = set(normalize_status_parts(row.get("Order Status", "")))
    return_statuses = set(normalize_status_parts(row.get("Return Status", "")))
    event_types = set(normalize_status_parts(row.get("Sales Event Sub Type", "")))

    if "CANCELLED" in raw_statuses or "REJECTED" in raw_statuses:
        if "RETURNED" in raw_statuses:
            return "Returned & Cancelled"
        return "Cancelled"

    if (
        "RETURNED" in raw_statuses
        or "COMPLETED" in return_statuses
        or ("RETURN" in event_types and "RETURN CANCELLATION" not in event_types)
    ):
        return "Returned"

    return "Delivered"


# ==========================================================
# LOGGING
# ==========================================================

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s",
    level=logging.INFO,
)
LOGGER = logging.getLogger(__name__)


# ==========================================================
# HEADER AND DATA HELPERS
# ==========================================================

def normalize_text(value: object) -> str:
    """Normalize sheet names and column names for tolerant matching."""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9%]+", " ", text)
    return text.strip()


def clean_po(value: object) -> Optional[str]:
    """Convert a PO number into a stable string key."""
    if value is None or pd.isna(value):
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    return text if text else None


def to_number(series: pd.Series) -> pd.Series:
    """Convert report values into numbers, treating blanks as zero."""
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def find_sheet_name(
    sheet_names: Iterable[str],
    aliases: Iterable[str],
) -> Optional[str]:
    normalized = {normalize_text(name): name for name in sheet_names}

    for alias in aliases:
        alias_norm = normalize_text(alias)

        if alias_norm in normalized:
            return normalized[alias_norm]

        for norm_name, original_name in normalized.items():
            if alias_norm in norm_name:
                return original_name

    return None


def find_column(
    dataframe: pd.DataFrame,
    aliases: Iterable[str],
    required: bool = True,
) -> Optional[str]:
    normalized_columns = {
        normalize_text(column): column for column in dataframe.columns
    }

    # Exact normalized match first
    for alias in aliases:
        alias_norm = normalize_text(alias)
        if alias_norm in normalized_columns:
            return normalized_columns[alias_norm]

    # Partial match second
    for alias in aliases:
        alias_norm = normalize_text(alias)
        for normalized_column, original_column in normalized_columns.items():
            if alias_norm and alias_norm in normalized_column:
                return original_column

    if required:
        raise KeyError(
            f"Required column not found. Expected one of: {list(aliases)}"
        )

    return None


def prepare_po_column(
    dataframe: pd.DataFrame,
    aliases: Iterable[str] = ("Po Number",),
) -> pd.DataFrame:
    dataframe = dataframe.copy()
    po_column = find_column(dataframe, aliases)
    dataframe["__po__"] = dataframe[po_column].map(clean_po)
    return dataframe[dataframe["__po__"].notna()].copy()


def aggregate_numeric(
    dataframe: pd.DataFrame,
    po_column: str,
    output_name: str,
    source_column: Optional[str],
) -> pd.DataFrame:
    if source_column is None:
        return pd.DataFrame(columns=[po_column, output_name])

    working = dataframe[[po_column, source_column]].copy()
    working[source_column] = to_number(working[source_column])

    return (
        working.groupby(po_column, as_index=False)[source_column]
        .sum()
        .rename(columns={source_column: output_name})
    )


def merge_frames(master: pd.DataFrame, frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    result = master.copy()

    for frame in frames:
        if frame is None or frame.empty:
            continue
        result = result.merge(frame, on="PO Number", how="left")

    return result


# ==========================================================
# SHEET PROCESSORS
# ==========================================================

def process_all_orders(dataframe: pd.DataFrame) -> pd.DataFrame:
    df = prepare_po_column(dataframe)

    qty_column = find_column(df, ("Quantity",))
    status_column = find_column(
        df,
        ("Order Item Status", "order_item_status"),
        required=False,
    )
    order_date_column = find_column(
        df,
        ("Order Date", "order_date"),
        required=False,
    )

    df["Order Qty"] = to_number(df[qty_column])

    output = (
        df.groupby("__po__", as_index=False)
        .agg(
            {
                "Order Qty": "sum",
                **(
                    {status_column: lambda s: ", ".join(
                        sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
                    )}
                    if status_column
                    else {}
                ),
                **(
                    {order_date_column: "min"}
                    if order_date_column
                    else {}
                ),
            }
        )
        .rename(columns={"__po__": "PO Number"})
    )

    if status_column:
        output = output.rename(columns={status_column: "Order Status"})
    else:
        output["Order Status"] = ""

    if order_date_column:
        output = output.rename(columns={order_date_column: "Order Date"})
    else:
        output["Order Date"] = pd.NaT

    return output


def process_sales_report(dataframe: pd.DataFrame) -> pd.DataFrame:
    df = prepare_po_column(dataframe)

    qty_column = find_column(df, ("Quantity",))
    subtype_column = find_column(
        df,
        ("Event Sub Type", "event_sub_type"),
        required=False,
    )
    order_date_column = find_column(
        df,
        ("Order Date", "order_date"),
        required=False,
    )

    df["Sales Report Qty"] = to_number(df[qty_column])

    output = (
        df.groupby("__po__", as_index=False)
        .agg(
            {
                "Sales Report Qty": "sum",
                **(
                    {subtype_column: lambda s: ", ".join(
                        sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
                    )}
                    if subtype_column
                    else {}
                ),
                **(
                    {order_date_column: "min"}
                    if order_date_column
                    else {}
                ),
            }
        )
        .rename(columns={"__po__": "PO Number"})
    )

    if subtype_column:
        output = output.rename(columns={subtype_column: "Sales Event Sub Type"})
    else:
        output["Sales Event Sub Type"] = ""

    if order_date_column:
        output = output.rename(columns={order_date_column: "Sales Report Order Date"})
    else:
        output["Sales Report Order Date"] = pd.NaT

    return output


def process_erp(dataframe: pd.DataFrame) -> pd.DataFrame:
    df = prepare_po_column(dataframe)

    quantity_column = find_column(df, ("Quantity",))
    amount_column = find_column(df, ("Line Amount",), required=False)
    sale_return_column = find_column(df, ("Sale/Return", "Sale Return"))
    invoice_column = find_column(df, ("Invoice No",), required=False)
    invoice_date_column = find_column(df, ("Invoice Date",), required=False)
    branch_column = find_column(df, ("Branch Code",), required=False)

    df["__qty__"] = to_number(df[quantity_column])
    df["__amount__"] = (
        to_number(df[amount_column]) if amount_column else 0.0
    )
    df["__type__"] = df[sale_return_column].astype(str).str.strip().str.lower()

    is_return = df["__type__"].str.contains(
        r"return|credit|cn",
        regex=True,
        na=False,
    )

    sales = df[~is_return].copy()
    returns = df[is_return].copy()

    sales_group = sales.groupby("__po__", as_index=False).agg(
        {
            "__qty__": "sum",
            "__amount__": "sum",
            **(
                {invoice_column: lambda s: ", ".join(
                    sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
                )}
                if invoice_column
                else {}
            ),
            **(
                {invoice_date_column: "min"}
                if invoice_date_column
                else {}
            ),
            **(
                {branch_column: lambda s: ", ".join(
                    sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
                )}
                if branch_column
                else {}
            ),
        }
    )

    sales_group = sales_group.rename(
        columns={
            "__po__": "PO Number",
            "__qty__": "ERP Billed Qty",
            "__amount__": "ERP Billing Amount",
        }
    )

    if invoice_column:
        sales_group = sales_group.rename(
            columns={invoice_column: "Invoice No"}
        )
    else:
        sales_group["Invoice No"] = ""

    if invoice_date_column:
        sales_group = sales_group.rename(
            columns={invoice_date_column: "Invoice Date"}
        )
    else:
        sales_group["Invoice Date"] = pd.NaT

    if branch_column:
        sales_group = sales_group.rename(
            columns={branch_column: "Branch Code"}
        )
    else:
        sales_group["Branch Code"] = ""

    returns_group = returns.groupby("__po__", as_index=False).agg(
        ERP_CN_Qty=("__qty__", "sum"),
        ERP_CN_Amount=("__amount__", "sum"),
    ).rename(
        columns={
            "__po__": "PO Number",
            "ERP_CN_Qty": "ERP CN Qty",
            "ERP_CN_Amount": "ERP CN Amount",
        }
    )

    return sales_group.merge(returns_group, on="PO Number", how="outer")


def process_returns(dataframe: pd.DataFrame) -> pd.DataFrame:
    df = prepare_po_column(dataframe)

    quantity_column = find_column(df, ("Quantity",))
    created_date_column = find_column(
        df,
        ("Return Requested Date", "return_requested_date", "Return Created Date"),
        required=False,
    )
    approval_date_column = find_column(
        df,
        ("Return Approval Date", "return_approval_date"),
        required=False,
    )
    return_type_column = find_column(
        df,
        ("Return Type", "return_type"),
        required=False,
    )
    return_status_column = find_column(
        df,
        ("Return Status", "return_status"),
        required=False,
    )

    df["Flipkart Return Qty"] = to_number(df[quantity_column]).abs()

    aggregations = {"Flipkart Return Qty": "sum"}

    if created_date_column:
        df[created_date_column] = pd.to_datetime(
            df[created_date_column], errors="coerce"
        )
        aggregations[created_date_column] = "min"

    if approval_date_column:
        df[approval_date_column] = pd.to_datetime(
            df[approval_date_column], errors="coerce"
        )
        aggregations[approval_date_column] = "min"

    if return_type_column:
        aggregations[return_type_column] = lambda s: ", ".join(
            sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
        )

    if return_status_column:
        aggregations[return_status_column] = lambda s: ", ".join(
            sorted({str(v).strip() for v in s.dropna() if str(v).strip()})
        )

    output = (
        df.groupby("__po__", as_index=False)
        .agg(aggregations)
        .rename(columns={"__po__": "PO Number"})
    )

    if created_date_column:
        output = output.rename(
            columns={created_date_column: "Return Created Date"}
        )
    else:
        output["Return Created Date"] = pd.NaT

    if approval_date_column:
        output = output.rename(
            columns={approval_date_column: "Return Approval Date"}
        )
    else:
        output["Return Approval Date"] = pd.NaT

    if return_type_column:
        output = output.rename(columns={return_type_column: "Return Type"})
    else:
        output["Return Type"] = ""

    if return_status_column:
        output = output.rename(columns={return_status_column: "Return Status"})
    else:
        output["Return Status"] = ""

    return output

def process_payments(dataframe: pd.DataFrame) -> pd.DataFrame:
    df = prepare_po_column(dataframe)

    payment_date_column = find_column(
        df,
        ("Payment Date",),
        required=False,
    )
    settlement_column = find_column(
        df,
        (
            "Payment Received",
            "Bank Settlement Value (Rs.)",
            "Bank Settlement Value",
        ),
    )
    protection_column = find_column(
        df,
        ("Protection Fund",),
        required=False,
    )
    refund_column = find_column(
        df,
        ("Refund",),
        required=False,
    )

    sale_amount_column = find_column(
        df,
        ("Sale Amount",),
        required=False,
    )
    marketplace_fee_column = find_column(
        df,
        ("Marketplace Fee",),
        required=False,
    )
    taxes_column = find_column(
        df,
        ("Taxes",),
        required=False,
    )
    offer_adjustments_column = find_column(
        df,
        ("Offer Adjustments", "Offer Adjustment"),
        required=False,
    )
    gst_tcs_credit_column = find_column(
        df,
        ("Input GST TCS Credits",),
        required=False,
    )
    income_tax_credit_column = find_column(
        df,
        ("Income Tax Credits",),
        required=False,
    )

    numeric_columns: Dict[str, Optional[str]] = {
        "Bank Settlement Value": settlement_column,
        "Protection Fund": protection_column,
        "Refund": refund_column,
        "Payment Sale Amount": sale_amount_column,
        "Marketplace Fee": marketplace_fee_column,
        "Payment Taxes": taxes_column,
        "Offer Adjustments": offer_adjustments_column,
        "GST/TCS Credits": gst_tcs_credit_column,
        "Income Tax Credits": income_tax_credit_column,
    }

    for output_column, source_column in numeric_columns.items():
        df[output_column] = (
            to_number(df[source_column]) if source_column else 0.0
        )

    aggregations = {
        output_column: "sum" for output_column in numeric_columns
    }

    if payment_date_column:
        df[payment_date_column] = pd.to_datetime(
            df[payment_date_column], errors="coerce"
        )
        aggregations[payment_date_column] = "min"

        # Refund Date is the payment date attached to a non-zero refund row.
        df["__refund_date__"] = df[payment_date_column].where(
            df["Refund"].abs() > AMOUNT_TOLERANCE
        )
        aggregations["__refund_date__"] = "max"

    output = (
        df.groupby("__po__", as_index=False)
        .agg(aggregations)
        .rename(columns={"__po__": "PO Number"})
    )

    if payment_date_column:
        output = output.rename(
            columns={
                payment_date_column: "Payment Date",
                "__refund_date__": "Refund Date",
            }
        )
    else:
        output["Payment Date"] = pd.NaT
        output["Refund Date"] = pd.NaT

    # Expected settlement is estimated from the fields available in the
    # Flipkart payment sheet. Since Flipkart may store fee values as either
    # positive deductions or negative values, abs() is used for deductions.
    output["Expected Settlement"] = (
        output["Payment Sale Amount"]
        - output["Marketplace Fee"].abs()
        - output["Payment Taxes"].abs()
        + output["Offer Adjustments"]
        + output["GST/TCS Credits"]
        + output["Income Tax Credits"]
        + output["Protection Fund"]
        - output["Refund"].abs()
    )

    output["Settlement Difference"] = (
        output["Bank Settlement Value"]
        - output["Expected Settlement"]
    )

    return output


# ==========================================================
# RECONCILIATION RULES
# ==========================================================

def determine_remark(row: pd.Series) -> str:
    """
    Determine the reconciliation remark using the approved business priority.

    Priority:
    1. Billing & Payment Pending
       - Net Order Qty > 0
       - ERP Billed Qty = 0
       - Net ERP Qty = 0
       - Received Amount = 0
       - Order is not Returned/Cancelled

    2. Billing Pending
       - ERP Billed Qty = 0
       - Payment has been received
       - Order is not Returned/Cancelled

    3. Reconciled
       - Net ERP Qty = 0 after excluding pending billing/payment cases, or
       - Pay Rece % is at least 50%.

    4. CN Pending
       - Refund exists and Net ERP Qty > 0, or
       - Considered status is Returned/Cancelled and Net ERP Qty > 0
         with payment not received, or
       - Returned quantity exceeds absolute ERP CN quantity.

    5. Payment Pending
       - Delivered/non-returned/non-cancelled order
       - Net ERP Qty > 0
       - Payment not received

    6. Extra Billing
    7. Short Payment Received
    8. Order Details Review
    9. Reconciled fallback
    """

    def number(column: str) -> float:
        value = row.get(column, 0)
        if value is None or pd.isna(value) or value == "":
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    return_qty = number("Flipkart Return Qty")
    net_order_qty = number("Net Order Qty")
    order_qty = number("Order Qty")
    sales_report_qty = number("Sales Report Qty")

    billed_qty = number("ERP Billed Qty")
    cn_qty = number("ERP CN Qty")
    net_erp_qty = number("Net ERP Qty")

    received_amount = number("Received Amount")
    refund_amount = number("Refund")
    net_erp_billing = number("Net ERP Billing")

    considered_status = str(
        row.get("Consider the Order Status", "") or ""
    ).strip().lower()
    order_status = str(
        row.get("Order Status", "") or ""
    ).strip().lower()
    return_status = str(
        row.get("Return Status", "") or ""
    ).strip().lower()

    returned_statuses = {
        "returned",
        "returned & cancelled",
    }
    cancelled_statuses = {
        "cancelled",
        "returned & cancelled",
    }

    is_returned = considered_status in returned_statuses
    is_cancelled = considered_status in cancelled_statuses
    is_return_or_cancel = is_returned or is_cancelled
    is_delivered = considered_status == "delivered"

    # v15.11: Return is already physically received, but no billing/payment
    # was ever created. There is nothing to collect or reverse, so close it.
    return_created_date = pd.to_datetime(
        row.get("Return Created Date"), errors="coerce"
    )
    return_received_date = pd.to_datetime(
        row.get("Return Approval Date"), errors="coerce"
    )
    has_received_return = (
        return_qty > QUANTITY_TOLERANCE
        and (
            pd.notna(return_received_date)
            or pd.notna(return_created_date)
        )
    )
    if (
        has_received_return
        and billed_qty <= QUANTITY_TOLERANCE
        and abs(received_amount) <= AMOUNT_TOLERANCE
        and abs(refund_amount) <= AMOUNT_TOLERANCE
    ):
        return "Reconciled"

    # ------------------------------------------------------
    # 1. BILLING & PAYMENT PENDING
    # ------------------------------------------------------
    # Order quantity exists, but neither ERP billing nor payment exists.
    # This must be evaluated before the Net ERP Qty = 0 reconciliation rule.
    if (
        net_order_qty > QUANTITY_TOLERANCE
        and billed_qty <= QUANTITY_TOLERANCE
        and abs(net_erp_qty) <= QUANTITY_TOLERANCE
        and abs(received_amount) <= AMOUNT_TOLERANCE
        and abs(refund_amount) <= AMOUNT_TOLERANCE
        and not is_return_or_cancel
    ):
        return "Billing & Payment Pending"

    # ------------------------------------------------------
    # 2. BILLING PENDING
    # ------------------------------------------------------
    # Payment received, but ERP invoice has not been raised.
    if (
        received_amount > AMOUNT_TOLERANCE
        and billed_qty <= QUANTITY_TOLERANCE
        and not is_return_or_cancel
    ):
        return "Billing Pending"

    # ------------------------------------------------------
    # 3. RECONCILED
    # ------------------------------------------------------
    # Once billing/payment-pending scenarios are excluded, no open ERP
    # quantity means the order is reconciled.
    if abs(net_erp_qty) <= QUANTITY_TOLERANCE:
        return "Reconciled"

    # Payment receipt is at least 50% of Net ERP Billing.
    if net_erp_billing > AMOUNT_TOLERANCE:
        payment_receipt_ratio = received_amount / net_erp_billing
        if payment_receipt_ratio >= MIN_PAYMENT_RECEIPT_RATIO:
            return "Reconciled"

    # ------------------------------------------------------
    # 4. CN PENDING
    # ------------------------------------------------------
    # A refund has been processed while ERP quantity/CN is still open.
    if (
        net_erp_qty > QUANTITY_TOLERANCE
        and refund_amount < -AMOUNT_TOLERANCE
    ):
        return "CN Pending"

    # Returned or Cancelled order with open ERP quantity and no payment.
    # No age condition is required.
    if (
        is_return_or_cancel
        and net_erp_qty > QUANTITY_TOLERANCE
        and received_amount <= AMOUNT_TOLERANCE
    ):
        return "CN Pending"

    # Returned quantity is greater than ERP CN quantity raised.
    if (
        is_returned
        and return_qty > abs(cn_qty) + QUANTITY_TOLERANCE
    ):
        return "CN Pending"

    # ------------------------------------------------------
    # 5. PAYMENT PENDING
    # ------------------------------------------------------
    # Only a Delivered/non-returned/non-cancelled order can be Payment Pending.
    if (
        not is_return_or_cancel
        and net_erp_qty > QUANTITY_TOLERANCE
        and received_amount <= AMOUNT_TOLERANCE
    ):
        return "Payment Pending"

    # ------------------------------------------------------
    # 6. EXTRA BILLING
    # ------------------------------------------------------
    cancelled_return_delivered = (
        is_delivered and "cancelled" in return_status
    )

    if (
        net_order_qty > QUANTITY_TOLERANCE
        and net_erp_qty > net_order_qty + QUANTITY_TOLERANCE
        and not cancelled_return_delivered
    ):
        return "Extra Billing"

    # ------------------------------------------------------
    # 7. SHORT PAYMENT RECEIVED
    # ------------------------------------------------------
    if (
        net_erp_billing > AMOUNT_TOLERANCE
        and received_amount > AMOUNT_TOLERANCE
    ):
        return "Short Payment Received"

    # ------------------------------------------------------
    # 8. ORDER DETAILS REVIEW
    # ------------------------------------------------------
    # Check only columns that are actually populated in the combined output.
    if (
        order_qty <= QUANTITY_TOLERANCE
        and sales_report_qty <= QUANTITY_TOLERANCE
        and net_order_qty <= QUANTITY_TOLERANCE
    ):
        return "Order Details Review"

    # ------------------------------------------------------
    # 9. FALLBACK
    # ------------------------------------------------------
    return "Reconciled"


def generate_reconciliation(input_file: str | Path, output_file: str | Path) -> Path:
    input_path = Path(input_file)
    output_path = Path(output_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Open only long enough to inspect sheet names, then release the
    # Windows file handle before processing and temporary-folder cleanup.
    with pd.ExcelFile(input_path) as excel_file:
        sheet_names = excel_file.sheet_names

    sheet_aliases = {
        "ERP Sales Register": ("ERP Sales Register", "ERP"),
        "Flipkart Payments": ("Flipkart Payments", "Payments"),
        "Flipkart Returns": ("Flipkart Returns", "Returns"),
        "Flipkart All Orders": ("Flipkart All Orders", "All Orders"),
        "Flipkart Sales Report": ("Flipkart Sales Report", "Sales Report"),
    }

    resolved_sheets: Dict[str, str] = {}

    for logical_name, aliases in sheet_aliases.items():
        actual_name = find_sheet_name(sheet_names, aliases)

        if not actual_name:
            raise ValueError(
                f"Required sheet '{logical_name}' was not found. "
                f"Available sheets: {sheet_names}"
            )

        resolved_sheets[logical_name] = actual_name

    erp_df = pd.read_excel(
        input_path,
        sheet_name=resolved_sheets["ERP Sales Register"],
    )
    payment_df = pd.read_excel(
        input_path,
        sheet_name=resolved_sheets["Flipkart Payments"],
    )
    return_df = pd.read_excel(
        input_path,
        sheet_name=resolved_sheets["Flipkart Returns"],
    )
    orders_df = pd.read_excel(
        input_path,
        sheet_name=resolved_sheets["Flipkart All Orders"],
    )
    sales_df = pd.read_excel(
        input_path,
        sheet_name=resolved_sheets["Flipkart Sales Report"],
    )

    orders_output = process_all_orders(orders_df)
    sales_output = process_sales_report(sales_df)
    erp_output = process_erp(erp_df)
    returns_output = process_returns(return_df)
    payments_output = process_payments(payment_df)

    # Master list is the union of unique PO numbers from both order sheets.
    master_po = pd.concat(
        [
            orders_output[["PO Number"]],
            sales_output[["PO Number"]],
        ],
        ignore_index=True,
    ).drop_duplicates()

    result = merge_frames(
        master_po,
        [
            orders_output,
            sales_output,
            erp_output,
            returns_output,
            payments_output,
        ],
    )

    numeric_output_columns = [
        "Order Qty",
        "Sales Report Qty",
        "ERP Billed Qty",
        "ERP Billing Amount",
        "Flipkart Return Qty",
        "ERP CN Qty",
        "ERP CN Amount",
        "Bank Settlement Value",
        "Protection Fund",
        "Refund",
        "Payment Sale Amount",
        "Marketplace Fee",
        "Payment Taxes",
        "Offer Adjustments",
        "GST/TCS Credits",
        "Income Tax Credits",
        "Expected Settlement",
        "Settlement Difference",
    ]

    for column in numeric_output_columns:
        if column not in result.columns:
            result[column] = 0.0
        result[column] = to_number(result[column])

    text_output_columns = [
        "Order Status",
        "Sales Event Sub Type",
        "Invoice No",
        "Branch Code",
        "Return Type",
        "Return Status",
    ]

    for column in text_output_columns:
        if column not in result.columns:
            result[column] = ""
        result[column] = result[column].fillna("")

    # Attached output format: use All Orders quantity first and Sales
    # Report quantity as fallback when the PO is missing from All Orders.
    result["Order Date or Sales Report Order Date"] = (
        pd.to_datetime(result["Order Date"], errors="coerce")
        .combine_first(
            pd.to_datetime(
                result["Sales Report Order Date"], errors="coerce"
            )
        )
    )

    result["Order Qty or Sales"] = result["Order Qty"].where(
        result["Order Qty"].abs() > QUANTITY_TOLERANCE,
        result["Sales Report Qty"],
    )

    result["Net Order Qty"] = (
        result["Order Qty or Sales"]
        - result["Flipkart Return Qty"]
    )

    # ERP CN quantities and amounts are normally negative in ERP. If the
    # source provides positive CN values, convert them to negative so the
    # net formulas remain K + L and Billing + CN.
    result["ERP CN Qty"] = -result["ERP CN Qty"].abs()
    result["ERP CN Amount"] = -result["ERP CN Amount"].abs()

    result["Net ERP Qty"] = (
        result["ERP Billed Qty"]
        + result["ERP CN Qty"]
    )

    result["Net ERP Billing"] = (
        result["ERP Billing Amount"]
        + result["ERP CN Amount"]
    )

    result["Received Amount"] = result["Bank Settlement Value"]

    # Payment receipt percentage used for reconciliation.
    # Stored as a decimal so Excel can display it as a percentage.
    result["Pay Rece%"] = 0.0
    positive_billing_mask = result["Net ERP Billing"].abs() > AMOUNT_TOLERANCE
    result.loc[positive_billing_mask, "Pay Rece%"] = (
        result.loc[positive_billing_mask, "Received Amount"]
        / result.loc[positive_billing_mask, "Net ERP Billing"]
    )

    result["Qty Difference"] = (
        result["Net Order Qty"]
        - result["Net ERP Qty"]
    )

    result["Price Difference"] = (
        result["Net ERP Billing"]
        - result["Received Amount"]
    )

    # Pending Since Days:
    # Default priority:
    #   Refund Date -> Return Created Date -> Return Approval Date
    #
    # For Billing & Payment Pending, Billing Pending, and Payment Pending:
    #   Refund Date -> Return Created Date -> Order Date
    #
    # Therefore, when neither Refund Date nor Return Created Date is
    # available for these three remarks, the Order Date is used.
    today = pd.Timestamp.now().normalize()

    refund_date = pd.to_datetime(
        result.get("Refund Date"), errors="coerce"
    )
    return_created_date = pd.to_datetime(
        result.get("Return Created Date"), errors="coerce"
    )
    return_approval_date = pd.to_datetime(
        result.get("Return Approval Date"), errors="coerce"
    )
    order_date = pd.to_datetime(
        result.get("Order Date or Sales Report Order Date"),
        errors="coerce",
    )

    default_pending_base_date = (
        refund_date
        .combine_first(return_created_date)
        .combine_first(return_approval_date)
    )

    billing_payment_pending_base_date = (
        refund_date
        .combine_first(return_created_date)
        .combine_first(order_date)
    )

    result["Consider the Order Status"] = result.apply(
        get_considered_order_status,
        axis=1,
    )

    result["Remarks"] = result.apply(determine_remark, axis=1)

    order_date_fallback_remarks = {
        "Billing & Payment Pending",
        "Billing Pending",
        "Payment Pending",
    }

    use_order_date_fallback = result["Remarks"].isin(
        order_date_fallback_remarks
    )

    pending_base_date = default_pending_base_date.copy()
    pending_base_date.loc[use_order_date_fallback] = (
        billing_payment_pending_base_date.loc[use_order_date_fallback]
    )

    pending_mask = result["Remarks"].ne("Reconciled")
    result["Pending Since Days"] = pd.NA

    valid_pending_date_mask = pending_mask & pending_base_date.notna()
    result.loc[
        valid_pending_date_mask,
        "Pending Since Days",
    ] = (
        today
        - pending_base_date.loc[
            valid_pending_date_mask
        ].dt.normalize()
    ).dt.days.clip(lower=0).astype("Int64")

    remark_rank = {
        remark: rank for rank, remark in enumerate(REMARK_PRIORITY, start=1)
    }
    result["__remark_rank__"] = result["Remarks"].map(remark_rank).fillna(999)
    result = result.sort_values(
        ["__remark_rank__", "PO Number"],
        ascending=[True, True],
    ).drop(columns="__remark_rank__")

    preferred_columns = [
        "PO Number",
        "Order Date or Sales Report Order Date",
        "Order Status",
        "Sales Event Sub Type",
        "Order Qty or Sales",
        "Flipkart Return Qty",
        "Return Approval Date",
        "Return Type",
        "Return Status",
        "Consider the Order Status",
        "Net Order Qty",
        "ERP Billed Qty",
        "ERP CN Qty",
        "Net ERP Qty",
        "Net ERP Billing",
        "Received Amount",
        "Pay Rece%",
        "Qty Difference",
        "Price Difference",
        "Protection Fund",
        "Refund",
        "Refund Date",
        "Remarks",
        "Pending Since Days",
    ]

    for column in preferred_columns:
        if column not in result.columns:
            result[column] = ""

    result = result[preferred_columns]

    summary = (
        result.groupby("Remarks", dropna=False)
        .size()
        .reset_index(name="PO Count")
    )

    summary["Priority"] = (
        summary["Remarks"].map(remark_rank).fillna(999)
    )
    summary = summary.sort_values("Priority").drop(columns="Priority")

    logic_rows = [
        ["Rule", "Definition"],
        [
            "Master PO",
            "Union of unique PO Numbers from Flipkart All Orders and Flipkart Sales Report",
        ],
        [
            "Consider the Order Status",
            "Derived from the supplied mapping using Order Status + Sales Event Sub Type + Return Type + Return Status",
        ],
        [
            "Extra Billing",
            "ERP billed quantity is greater than the maximum quantity found in All Orders or Sales Report",
        ],
        [
            "CN Pending",
            "Flipkart return quantity is greater than ERP credit-note quantity",
        ],
        [
            "Billing & Payment Pending",
            "ERP billed quantity is zero and bank settlement is zero",
        ],
        [
            "Billing Pending",
            "ERP billed quantity is zero but payment exists",
        ],
        [
            "Payment Pending",
            "ERP billing exists but bank settlement is zero",
        ],
        [
            "Short Payment Received",
            "Bank settlement is lower than estimated expected settlement beyond the configured tolerance",
        ],
        [
            "Order Details Review",
            "PO exists in only one of the two order-detail sheets",
        ],
        [
            "Reconciled",
            "No billing, CN, payment, or source-order exception is detected",
        ],
        [
            "Zero Billing/Payment Status Exception",
            "Configured cancelled, returned, rejected, return-requested, approval-hold and related status combinations are Reconciled when ERP billed quantity and received amount are both zero",
        ],
        [
            "Pending Since Days",
            "Current date minus Refund Date; if missing, Return Created Date; if still missing, Return Approval Date",
        ],
        [
            "Amount tolerance",
            str(AMOUNT_TOLERANCE),
        ],
        [
            "Quantity tolerance",
            str(QUANTITY_TOLERANCE),
        ],
    ]

    logic_dataframe = pd.DataFrame(logic_rows[1:], columns=logic_rows[0])

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        result.to_excel(
            writer,
            sheet_name="Reconciliation",
            index=False,
        )
        summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        logic_dataframe.to_excel(
            writer,
            sheet_name="Logic & Remarks",
            index=False,
        )
        status_mapping_dataframe = pd.DataFrame(
            [
                {
                    "Order Status": ", ".join(key[0]),
                    "Sales Event Sub Type": ", ".join(key[1]),
                    "Return Type": ", ".join(key[2]),
                    "Return Status": ", ".join(key[3]),
                    "Consider the Order Status": value,
                }
                for key, value in STATUS_CONSIDERATION_MAP.items()
            ]
        )
        status_mapping_dataframe.to_excel(
            writer,
            sheet_name="Status Mapping",
            index=False,
        )

    format_output_workbook(output_path)

    LOGGER.info("Reconciliation completed: %s", output_path)
    return output_path


# ==========================================================
# EXCEL FORMATTING
# ==========================================================

def format_output_workbook(output_file: str | Path) -> None:
    workbook = load_workbook(output_file)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            column_index = column_cells[0].column
            max_length = 0

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(max_length + 2, 12), 40)

    reconciliation_sheet = workbook["Reconciliation"]

    if "Remarks" in [
        cell.value for cell in reconciliation_sheet[1]
    ]:
        remarks_column = next(
            cell.column
            for cell in reconciliation_sheet[1]
            if cell.value == "Remarks"
        )

        fill_map = {
            "Reconciled": "C6E0B4",
            "Billing Pending": "FFF2CC",
            "Payment Pending": "FCE4D6",
            "Billing & Payment Pending": "F8CBAD",
            "CN Pending": "FFD966",
            "Short Payment Received": "F4B084",
            "Extra Billing": "FF9999",
            "Order Details Review": "D9EAF7",
        }

        for row_number in range(2, reconciliation_sheet.max_row + 1):
            cell = reconciliation_sheet.cell(
                row=row_number,
                column=remarks_column,
            )

            if cell.value in fill_map:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_map[cell.value],
                )
                cell.font = Font(bold=True)

    workbook.save(output_file)
    workbook.close()


# ==========================================================
# TELEGRAM BOT HANDLERS
# ==========================================================

async def start_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    message = (
        "Welcome to the Flipkart Reconciliation Bot.\n\n"
        "Upload one Excel workbook containing these sheets:\n"
        "• ERP Sales Register\n"
        "• Flipkart Payments\n"
        "• Flipkart Returns\n"
        "• Flipkart All Orders\n"
        "• Flipkart Sales Report\n\n"
        "The bot will return a PO-wise reconciliation file with remarks."
    )

    if update.message:
        await update.message.reply_text(message)


async def help_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    message = (
        "Send an .xlsx or .xlsm Flipkart workbook.\n\n"
        "The master PO list is created from the union of unique PO Numbers "
        "in All Orders and Sales Report."
    )

    if update.message:
        await update.message.reply_text(message)


async def handle_document(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    if not update.message or not update.message.document:
        return

    document = update.message.document
    file_name = document.file_name or "flipkart_input.xlsx"
    extension = Path(file_name).suffix.lower()

    if extension not in {".xlsx", ".xlsm"}:
        await update.message.reply_text(
            "Please upload an Excel file in .xlsx or .xlsm format."
        )
        return

    await update.message.reply_text(
        "File received. Reconciliation is being processed."
    )

    try:
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            input_path = temporary_path / file_name
            output_path = temporary_path / OUTPUT_FILE_NAME

            telegram_file = await document.get_file()
            await telegram_file.download_to_drive(
                custom_path=str(input_path)
            )

            generate_reconciliation(
                input_file=input_path,
                output_file=output_path,
            )

            with output_path.open("rb") as output_document:
                await update.message.reply_document(
                    document=output_document,
                    filename=OUTPUT_FILE_NAME,
                    caption=(
                        "Flipkart reconciliation completed successfully."
                    ),
                )

    except Exception as error:
        LOGGER.exception("Telegram reconciliation failed")

        await update.message.reply_text(
            "Reconciliation failed.\n\n"
            f"Error: {error}"
        )


async def error_handler(
    update: object,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    LOGGER.exception(
        "Unhandled Telegram error",
        exc_info=context.error,
    )


def run_telegram_bot() -> None:
    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        raise ValueError(
            "Update BOT_TOKEN near the top of the Python file before "
            "starting the Telegram bot."
        )

    application = (
        Application.builder()
        .token(BOT_TOKEN)
        .build()
    )

    application.add_handler(
        CommandHandler("start", start_command)
    )
    application.add_handler(
        CommandHandler("help", help_command)
    )
    application.add_handler(
        MessageHandler(
            filters.Document.ALL,
            handle_document,
        )
    )
    application.add_error_handler(error_handler)

    LOGGER.info("Telegram bot started")
    application.run_polling(
        allowed_updates=Update.ALL_TYPES
    )


# ==========================================================
# COMMAND-LINE SUPPORT
# ==========================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Flipkart PO-wise reconciliation"
    )

    parser.add_argument(
        "input_file",
        nargs="?",
        help="Path to Flipkart Excel workbook",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=OUTPUT_FILE_NAME,
        help=f"Output workbook path. Default: {OUTPUT_FILE_NAME}",
    )
    parser.add_argument(
        "--telegram",
        action="store_true",
        help="Start Telegram bot mode",
    )

    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()

    if arguments.telegram:
        run_telegram_bot()
        return

    if arguments.input_file:
        generated_file = generate_reconciliation(
            input_file=arguments.input_file,
            output_file=arguments.output,
        )
        print(f"Created: {generated_file}")
        return

    # Default mode when no workbook path is supplied
    run_telegram_bot()


if __name__ == "__main__":
    main()
