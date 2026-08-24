
from __future__ import annotations

import io
import json
import re
import os
import hashlib
import sqlite3
import psycopg2
from psycopg2.extras import execute_batch
import shutil
import tempfile
import warnings
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
from datetime import date, datetime
from time import perf_counter, sleep
from pathlib import Path

import pandas as pd
import requests
import msal
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Protection, PatternFill, Font

import amazon_engine
import flipkart_engine

warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable*",
    category=UserWarning,
)

APP_VERSION = "15.18-management-mis"

st.set_page_config(
    page_title="E-Commerce Reconciliation Control Tower",
    page_icon="📊",
    layout="wide",
)

# ============================================================
# PERSISTENT STORAGE
# ============================================================
# Production / Streamlit Cloud:
#   DATABASE_URL in Streamlit Secrets -> Supabase PostgreSQL.
# Local Windows fallback:
#   SQLite under ~/ReconciliationTowerData.
#
# Reconciliation, pending-task, MIR/TEI, ticket, completion, upload-history
# and audit data are stored in PostgreSQL when DATABASE_URL is configured.
DATA_DIR = Path.home() / "ReconciliationTowerData"
DATA_DIR.mkdir(parents=True, exist_ok=True)

SOURCE_DIR = DATA_DIR / "source_snapshots"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / "reconciliation_tower.db"
M365_CONFIG = DATA_DIR / "m365_config.json"
M365_TOKEN_CACHE = DATA_DIR / "m365_token_cache.json"
SECURITY_CONFIG = DATA_DIR / "security.json"

ATTACHED_COLUMNS_INTERNAL = [
    "Order No",
    "Order Date/Shipment Date",
    "Order Item",
    "Order Qty",
    "Order Price",
    "Branch Code",
    "Order Status",
    "Order Return Date",
    "Return Completed Date",
    "Return Status",
    "Payment Status",
    "Refund",
    "Reimbursement",
    "Invoice",
    "Invoice Date",
    "Invoice Qty",
    "Invoice Price",
    "Credit Note",
    "CN Date",
    "CN Qty",
    "CN Price",
    "Pending Remarks",
    "Pending Task Created Date",
    "MIR No",
    "MIR Date",
    "TEI No",
    "TEI Date",
    "TEI Qty",
    "Ticket Raised If Any",
    "Ticket Raised Date",
    "Team Remarks",
    "Task Completed Date",
    "Task Status",
    "Aging Days of Task Pending/Completed",
]

# Excel export preserves the user's attached duplicate Date / Qty / Price headers.
ATTACHED_COLUMNS_EXPORT = [
    "Order No",
    "Order Date/Shipment Date",
    "Order Item",
    "Order Qty",
    "Order Price",
    "Branch Code",
    "Order Status",
    "Order Return Date",
    "Return Completed Date",
    "Return Status",
    "Payment Status",
    "Refund",
    "Reimbursement",
    "Invoice",
    "Date",
    "Qty",
    "Price",
    "Credit Note",
    "Date",
    "Qty",
    "Price",
    "Pending Remarks",
    "Pending Task Created Date",
    "MIR No",
    "MIR Date",
    "TEI No",
    "TEI Date",
    "TEI Qty",
    "Ticket Raised If Any",
    "Ticket Raised Date",
    "Team Remarks",
    "Task Completed Date",
    "Task Status",
    "Aging Days of Task Pending/Completed",
]

MIR_UPLOAD_COLUMNS = [
    "Order No","Invoice No","Inv Date","Inv Qty","Price","Product Code",
    "MIR No","MIR Date","TEI No","TEI Date","SR No","SR Date","Remarks"
]

MIR_BRANCH_PEOPLE = {
    "1600": "Rajan & Sashikant",
    "5400": "Karthik & Sonia",
    "5600": "Jitendra & Sonia",
    "7800": "Sachin & Sonia",
}

TASK_KEYWORDS = [
    "CN & Payment Pending",
    "Billing & Payment Pending",
    "Short Payment Received",
    "Short Payment Received",
    "Billing Pending",
    "Payment Pending",
    "CN Pending",
    "MIR Pending",
    "TEI Pending",
    "Extra Billing",
    "Order Details Review",
    "Billing Entry Review",
]


def load_security():
    # Cloud-first: keep Admin PIN across Streamlit restart/redeploy.
    try:
        if "db" in globals():
            with db() as c:
                row = c.execute(
                    "SELECT setting_value FROM app_settings WHERE setting_key=?",
                    ("admin_pin_hash",)
                ).fetchone()
            if row and txt(row[0]):
                return {"admin_pin_hash": txt(row[0])}
    except Exception:
        pass

    # Local fallback for offline Windows use.
    if not SECURITY_CONFIG.exists():
        return {}
    try:
        return json.loads(SECURITY_CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}

def pin_hash(pin):
    return hashlib.sha256(str(pin).encode("utf-8")).hexdigest()

def save_admin_pin(pin):
    pin = str(pin).strip()
    if len(pin) < 4:
        raise ValueError("Admin PIN must contain at least 4 characters.")

    hashed = pin_hash(pin)

    # Persist to Supabase/PostgreSQL (or SQLite fallback DB).
    with db() as c:
        c.execute("""
            INSERT INTO app_settings(setting_key,setting_value,updated_at)
            VALUES(?,?,?)
            ON CONFLICT(setting_key) DO UPDATE SET
                setting_value=excluded.setting_value,
                updated_at=excluded.updated_at
        """,(
            "admin_pin_hash",
            hashed,
            datetime.now().isoformat(timespec="seconds")
        ))
        c.commit()

    # Keep a local fallback copy for offline Windows use.
    try:
        SECURITY_CONFIG.write_text(
            json.dumps({"admin_pin_hash": hashed}, indent=2),
            encoding="utf-8"
        )
    except Exception:
        pass

def is_admin_unlocked():
    return bool(st.session_state.get("admin_unlocked", False))

# ============================================================
# DATABASE
# ============================================================
def get_database_url():
    """Return Supabase/PostgreSQL DATABASE_URL when configured."""
    try:
        value = st.secrets.get("DATABASE_URL", "")
        if value:
            return str(value).strip()
    except Exception:
        pass
    return os.getenv("DATABASE_URL", "").strip()


def using_postgres():
    return bool(get_database_url())


def invalidate_read_cache():
    """Refresh cached dashboard reads after any database write."""
    try:
        st.cache_data.clear()
    except Exception:
        pass


class _PgResult:
    def __init__(self, cursor=None, rows=None):
        self.cursor = cursor
        self.rows = rows

    def fetchall(self):
        if self.rows is not None:
            return self.rows
        return self.cursor.fetchall()

    def fetchone(self):
        if self.rows is not None:
            return self.rows[0] if self.rows else None
        return self.cursor.fetchone()


class _PgCursorCompat:
    """DB-API cursor wrapper that translates SQLite qmark placeholders."""
    def __init__(self, conn):
        self._cursor = conn.cursor()

    @property
    def description(self):
        return self._cursor.description

    @property
    def rowcount(self):
        return self._cursor.rowcount

    def execute(self, sql, params=None):
        text = str(sql)

        # pandas.read_sql_query() may send SQLite-style '?' placeholders.
        text = text.replace("?", "%s")

        self._cursor.execute(text, params or ())
        return self

    def executemany(self, sql, seq_of_params):
        text = str(sql).replace("?", "%s")
        self._cursor.executemany(text, seq_of_params)
        return self

    def fetchall(self):
        return self._cursor.fetchall()

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchmany(self, size=None):
        if size is None:
            return self._cursor.fetchmany()
        return self._cursor.fetchmany(size)

    def close(self):
        self._cursor.close()


def _postgres_url_candidate(url, port_override=None):
    """Return a hardened PostgreSQL URI without ever logging credentials."""
    try:
        parts = urlsplit(str(url).strip())
        host = parts.hostname or ""
        port = int(port_override or parts.port or 5432)
        user = parts.username or ""
        password = parts.password or ""

        # Preserve percent-escaped credentials from the original URI when possible.
        netloc = ""
        if user:
            from urllib.parse import quote
            netloc = quote(user, safe=".")
            if password:
                netloc += ":" + quote(password, safe="")
            netloc += "@"
        netloc += host
        if port:
            netloc += f":{port}"

        query = dict(parse_qsl(parts.query, keep_blank_values=True))
        query.setdefault("sslmode", "require")
        query.setdefault("gssencmode", "disable")
        query.setdefault("client_encoding", "UTF8")
        return urlunsplit((parts.scheme or "postgresql", netloc, parts.path or "/postgres", urlencode(query), ""))
    except Exception:
        # If parsing ever fails, keep the user's supplied URI unchanged.
        return str(url).strip()


def _postgres_connection_candidates(url):
    """Session pooler first; transaction pooler is a controlled fallback."""
    candidates = [("configured", _postgres_url_candidate(url))]
    try:
        parts = urlsplit(str(url).strip())
        host = (parts.hostname or "").lower()
        port = parts.port or 5432
        if host.endswith(".pooler.supabase.com") and int(port) == 5432:
            candidates.append(("transaction-pooler-fallback", _postgres_url_candidate(url, 6543)))
    except Exception:
        pass
    return candidates


class PostgresCompat:
    """
    DB-API compatibility layer for PostgreSQL/Supabase.

    The configured Session Pooler is tried first. If Supabase's Session Pooler
    is temporarily unable to route to the project database, the same Supavisor
    host on transaction-pooler port 6543 is tried as a controlled fallback.
    No automatic SQLite fallback is used while DATABASE_URL is configured, so
    production data can never silently split across two databases.
    """
    def __init__(self, url):
        last_error = None
        self.connection_mode = ""
        for mode, candidate in _postgres_connection_candidates(url):
            for attempt in range(2):
                try:
                    self.conn = psycopg2.connect(
                        candidate,
                        connect_timeout=20,
                        application_name="glen_reconciliation_tower",
                        keepalives=1,
                        keepalives_idle=30,
                        keepalives_interval=10,
                        keepalives_count=3,
                    )
                    self.connection_mode = mode
                    return
                except psycopg2.OperationalError as exc:
                    last_error = exc
                    if attempt == 0:
                        sleep(1.5)
        raise last_error

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if exc_type is None:
                self.conn.commit()
            else:
                self.conn.rollback()
        finally:
            self.conn.close()
        return False

    def cursor(self):
        # Required by pandas.read_sql_query/read_sql.
        return _PgCursorCompat(self.conn)

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()

    @staticmethod
    def _convert_placeholders(sql):
        # Existing application SQL uses qmark placeholders.
        # None of the business queries contain literal '?' characters.
        return sql.replace("?", "%s")

    @staticmethod
    def _convert_ddl(sql):
        # PostgreSQL equivalent of SQLite AUTOINCREMENT.
        sql = re.sub(
            r"\bid\s+INTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT\b",
            "id BIGSERIAL PRIMARY KEY",
            sql,
            flags=re.I,
        )
        return sql

    def execute(self, sql, params=None):
        text = str(sql).strip()

        # Emulate SQLite PRAGMA table_info(table) used by migrations.
        pragma = re.fullmatch(
            r"PRAGMA\s+table_info\(([^)]+)\)",
            text,
            flags=re.I,
        )
        if pragma:
            table = pragma.group(1).strip().strip('"').strip("'")
            cur = self.conn.cursor()
            cur.execute(
                """
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name=%s
                ORDER BY ordinal_position
                """,
                (table,),
            )
            cols = cur.fetchall()
            cur.close()
            # Match SQLite PRAGMA tuple shape enough for existing row[1] usage:
            # (cid, name, type, notnull, default, pk)
            rows = [
                (
                    idx,
                    row[0],
                    row[1],
                    1 if row[2] == "NO" else 0,
                    row[3],
                    0,
                )
                for idx, row in enumerate(cols)
            ]
            return _PgResult(rows=rows)

        text = self._convert_ddl(text)
        text = self._convert_placeholders(text)

        cur = self.conn.cursor()
        cur.execute(text, params or ())
        return _PgResult(cursor=cur)

    def executemany(self, sql, seq_of_params):
        text = self._convert_ddl(str(sql))
        text = self._convert_placeholders(text)
        cur = self.conn.cursor()
        rows = list(seq_of_params)
        if rows:
            # psycopg2 cursor.executemany() still performs many server round trips.
            # execute_batch groups statements into far fewer network calls, which is
            # substantially faster for Supabase/remote PostgreSQL.
            execute_batch(cur, text, rows, page_size=1000)
        return _PgResult(cursor=cur)


def db():
    """
    Cloud-first database connection.
    DATABASE_URL configured -> Supabase/PostgreSQL.
    Otherwise local SQLite is used for Windows/offline development.
    """
    url = get_database_url()
    if url:
        return PostgresCompat(url)
    return sqlite3.connect(DB_PATH, check_same_thread=False)


@st.cache_resource(show_spinner=False)
def init_db():
    with db() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS reconciliation_master(
            portal TEXT NOT NULL,
            order_no TEXT NOT NULL,
            order_date TEXT,
            order_item TEXT,
            order_qty REAL,
            repl_quantity REAL,
            order_price REAL,
            replacement_item_price REAL,
            courier_customer_return_qty REAL,
            branch_code TEXT,
            order_status TEXT,
            order_return_date TEXT,
            return_completed_date TEXT,
            return_status TEXT,
            payment_status TEXT,
            refund REAL,
            reimbursement REAL,
            adjustment REAL,
            invoice_no TEXT,
            invoice_date TEXT,
            invoice_qty REAL,
            invoice_price REAL,
            cn_no TEXT,
            cn_date TEXT,
            cn_qty REAL,
            cn_price REAL,
            pending_remarks TEXT,
            replacement_order_id TEXT,
            net_pay_received REAL,
            deferred_amount REAL,
            total_deductions REAL,
            transaction_status TEXT,
            payment_date TEXT,
            refund_date TEXT,
            payment_reference TEXT,
            reimbursement_reason TEXT,
            source_file TEXT,
            source_uploaded_at TEXT,
            PRIMARY KEY(portal, order_no)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS pending_tasks(
            portal TEXT NOT NULL,
            order_no TEXT NOT NULL,
            task_type TEXT NOT NULL,
            branch_code TEXT,
            task_created_date TEXT,
            working_date TEXT,
            task_completed_date TEXT,
            task_status TEXT,
            team_remarks TEXT,
            ticket_raised TEXT,
            raised_date TEXT,
            last_update TEXT,
            PRIMARY KEY(portal, order_no, task_type)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS mir_details(
            branch_code TEXT NOT NULL,
            order_no TEXT NOT NULL,
            invoice_no TEXT,
            inv_date TEXT,
            inv_qty REAL,
            price REAL,
            product_code TEXT,
            mir_no TEXT,
            mir_date TEXT,
            tei_no TEXT,
            tei_date TEXT,
            tei_qty REAL,
            sr_no TEXT,
            sr_date TEXT,
            remarks TEXT,
            responsible_persons TEXT,
            updated_at TEXT,
            PRIMARY KEY(branch_code, order_no)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS branch_owner_rules(
            branch_code TEXT NOT NULL,
            task_type TEXT NOT NULL,
            owner_name TEXT,
            owner_email TEXT NOT NULL,
            PRIMARY KEY(branch_code, task_type, owner_email)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS upload_history(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portal TEXT,
            filename TEXT,
            uploaded_at TEXT,
            rows_processed INTEGER
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS source_kpis(
            portal TEXT PRIMARY KEY,
            adjustment_total REAL DEFAULT 0,
            updated_at TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS source_workbooks(
            portal TEXT PRIMARY KEY,
            filename TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            workbook_data BYTEA NOT NULL,
            workbook_sha256 TEXT,
            workbook_size INTEGER
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS app_settings(
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT,
            updated_at TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS activity_audit(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            activity_time TEXT NOT NULL,
            action_type TEXT NOT NULL,
            portal TEXT,
            branch_code TEXT,
            order_no TEXT,
            task_type TEXT,
            filename TEXT,
            details TEXT
        )
        """)
        # Schema migration for v10 settlement fields.
        existing_cols = {
            row[1] for row in c.execute(
                "PRAGMA table_info(reconciliation_master)"
            ).fetchall()
        }
        migrations = {
            "replacement_order_id": "TEXT",
            "repl_quantity": "REAL",
            "replacement_item_price": "REAL",
            "courier_customer_return_qty": "REAL",
            "net_pay_received": "REAL",
            "deferred_amount": "REAL",
            "adjustment": "REAL",
            "total_deductions": "REAL",
            "transaction_status": "TEXT",
            "payment_date": "TEXT",
            "refund_date": "TEXT",
            "payment_reference": "TEXT",
            "reimbursement_reason": "TEXT",
        }
        for col, sql_type in migrations.items():
            if col not in existing_cols:
                c.execute(
                    f"ALTER TABLE reconciliation_master ADD COLUMN {col} {sql_type}"
                )

        # Task-table migration for Ticket Raised If Any.
        task_existing_cols = {
            row[1] for row in c.execute(
                "PRAGMA table_info(pending_tasks)"
            ).fetchall()
        }
        if "ticket_raised" not in task_existing_cols:
            c.execute(
                "ALTER TABLE pending_tasks ADD COLUMN ticket_raised TEXT"
            )
        if "raised_date" not in task_existing_cols:
            c.execute(
                "ALTER TABLE pending_tasks ADD COLUMN raised_date TEXT"
            )

        # v15.7 MIR/TEI quantity migration.
        mir_existing_cols = {
            row[1] for row in c.execute(
                "PRAGMA table_info(mir_details)"
            ).fetchall()
        }
        if "tei_qty" not in mir_existing_cols:
            c.execute(
                "ALTER TABLE mir_details ADD COLUMN tei_qty REAL"
            )

        # Source snapshot diagnostics migration (safe for existing databases).
        source_existing_cols = {
            row[1] for row in c.execute(
                "PRAGMA table_info(source_workbooks)"
            ).fetchall()
        }
        if "workbook_sha256" not in source_existing_cols:
            c.execute(
                "ALTER TABLE source_workbooks ADD COLUMN workbook_sha256 TEXT"
            )
        if "workbook_size" not in source_existing_cols:
            c.execute(
                "ALTER TABLE source_workbooks ADD COLUMN workbook_size INTEGER"
            )

        # v15.8 query-performance indexes. These are safe/idempotent and speed
        # branch dashboards, pending-task lookups, team uploads and audit/history
        # reads as the persistent database grows.
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_reco_branch_order
            ON reconciliation_master(branch_code, order_no)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_reco_portal_uploaded
            ON reconciliation_master(portal, source_uploaded_at)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_tasks_branch_order
            ON pending_tasks(branch_code, order_no)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_tasks_branch_status
            ON pending_tasks(branch_code, task_status)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_tasks_status_created
            ON pending_tasks(task_status, task_created_date)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_mir_order
            ON mir_details(order_no)
        """)
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_audit_activity_time
            ON activity_audit(activity_time)
        """)

        c.commit()

# Initialize the database without allowing a remote connection outage to dump a
# full traceback in the browser. This is deliberately fail-closed: when a cloud
# DATABASE_URL exists we do not silently switch to SQLite, because that could
# split live reconciliation history across two stores.
try:
    init_db()
except psycopg2.OperationalError as _startup_db_exc:
    _msg = str(_startup_db_exc).lower()
    st.error("Cloud database connection could not be established.")
    if "nxdomain" in _msg:
        st.warning(
            "Supabase's connection pooler can be reached, but it is currently "
            "unable to route to this project's database (NXDOMAIN). The app has "
            "stopped safely and has not switched to a different database."
        )
    elif "client encoding" in _msg:
        st.warning(
            "The PostgreSQL handshake reached the pooler but did not complete. "
            "This build already retries with SSL, UTF-8 client encoding and the "
            "Supabase transaction-pooler port."
        )
    else:
        st.warning(
            "Check the configured DATABASE_URL, database password and Supabase "
            "pooler availability. No reconciliation data has been deleted."
        )
    if st.button("Retry cloud database", type="primary"):
        try:
            init_db.clear()
        except Exception:
            pass
        st.rerun()
    st.stop()

# Visible persistence diagnostic.
if using_postgres():
    st.caption("Storage: Supabase PostgreSQL — persistent cloud database")
else:
    st.warning(
        "Storage: Local SQLite fallback. For Streamlit Cloud persistence, "
        "configure DATABASE_URL in Streamlit Secrets."
    )


def log_activity(action_type, portal="", branch_code="", order_no="", task_type="", filename="", details=""):
    with db() as c:
        c.execute("""
            INSERT INTO activity_audit(
                activity_time,action_type,portal,branch_code,order_no,
                task_type,filename,details
            ) VALUES(?,?,?,?,?,?,?,?)
        """,(
            datetime.now().isoformat(timespec="seconds"),
            txt(action_type),txt(portal),txt(branch_code),txt(order_no),
            txt(task_type),txt(filename),txt(details)
        ))
        c.commit()

def load_audit():
    with db() as c:
        return pd.read_sql_query("""
            SELECT activity_time,action_type,portal,branch_code,order_no,
                   task_type,filename,details
            FROM activity_audit
            ORDER BY id DESC
        """,c)

# ============================================================
# UTILITIES
# ============================================================

def safe_series(df, column, default=0):
    if column in df.columns:
        return df[column]
    return pd.Series([default] * len(df), index=df.index)

def source_adjustment_total(portal=None):
    try:
        with db() as c:
            if portal:
                row=c.execute(
                    "SELECT adjustment_total FROM source_kpis WHERE portal=?",
                    (portal,)
                ).fetchone()
            else:
                row=c.execute(
                    "SELECT COALESCE(SUM(adjustment_total),0) FROM source_kpis"
                ).fetchone()
        return float(row[0] or 0) if row else 0.0
    except Exception:
        return 0.0


def safe_numeric(df, column):
    return pd.to_numeric(safe_series(df, column, 0), errors="coerce").fillna(0)

def safe_text(df, column):
    return safe_series(df, column, "").fillna("").astype(str)

def txt(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()

def num(v):
    x = pd.to_numeric(v, errors="coerce")
    return 0.0 if pd.isna(x) else float(x)

def iso_date(v):
    x = pd.to_datetime(v, errors="coerce")
    return None if pd.isna(x) else x.date().isoformat()

def norm_col(v):
    return re.sub(r"[^a-z0-9]+", "", str(v).lower())

def clean_id(v):
    if v is None or pd.isna(v):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def find_col(df, aliases):
    mapping = {norm_col(c): c for c in df.columns}
    for a in aliases:
        k = norm_col(a)
        if k in mapping:
            return mapping[k]
    for a in aliases:
        k = norm_col(a)
        for nk, original in mapping.items():
            if k and k in nk:
                return original
    return None


# ============================================================
# FAST SOURCE WORKBOOK CACHE
# ============================================================
# A 15 MB marketplace workbook may contain several large sheets. Re-opening
# and re-parsing the same sheet many times during a single Run & Save cycle
# is expensive. Cache parsed source sheets only for the current app process.
_SOURCE_EXCEL_CACHE = {}
_SOURCE_SHEET_CACHE = {}


def _source_key(path):
    try:
        return str(Path(path).resolve())
    except Exception:
        return str(path)


def source_excel(path):
    key = _source_key(path)
    xl = _SOURCE_EXCEL_CACHE.get(key)
    if xl is None:
        # Load workbook bytes into memory before giving them to pandas.
        # This prevents Windows from keeping the uploaded temporary .xlsx
        # file open/locked while the reconciliation pipeline is running.
        raw = Path(path).read_bytes()
        xl = pd.ExcelFile(io.BytesIO(raw))
        _SOURCE_EXCEL_CACHE[key] = xl
    return xl


def source_sheet(path, aliases, fuzzy=False):
    """Read a source sheet once and reuse the parsed DataFrame."""
    xl = source_excel(path)
    sheet = None

    if fuzzy:
        try:
            sheet = flipkart_engine.find_sheet_name(xl.sheet_names, aliases)
        except Exception:
            sheet = None
    else:
        wanted = {norm_col(a) for a in aliases}
        for name in xl.sheet_names:
            if norm_col(name) in wanted:
                sheet = name
                break

    if not sheet:
        return pd.DataFrame()

    key = (_source_key(path), str(sheet))
    if key not in _SOURCE_SHEET_CACHE:
        _SOURCE_SHEET_CACHE[key] = pd.read_excel(xl, sheet_name=sheet)
    return _SOURCE_SHEET_CACHE[key]


def clear_source_cache(path=None):
    """Release workbook handles/DataFrames after a reconciliation cycle."""
    if path is None:
        keys = list(_SOURCE_EXCEL_CACHE.keys())
    else:
        keys = [_source_key(path)]

    for key in keys:
        xl = _SOURCE_EXCEL_CACHE.pop(key, None)
        if xl is not None:
            try:
                xl.close()
            except Exception:
                pass

        for skey in [x for x in list(_SOURCE_SHEET_CACHE) if x[0] == key]:
            _SOURCE_SHEET_CACHE.pop(skey, None)


def task_type_from_remark(remark):
    r = txt(remark)
    if not r or r.lower().startswith("reconciled"):
        return ""
    for key in TASK_KEYWORDS:
        if key.lower() in r.lower():
            return "Short Payment Received" if key == "Short Payment Received" else key
    return "Review Required"

def detect_portal(path):
    sheets = source_excel(path).sheet_names
    normalized = {norm_col(x) for x in sheets}

    amazon = {norm_col(x) for x in [
        "Orders","Reverse","Payments","Sales","Replacement_Map","Reimbursement"
    ]}
    if amazon.issubset(normalized):
        return "Amazon"

    aliases = [
        ("ERP Sales Register","ERP"),
        ("Flipkart Payments","Payments"),
        ("Flipkart Returns","Returns"),
        ("Flipkart All Orders","All Orders"),
        ("Flipkart Sales Report","Sales Report"),
    ]
    if all(flipkart_engine.find_sheet_name(sheets, a) for a in aliases):
        return "Flipkart"

    return None

def read_sheet(path, aliases):
    return source_sheet(path, aliases, fuzzy=False)

# ============================================================
# SOURCE ENRICHMENT
# ============================================================

def amazon_related_order_map(path):
    """Original Amazon order -> original + replacement order IDs."""
    mapping = {}
    try:
        repl = read_sheet(path,["Replacement_Map"])
        if repl.empty:
            return mapping
        original = find_col(repl,["Original Order Id"])
        replacement = find_col(repl,["Replacement Order Id"])
        if not original or not replacement:
            return mapping
        for _, r in repl.iterrows():
            o = clean_id(r.get(original))
            rp = clean_id(r.get(replacement))
            if not o or not rp:
                continue
            mapping.setdefault(o,[o])
            if rp not in mapping[o]:
                mapping[o].append(rp)
    except Exception:
        pass
    return mapping

def amazon_enrichment(path):
    """Order/Sales/Reverse enrichment using the headers in Amazon FY 26-27."""
    result = {}

    # Orders: SKU, shipment date and branch.
    try:
        orders = read_sheet(path,["Orders"])
        if not orders.empty:
            oid = find_col(orders,["Amazon Order Id"])
            sku = find_col(orders,["Merchant SKU","SKU"])
            ship = find_col(orders,["Shipment Date"])
            branch = find_col(orders,["Branch Code"])
            if oid:
                x=orders.copy(); x["__id__"]=x[oid].map(clean_id)
                for order_id,g in x.groupby("__id__"):
                    if not order_id: continue
                    d=result.setdefault(order_id,{})
                    if sku:
                        d["order_item"]=" | ".join(dict.fromkeys([txt(v) for v in g[sku] if txt(v)]))
                    if ship:
                        ds=pd.to_datetime(g[ship],errors="coerce").dropna()
                        if not ds.empty: d["shipment_date"]=ds.min().date().isoformat()
                    if branch:
                        vals=[txt(v) for v in g[branch] if txt(v)]
                        if vals: d["branch_code"]=vals[0]
    except Exception:
        pass

    # Reverse: operational return creation/delivery/status.
    try:
        rev=read_sheet(path,["Reverse"])
        if not rev.empty:
            oid=find_col(rev,["Order ID"])
            created=find_col(rev,["Return Created Date"])
            delivered=find_col(rev,[" Return Delivery Date","Return Delivery Date"])
            status=find_col(rev,["Return status","Return Status"])
            rtype=find_col(rev,["Return Type"])
            if oid:
                x=rev.copy(); x["__id__"]=x[oid].map(clean_id)
                for order_id,g in x.groupby("__id__"):
                    if not order_id: continue
                    d=result.setdefault(order_id,{})
                    if created:
                        ds=pd.to_datetime(g[created],errors="coerce").dropna()
                        if not ds.empty: d["return_created_date"]=ds.max().date().isoformat()
                    if delivered:
                        ds=pd.to_datetime(g[delivered],errors="coerce").dropna()
                        if not ds.empty: d["return_delivery_date"]=ds.max().date().isoformat()
                    if status:
                        d["return_status"]=" | ".join(dict.fromkeys([txt(v) for v in g[status] if txt(v)]))
                    if rtype:
                        d["return_type"]=" | ".join(dict.fromkeys([txt(v) for v in g[rtype] if txt(v)]))
    except Exception:
        pass

    # Sales: exact Invoice/CN numbers, dates, quantities and values.
    try:
        sales=read_sheet(path,["Sales"])
        if not sales.empty:
            oid=find_col(sales,["Po Number","PO Number"])
            inv=find_col(sales,["Invoice No"])
            invd=find_col(sales,["Invoice Date"])
            qty=find_col(sales,["Quantity"])
            price=find_col(sales,["Item Price","Line Amount","Gross Amount"])
            event=find_col(sales,["Sale/Return","Document Type"])
            item=find_col(sales,["Product/Item No","Item Description"])
            if oid:
                x=sales.copy(); x["__id__"]=x[oid].map(clean_id)
                for order_id,g in x.groupby("__id__"):
                    if not order_id: continue
                    d=result.setdefault(order_id,{})
                    is_return=pd.Series(False,index=g.index)
                    if event:
                        is_return=g[event].astype(str).str.contains("RETURN|CREDIT|CN",case=False,na=False)
                    elif qty:
                        is_return=pd.to_numeric(g[qty],errors="coerce").fillna(0)<0
                    sale_g=g[~is_return]; ret_g=g[is_return]
                    if item and not d.get("order_item"):
                        vals=[txt(v) for v in sale_g[item] if txt(v)]
                        if vals: d["order_item"]=" | ".join(dict.fromkeys(vals))
                    if inv and not sale_g.empty:
                        d["invoice_no"]=" | ".join(dict.fromkeys([txt(v) for v in sale_g[inv] if txt(v)]))
                    if invd and not sale_g.empty:
                        ds=pd.to_datetime(sale_g[invd],errors="coerce").dropna()
                        if not ds.empty: d["invoice_date"]=ds.max().date().isoformat()
                    if qty and not sale_g.empty:
                        d["invoice_qty"]=abs(float(pd.to_numeric(sale_g[qty],errors="coerce").fillna(0).sum()))
                    if price and not sale_g.empty:
                        d["invoice_price"]=abs(float(pd.to_numeric(sale_g[price],errors="coerce").fillna(0).sum()))
                    if inv and not ret_g.empty:
                        d["cn_no"]=" | ".join(dict.fromkeys([txt(v) for v in ret_g[inv] if txt(v)]))
                    if invd and not ret_g.empty:
                        ds=pd.to_datetime(ret_g[invd],errors="coerce").dropna()
                        if not ds.empty: d["cn_date"]=ds.max().date().isoformat()
                    if qty and not ret_g.empty:
                        d["cn_qty"]=abs(float(pd.to_numeric(ret_g[qty],errors="coerce").fillna(0).sum()))
                    if price and not ret_g.empty:
                        d["cn_price"]=abs(float(pd.to_numeric(ret_g[price],errors="coerce").fillna(0).sum()))
    except Exception:
        pass

    # Replacement IDs are not primary reconciliation rows.
    # Roll every applicable source detail back to the original order.
    related_map = amazon_related_order_map(path)

    def combine_text(values):
        out = []
        for value in values:
            if not value:
                continue
            for part in str(value).split("|"):
                part = part.strip()
                if part and part not in out:
                    out.append(part)
        return " | ".join(out)

    def latest_date(values):
        parsed = pd.to_datetime(
            [v for v in values if v],
            errors="coerce"
        )
        parsed = pd.Series(parsed).dropna()
        if parsed.empty:
            return None
        return parsed.max().date().isoformat()

    for original, related_ids in related_map.items():
        dest = result.setdefault(original,{})
        dest["replacement_order_id"] = " | ".join(related_ids[1:])

        related_details = [
            result.get(rid,{})
            for rid in related_ids
        ]

        # Text fields can legitimately exist on original and replacement IDs.
        for key in [
            "order_item","return_status","return_type",
            "invoice_no","cn_no"
        ]:
            combined = combine_text(
                [d.get(key) for d in related_details]
            )
            if combined:
                dest[key] = combined

        # Use latest operational/accounting date across related IDs.
        for key in [
            "return_created_date","return_delivery_date",
            "invoice_date","cn_date"
        ]:
            latest = latest_date(
                [d.get(key) for d in related_details]
            )
            if latest:
                dest[key] = latest

        # Sum invoice/CN quantities and values across original + replacements.
        for key in [
            "invoice_qty","invoice_price","cn_qty","cn_price"
        ]:
            values = [
                float(d.get(key,0) or 0)
                for d in related_details
            ]
            if any(values):
                dest[key] = sum(values)

    return result



def amazon_adjustment_source_total(path):
    """Exact Released Adjustment total from Amazon Payments C/AA."""
    try:
        p=read_sheet(path,["Payments"])
        if p.empty or p.shape[1] < 27:
            return 0.0
        mask=(
            p.iloc[:,2].fillna("").astype(str)
            .str.strip().str.casefold().eq("adjustment")
        )
        status=find_col(p,["Transaction Status"])
        if status:
            mask &= (
                p[status].fillna("").astype(str)
                .str.strip().str.casefold().eq("released")
            )
        return float(
            pd.to_numeric(p.loc[mask,p.columns[26]],errors="coerce")
            .fillna(0).sum()
        )
    except Exception:
        return 0.0



def amazon_adjustment_mapping_exact(path):
    """Exact Amazon Adjustment from Payments C/D/AA/AB."""
    try:
        p = read_sheet(path, ["Payments"])
    except Exception:
        return {}, 0.0

    if p.empty or p.shape[1] < 28:
        return {}, 0.0

    type_col = p.columns[2]      # C = Type
    order_col = p.columns[3]     # D = Order ID
    amount_col = p.columns[26]   # AA = Payment Received
    status_col = p.columns[27]   # AB = Transaction Status

    mask = (
        p[type_col].fillna("").astype(str)
        .str.strip().str.casefold().eq("adjustment")
    )
    mask &= (
        p[status_col].fillna("").astype(str)
        .str.strip().str.casefold().eq("released")
    )

    x = p.loc[mask, [order_col, amount_col]].copy()
    x["__order__"] = x[order_col].map(clean_id)
    x["__amount__"] = pd.to_numeric(
        x[amount_col], errors="coerce"
    ).fillna(0.0)

    source_total = float(x["__amount__"].sum())

    valid = x[x["__order__"] != ""]
    direct = (
        valid.groupby("__order__")["__amount__"].sum().to_dict()
        if not valid.empty else {}
    )
    direct = {
        clean_id(k): float(v)
        for k, v in direct.items()
        if clean_id(k)
    }
    return direct, source_total


def amazon_adjustment_rollup_exact(path):
    direct, source_total = amazon_adjustment_mapping_exact(path)
    rolled = dict(direct)

    try:
        related_map = amazon_related_order_map(path)
    except Exception:
        related_map = {}

    for original, related_ids in related_map.items():
        amount = sum(
            float(direct.get(clean_id(rid), 0.0) or 0.0)
            for rid in related_ids
        )
        if amount != 0:
            rolled[clean_id(original)] = amount

    return rolled, source_total


def backfill_amazon_adjustment_from_source(path):
    """
    Directly updates persisted Amazon order-wise Adjustment.
    KPI retains the complete source total, including unallocated Adjustment.
    """
    rolled, source_total = amazon_adjustment_rollup_exact(path)

    with db() as c:
        c.execute(
            "UPDATE reconciliation_master SET adjustment=0 WHERE portal='Amazon'"
        )

        db_orders = {
            clean_id(r[0])
            for r in c.execute(
                "SELECT order_no FROM reconciliation_master WHERE portal='Amazon'"
            ).fetchall()
            if clean_id(r[0])
        }

        matched_total = 0.0
        matched_orders = 0
        adjustment_rows = []
        for order_no, amount in rolled.items():
            if order_no not in db_orders:
                continue
            adjustment_rows.append((float(amount), order_no))
            matched_total += float(amount)
            matched_orders += 1

        if adjustment_rows:
            c.executemany(
                """
                UPDATE reconciliation_master
                SET adjustment=?
                WHERE portal='Amazon' AND order_no=?
                """,
                adjustment_rows
            )

        c.execute("""
            CREATE TABLE IF NOT EXISTS source_kpis(
                portal TEXT PRIMARY KEY,
                adjustment_total REAL DEFAULT 0,
                updated_at TEXT
            )
        """)
        c.execute("""
            INSERT INTO source_kpis(portal,adjustment_total,updated_at)
            VALUES(?,?,?)
            ON CONFLICT(portal) DO UPDATE SET
                adjustment_total=excluded.adjustment_total,
                updated_at=excluded.updated_at
        """,(
            "Amazon",
            float(source_total),
            datetime.now().isoformat(timespec="seconds")
        ))

        c.commit()

    return {
        "source_total": float(source_total),
        "matched_total": float(matched_total),
        "matched_orders": int(matched_orders),
        "unmatched_total": float(source_total - matched_total),
    }


def ensure_amazon_adjustment_backfilled():
    source = saved_source_path("Amazon")
    if not source.exists():
        return None

    try:
        with db() as c:
            row = c.execute("""
                SELECT COUNT(*), COALESCE(SUM(ABS(adjustment)),0)
                FROM reconciliation_master
                WHERE portal='Amazon'
            """).fetchone()
        amazon_rows = int(row[0] or 0)
        adjustment_abs = float(row[1] or 0)
    except Exception:
        return None

    if amazon_rows > 0 and adjustment_abs <= 0.001:
        return backfill_amazon_adjustment_from_source(source)
    return None


def amazon_payment_enrichment(path):
    """
    Vectorized Amazon Payments enrichment.

    The old implementation looped through every Order ID group and repeatedly
    converted the same columns to numeric/date/string values. On large Amazon
    payment sheets this was a major part of the 278s enrichment stage.
    """
    result = {}
    try:
        p = read_sheet(path, ["Payments"])
        if p.empty:
            return result

        oid = find_col(p, ["Order ID","Amazon Order Id","Amazon Order ID"])
        amount = next(
            (c for c in p.columns
             if str(c).strip() in {"Payment Received","Rece Amount"}),
            None
        )
        typ = find_col(p, ["Type"])
        status = find_col(p, ["Transaction Status"])
        pdate = find_col(p, ["Transaction Release Date","date/time"])
        settlement = find_col(p, ["settlement id"])

        adjustment_type_col = p.columns[2] if len(p.columns) >= 3 else typ
        adjustment_amount_col = p.columns[26] if len(p.columns) >= 27 else amount

        fee_cols = [
            find_col(p, ["selling fees"]),
            find_col(p, ["fba fees"]),
            find_col(p, ["other transaction fees"]),
            find_col(p, ["TCS-CGST"]),
            find_col(p, ["TCS-SGST"]),
            find_col(p, ["TCS-IGST"]),
            find_col(p, ["TDS (Section 194-O)"]),
        ]
        fee_cols = [c for c in fee_cols if c]

        if not oid:
            return result

        x = p.copy()
        x["__id__"] = x[oid].map(clean_id)
        x = x[x["__id__"].ne("")].copy()

        if status:
            released = (
                x[status].fillna("").astype(str)
                .str.contains("RELEASED", case=False, na=False)
            )
            x = x[released].copy()

        if x.empty:
            return result

        # Convert each expensive source column only once.
        if amount:
            x["__amount__"] = pd.to_numeric(x[amount], errors="coerce").fillna(0.0)
        else:
            x["__amount__"] = 0.0

        if adjustment_amount_col is not None:
            x["__adj_amount__"] = pd.to_numeric(
                x[adjustment_amount_col], errors="coerce"
            ).fillna(0.0)
        else:
            x["__adj_amount__"] = 0.0

        if typ:
            type_text = x[typ].fillna("").astype(str)
            x["__refund_amount__"] = x["__amount__"].where(
                type_text.str.contains("refund|reversal", case=False, na=False),
                0.0,
            )
            x["__refund_flag__"] = type_text.str.contains(
                "refund|reversal", case=False, na=False
            )
        else:
            x["__refund_amount__"] = x["__amount__"].where(x["__amount__"] < 0, 0.0)
            x["__refund_flag__"] = x["__amount__"] < 0

        if adjustment_type_col is not None:
            adj_flag = (
                x[adjustment_type_col].fillna("").astype(str)
                .str.strip().str.casefold().eq("adjustment")
            )
            x["__adjustment__"] = x["__adj_amount__"].where(adj_flag, 0.0)
        else:
            x["__adjustment__"] = 0.0

        x["__deductions__"] = 0.0
        for col in fee_cols:
            x["__deductions__"] += pd.to_numeric(
                x[col], errors="coerce"
            ).fillna(0.0).abs()

        if pdate:
            x["__pdate__"] = pd.to_datetime(x[pdate], errors="coerce")
            x["__refund_date__"] = x["__pdate__"].where(x["__refund_flag__"])
        else:
            x["__pdate__"] = pd.NaT
            x["__refund_date__"] = pd.NaT

        # Numeric/date aggregation is now done by pandas in a few groupby calls.
        numeric = x.groupby("__id__", sort=False).agg(
            net_pay_received=("__amount__", "sum"),
            refund=("__refund_amount__", "sum"),
            adjustment=("__adjustment__", "sum"),
            total_deductions=("__deductions__", "sum"),
            payment_date=("__pdate__", "max"),
            refund_date=("__refund_date__", "max"),
        )

        # Only text fields need custom unique concatenation.
        refs = {}
        statuses = {}
        if settlement:
            tmp = x.loc[:, ["__id__", settlement]].copy()
            tmp[settlement] = tmp[settlement].map(txt)
            tmp = tmp[tmp[settlement].ne("")]
            if not tmp.empty:
                refs = (
                    tmp.groupby("__id__", sort=False)[settlement]
                    .agg(lambda v: " | ".join(dict.fromkeys(v)))
                    .to_dict()
                )

        if status:
            tmp = x.loc[:, ["__id__", status]].copy()
            tmp[status] = tmp[status].map(txt)
            tmp = tmp[tmp[status].ne("")]
            if not tmp.empty:
                statuses = (
                    tmp.groupby("__id__", sort=False)[status]
                    .agg(lambda v: " | ".join(dict.fromkeys(v)))
                    .to_dict()
                )

        for order_id, row in numeric.iterrows():
            d = {
                "net_pay_received": float(row["net_pay_received"] or 0.0),
                "refund": float(row["refund"] or 0.0),
                "adjustment": float(row["adjustment"] or 0.0),
                "total_deductions": float(row["total_deductions"] or 0.0),
            }
            if pd.notna(row["payment_date"]):
                d["payment_date"] = row["payment_date"].date().isoformat()
            if pd.notna(row["refund_date"]):
                d["refund_date"] = row["refund_date"].date().isoformat()
            if order_id in refs:
                d["payment_reference"] = refs[order_id]
            if order_id in statuses:
                d["payment_transaction_status"] = statuses[order_id]
            result[order_id] = d

    except Exception:
        pass

    # Roll replacement-order settlement activity into the original order.
    related_map = amazon_related_order_map(path)
    for original, related_ids in related_map.items():
        dest = result.setdefault(original, {})
        related = [result.get(rid, {}) for rid in related_ids]

        dest["net_pay_received"] = sum(
            float(d.get("net_pay_received", 0) or 0) for d in related
        )
        dest["refund"] = sum(float(d.get("refund", 0) or 0) for d in related)
        dest["adjustment"] = sum(float(d.get("adjustment", 0) or 0) for d in related)
        dest["total_deductions"] = sum(
            float(d.get("total_deductions", 0) or 0) for d in related
        )

        refs = []
        dates = []
        refund_dates = []
        for d in related:
            if d.get("payment_reference"):
                refs.extend(
                    x.strip()
                    for x in str(d["payment_reference"]).split("|")
                    if x.strip()
                )
            if d.get("payment_date"):
                dates.append(d["payment_date"])
            if d.get("refund_date"):
                refund_dates.append(d["refund_date"])

        if refs:
            dest["payment_reference"] = " | ".join(dict.fromkeys(refs))
        if dates:
            dest["payment_date"] = max(dates)
        if refund_dates:
            dest["refund_date"] = max(refund_dates)

    return result


def amazon_reimbursement_enrichment(path):
    result={}
    try:
        r=read_sheet(path,["Reimbursement"])
        if r.empty: return result
        oid=find_col(r,["Order Id"]); reason=find_col(r,["Reason"]); approval=find_col(r,["approval-date"])
        if not oid: return result
        x=r.copy(); x["__id__"]=x[oid].map(clean_id)
        for order_id,g in x.groupby("__id__"):
            if not order_id: continue
            d=result.setdefault(order_id,{})
            if reason: d["reason"]=" | ".join(dict.fromkeys([txt(v) for v in g[reason] if txt(v)]))
            if approval:
                ds=pd.to_datetime(g[approval],errors="coerce").dropna()
                if not ds.empty: d["approval_date"]=ds.max().date().isoformat()
    except Exception: pass
    return result


def flipkart_enrichment(path):
    """ERP invoice/CN plus order/return details from Flipkart FY 26-27."""
    result={}
    # ERP Sales Register
    try:
        erp = source_sheet(path, ("ERP Sales Register","ERP"), fuzzy=True)
        if not erp.empty:
            po=find_col(erp,["Po Number"]); inv=find_col(erp,["Invoice No"]); invd=find_col(erp,["Invoice Date"])
            qty=find_col(erp,["Quantity"]); price=find_col(erp,["Line Amount","Gross Amount"])
            branch=find_col(erp,["Branch Code"]); tx=find_col(erp,["Sale/Return","Document Type"]); item=find_col(erp,["Product/Item No","Item Description"])
            if po:
                x=erp.copy(); x["__id__"]=x[po].map(clean_id)
                for order_id,g in x.groupby("__id__"):
                    if not order_id: continue
                    d=result.setdefault(order_id,{})
                    if branch:
                        vals=[txt(v) for v in g[branch] if txt(v)]
                        if vals: d["branch_code"]=vals[0]
                    is_cn=pd.Series(False,index=g.index)
                    if tx: is_cn=g[tx].astype(str).str.contains("RETURN|CREDIT|CN",case=False,na=False)
                    elif qty: is_cn=pd.to_numeric(g[qty],errors="coerce").fillna(0)<0
                    sale_g=g[~is_cn]; cn_g=g[is_cn]
                    if item:
                        vals=[txt(v) for v in sale_g[item] if txt(v)]
                        if vals: d["order_item"]=" | ".join(dict.fromkeys(vals))
                    if inv and not sale_g.empty: d["invoice_no"]=" | ".join(dict.fromkeys([txt(v) for v in sale_g[inv] if txt(v)]))
                    if invd and not sale_g.empty:
                        ds=pd.to_datetime(sale_g[invd],errors="coerce").dropna()
                        if not ds.empty: d["invoice_date"]=ds.max().date().isoformat()
                    if qty and not sale_g.empty: d["invoice_qty"]=abs(float(pd.to_numeric(sale_g[qty],errors="coerce").fillna(0).sum()))
                    if price and not sale_g.empty: d["invoice_price"]=abs(float(pd.to_numeric(sale_g[price],errors="coerce").fillna(0).sum()))
                    if inv and not cn_g.empty: d["cn_no"]=" | ".join(dict.fromkeys([txt(v) for v in cn_g[inv] if txt(v)]))
                    if invd and not cn_g.empty:
                        ds=pd.to_datetime(cn_g[invd],errors="coerce").dropna()
                        if not ds.empty: d["cn_date"]=ds.max().date().isoformat()
                    if qty and not cn_g.empty: d["cn_qty"]=abs(float(pd.to_numeric(cn_g[qty],errors="coerce").fillna(0).sum()))
                    if price and not cn_g.empty: d["cn_price"]=abs(float(pd.to_numeric(cn_g[price],errors="coerce").fillna(0).sum()))
    except Exception: pass

    # All Orders: product item + delivery date/status.
    try:
        orders = source_sheet(path, ("Flipkart All Orders","All Orders"), fuzzy=True)
        po=find_col(orders,["Po Number"]); item=find_col(orders,["sku","product_title"]); status=find_col(orders,["Order Item Status"]); delivery=find_col(orders,["order_delivery_date"])
        if po:
            x=orders.copy(); x["__id__"]=x[po].map(clean_id)
            for order_id,g in x.groupby("__id__"):
                if not order_id: continue
                d=result.setdefault(order_id,{})
                if item and not d.get("order_item"):
                    d["order_item"]=" | ".join(dict.fromkeys([txt(v) for v in g[item] if txt(v)]))
                if status: d["order_status"]=" | ".join(dict.fromkeys([txt(v) for v in g[status] if txt(v)]))
                if delivery:
                    ds=pd.to_datetime(g[delivery],errors="coerce").dropna()
                    if not ds.empty: d["delivery_date"]=ds.max().date().isoformat()
    except Exception: pass

    # Returns: approval and completion dates/status.
    try:
        ret = source_sheet(path, ("Flipkart Returns","Returns"), fuzzy=True)
        po=find_col(ret,["Po Number"]); approval=find_col(ret,["Return Approval Date"]); complete=find_col(ret,["return_completion_date"]); status=find_col(ret,["return_status"]); rtype=find_col(ret,["Return Type"])
        if po:
            x=ret.copy(); x["__id__"]=x[po].map(clean_id)
            for order_id,g in x.groupby("__id__"):
                if not order_id: continue
                d=result.setdefault(order_id,{})
                if approval:
                    ds=pd.to_datetime(g[approval],errors="coerce").dropna()
                    if not ds.empty: d["return_approval_date"]=ds.max().date().isoformat()
                if complete:
                    ds=pd.to_datetime(g[complete],errors="coerce").dropna()
                    if not ds.empty: d["return_completion_date"]=ds.max().date().isoformat()
                if status: d["return_status"]=" | ".join(dict.fromkeys([txt(v) for v in g[status] if txt(v)]))
                if rtype: d["return_type"]=" | ".join(dict.fromkeys([txt(v) for v in g[rtype] if txt(v)]))
    except Exception: pass
    return result



def flipkart_return_approval_enrichment(path):
    """
    Flipkart Return Delivery Date source of truth:
    Flipkart Returns sheet -> Return Approval Date, grouped by PO/Order ID.
    """
    result = {}
    try:
        returns = source_sheet(
            path, ("Flipkart Returns","Returns"), fuzzy=True
        )

        if returns.empty:
            return result

        oid_col = find_col(
            returns,
            ["PO Number","Po Number","Order ID","Order No"]
        )

        approval_col = None
        for c in returns.columns:
            if str(c).strip() == "Return Approval Date":
                approval_col = c
                break

        if not oid_col or approval_col is None:
            return result

        x = returns.copy()
        x["__id__"] = x[oid_col].map(clean_id)
        x["__approval__"] = pd.to_datetime(
            x[approval_col], errors="coerce"
        )

        grouped = x.groupby("__id__", as_index=False)["__approval__"].max()

        for _, r in grouped.iterrows():
            oid = clean_id(r["__id__"])
            dt = r["__approval__"]
            if oid and pd.notna(dt):
                result[oid] = dt.date().isoformat()

    except Exception:
        pass

    return result

def flipkart_payment_enrichment(path):
    result={}
    try:
        p = source_sheet(path, ("Flipkart Payments","Payments"), fuzzy=True)
        if p.empty: return result
        oid=find_col(p,["Po Number","PO Number","Order ID"])
        received=next(
            (
                c for c in p.columns
                if str(c).strip() in {
                    "Payment Received",
                    "Bank Settlement Value (Rs.)",
                    "Bank Settlement Value",
                }
            ),
            None
        )
        refund=find_col(p,["Refund"])
        # Exact Flipkart source mapping verified from supplied workbook:
        # Flipkart Payments Column P = Reimbursement.
        reimbursement = (
            p.columns[15]
            if p.shape[1] >= 16
            else find_col(p,["Reimbursement"])
        )
        pdate=find_col(p,["Payment Date"]); neft=find_col(p,["NEFT ID"]); neft_type=find_col(p,["Neft Type"])
        marketplace_fee=find_col(p,["Marketplace Fee"]); taxes=find_col(p,["Taxes"])
        if not oid: return result
        x=p.copy(); x["__id__"]=x[oid].map(clean_id)
        for order_id,g in x.groupby("__id__"):
            if not order_id: continue
            d=result.setdefault(order_id,{})
            if received: d["net_pay_received"]=float(pd.to_numeric(g[received],errors="coerce").fillna(0).sum())
            if refund:
                d["refund"]=float(
                    pd.to_numeric(g[refund],errors="coerce").fillna(0).sum()
                )
            if reimbursement:
                d["reimbursement"]=float(
                    pd.to_numeric(
                        g[reimbursement],errors="coerce"
                    ).fillna(0).sum()
                )
            total=0.0
            for col in [marketplace_fee,taxes]:
                if col: total += float(pd.to_numeric(g[col],errors="coerce").fillna(0).abs().sum())
            d["total_deductions"]=total
            if pdate:
                ds=pd.to_datetime(g[pdate],errors="coerce").dropna()
                if not ds.empty:
                    d["payment_date"]=ds.max().date().isoformat()
                    refund_mask=pd.to_numeric(g[refund],errors="coerce").fillna(0).abs()>1 if refund else pd.Series(False,index=g.index)
                    rds=pd.to_datetime(g.loc[refund_mask,pdate],errors="coerce").dropna()
                    if not rds.empty: d["refund_date"]=rds.max().date().isoformat()
            if neft: d["payment_reference"]=" | ".join(dict.fromkeys([txt(v) for v in g[neft] if txt(v)]))
            if neft_type: d["payment_transaction_status"]=" | ".join(dict.fromkeys([txt(v) for v in g[neft_type] if txt(v)]))
    except Exception: pass
    return result


def payment_status_from_amount(received_amount, remark=""):
    """Return payment status from actual settlement amount, never by default.

    Zero/blank payment must never be labelled Received.  Remarks still take
    priority for short-payment wording, while a positive receipt is the only
    condition that can produce Received.
    """
    received = num(received_amount)
    r = txt(remark).lower()

    # v15.5: Negative net receipt is a refund and must never be classified
    # as Received or Short Payment.
    if received < -0.000001:
        return "Refund"
    if "short payment" in r:
        return "Short Payment"
    if received > 0.000001:
        return "Received"
    return "Pending"


def settlement_status_from_remark(remark, payment_status=""):
    r = txt(remark)
    if r.lower().startswith("reconciled"):
        return "Reconciled"
    if r:
        return r
    if txt(payment_status):
        return txt(payment_status)
    return ""

# ============================================================
# NORMALIZE PORTAL RECONCILIATION
# ============================================================

def amazon_reverse_return_type_map(source_path):
    """
    Exact Amazon Return Type mapping.

    Reverse sheet:
      Column A = Customer Return / Courier Return
      Order ID = named Order ID column, with Column D fallback.

    Replacement-order return activity is rolled into the original Amazon order.
    """
    try:
        reverse = source_sheet(source_path, ["Reverse"], fuzzy=False)
    except Exception:
        return {}

    if reverse.empty:
        return {}

    reverse.columns = [
        re.sub(r"\s+"," ",str(c)).strip()
        for c in reverse.columns
    ]

    # Exact source requested by user: Reverse Column A.
    return_type_series = (
        reverse.iloc[:,0].fillna("").astype(str).str.strip()
    )

    order_col = None
    for candidate in [
        "Order ID","Order Id","Amazon Order Id","Order No","Order Number"
    ]:
        if candidate in reverse.columns:
            order_col = candidate
            break

    if order_col is not None:
        order_series = reverse[order_col]
    elif reverse.shape[1] >= 4:
        order_series = reverse.iloc[:,3]  # Column D fallback
    else:
        return {}

    direct = {}

    for oid, raw_type in zip(order_series, return_type_series):
        oid = clean_id(oid)
        raw_type = txt(raw_type)

        if not oid or not raw_type:
            continue

        upper = raw_type.upper()

        if "CUSTOMER" in upper:
            value = "Customer Return"
        elif "COURIER" in upper:
            value = "Courier Return"
        else:
            # Ignore unrelated Column-A values rather than inventing a type.
            continue

        existing = direct.get(oid, "")
        if not existing:
            direct[oid] = value
        else:
            values = [x.strip() for x in existing.split("|") if x.strip()]
            if value not in values:
                values.append(value)
            direct[oid] = " | ".join(values)

    # Roll replacement order activity back to original order.
    try:
        related_map = amazon_related_order_map(source_path)
    except Exception:
        related_map = {}

    rolled = dict(direct)

    for original, related_ids in related_map.items():
        values = []

        for rid in related_ids:
            value = direct.get(clean_id(rid), "")
            if not value:
                continue

            for part in value.split("|"):
                part = part.strip()
                if part and part not in values:
                    values.append(part)

        if values:
            rolled[clean_id(original)] = " | ".join(values)

    return rolled


def backfill_amazon_return_type_from_source(source_path):
    """
    Write exact Reverse Column-A Return Type into the persistent Amazon rows.
    """
    mapping = amazon_reverse_return_type_map(source_path)

    updated = 0
    customer = 0
    courier = 0

    with db() as c:
        # Clear only Amazon Return Type before rebuilding from the current source.
        c.execute(
            "UPDATE reconciliation_master SET return_status='' WHERE portal='Amazon'"
        )

        orders = {
            clean_id(row[0])
            for row in c.execute(
                "SELECT order_no FROM reconciliation_master WHERE portal='Amazon'"
            ).fetchall()
            if clean_id(row[0])
        }

        update_rows = []
        for order_no, return_type in mapping.items():
            order_no = clean_id(order_no)
            if not order_no or order_no not in orders:
                continue

            update_rows.append((return_type, order_no))
            updated += 1

            if "Customer Return" in return_type:
                customer += 1
            if "Courier Return" in return_type:
                courier += 1

        if update_rows:
            c.executemany(
                """
                UPDATE reconciliation_master
                SET return_status=?
                WHERE portal='Amazon' AND order_no=?
                """,
                update_rows
            )

        c.commit()

    return {
        "updated_orders": updated,
        "customer_return_orders": customer,
        "courier_return_orders": courier,
    }


def ensure_amazon_return_type_backfilled():
    """
    Repair older persistent databases when Amazon Return Type is still blank.
    """
    source = saved_source_path("Amazon")

    if not source.exists():
        return None

    try:
        with db() as c:
            row = c.execute("""
                SELECT
                    COUNT(*),
                    SUM(
                        CASE
                            WHEN COALESCE(TRIM(return_status),'')<>'' THEN 1
                            ELSE 0
                        END
                    )
                FROM reconciliation_master
                WHERE portal='Amazon'
            """).fetchone()

        amazon_rows = int(row[0] or 0)
        populated = int(row[1] or 0)

    except Exception:
        return None

    if amazon_rows > 0 and populated == 0:
        return backfill_amazon_return_type_from_source(source)

    return None



def normalize_amazon(reco, source_path):
    t0 = perf_counter()
    reverse_return_type_map = amazon_reverse_return_type_map(source_path)
    t1 = perf_counter()
    enrichment = amazon_enrichment(source_path)
    t2 = perf_counter()
    payment_enrichment = amazon_payment_enrichment(source_path)
    t3 = perf_counter()
    reimbursement_enrichment = amazon_reimbursement_enrichment(source_path)
    t4 = perf_counter()
    st.session_state["_last_amazon_enrichment_timing"] = {
        "Return Type": t1-t0,
        "Orders/Sales/Reverse": t2-t1,
        "Payments": t3-t2,
        "Reimbursement": t4-t3,
    }
    rows = []

    for _, r in reco.iterrows():
        oid = clean_id(r.get("Amazon Order Id"))
        if not oid:
            continue
        e = enrichment.get(oid,{})
        remark = txt(r.get("Remarks"))

        received = num(r.get("Rece Amount"))
        net_billing = num(r.get("Net Sale Item Price"))

        payment_status = payment_status_from_amount(received, remark)

        return_qty = num(r.get("Courier & Customer Return Qty"))

        rows.append({
            "portal": "Amazon",
            "order_no": oid,
            "order_date": e.get("shipment_date") or iso_date(r.get("Shipment Date")),
            "order_item": e.get("order_item",""),
            "order_qty": num(r.get("Shipped Quantity")),
            "repl_quantity": num(r.get("Repl Quantity")),
            "order_price": num(r.get("Item Price")),
            "replacement_item_price": num(r.get("Replacement Item Price")),
            "courier_customer_return_qty": num(r.get("Courier & Customer Return Qty")),
            "branch_code": e.get("branch_code") or txt(r.get("Order Branch Code")),
            "order_status": "Returned" if return_qty > 0 else "Delivered",
            "order_return_date": e.get("return_created_date") or iso_date(r.get("Return Created Date")),
            "return_completed_date": e.get("return_delivery_date"),
            "return_status": (
                reverse_return_type_map.get(oid,"")
                or next(
                    (
                        reverse_return_type_map.get(clean_id(x),"")
                        for x in str(r.get("Replacement Order Id","")).split(",")
                        if reverse_return_type_map.get(clean_id(x),"")
                    ),
                    ""
                )
            ),
            "payment_status": payment_status,
            "refund": abs(float(payment_enrichment.get(oid,{}).get("refund",0.0) or 0.0)),
            "reimbursement": abs(num(r.get("Reimbursement Amount"))),
            "adjustment": float(
                payment_enrichment.get(oid,{}).get("adjustment",0.0) or 0.0
            ),
            "invoice_no": e.get("invoice_no",""),
            "invoice_date": e.get("invoice_date"),
            "invoice_qty": num(r.get("Sale Qty")),
            "invoice_price": (
                num(r.get("Sale Amount"))
                if "Sale Amount" in reco.columns
                else e.get("invoice_price", net_billing)
            ),
            "cn_no": e.get("cn_no",""),
            "cn_date": e.get("cn_date"),
            "cn_qty": num(r.get("Return Qty")),
            "cn_price": (
                num(r.get("Return CN Amount"))
                if "Return CN Amount" in reco.columns
                else e.get("cn_price", 0.0)
            ),
            "replacement_order_id": txt(r.get("Replacement Order Id")) or e.get("replacement_order_id",""),
            # Core Amazon engine already aggregates Payment Received
            # across the original order + all replacement order IDs.
            "net_pay_received": num(r.get("Rece Amount")),
            "deferred_amount": num(r.get("Deferred Amount")),
            "total_deductions": payment_enrichment.get(oid,{}).get(
                "total_deductions", None
            ),
            "transaction_status": settlement_status_from_remark(remark,payment_status),
            "payment_date": payment_enrichment.get(oid,{}).get("payment_date"),
            "refund_date": payment_enrichment.get(oid,{}).get("refund_date"),
            "payment_reference": payment_enrichment.get(oid,{}).get("payment_reference",""),
            "reimbursement_reason": reimbursement_enrichment.get(oid,{}).get("reason",txt(r.get("Reason"))),
            "pending_remarks": remark,
        })

    return pd.DataFrame(rows)

def normalize_flipkart(reco, source_path):
    enrichment = flipkart_enrichment(source_path)
    return_approval = flipkart_return_approval_enrichment(source_path)
    payment_enrichment = flipkart_payment_enrichment(source_path)
    rows = []

    for _, r in reco.iterrows():
        oid = clean_id(r.get("PO Number"))
        if not oid:
            continue
        e = enrichment.get(oid,{})
        remark = txt(r.get("Remarks"))

        received = num(r.get("Received Amount"))
        billing = num(r.get("Net ERP Billing"))

        payment_status = payment_status_from_amount(received, remark)

        rows.append({
            "portal": "Flipkart",
            "order_no": oid,
            "order_date": iso_date(r.get("Order Date or Sales Report Order Date")),
            "order_item": e.get("order_item",""),
            "order_qty": num(r.get("Order Qty or Sales")),
            "repl_quantity": 0.0,
            "order_price": billing,
            "replacement_item_price": 0.0,
            "courier_customer_return_qty": num(r.get("Flipkart Return Qty")),
            "branch_code": txt(r.get("Branch Code")) or e.get("branch_code",""),
            "order_status": txt(r.get("Consider the Order Status") or r.get("Considered Order Status") or e.get("order_status") or r.get("Order Status")),
            "order_return_date": e.get("return_approval_date") or iso_date(r.get("Return Approval Date")),
            "return_completed_date": (
                iso_date(r.get("Return Approval Date"))
                or return_approval.get(oid)
            ),
            "return_status": e.get("return_status") or e.get("return_type") or txt(r.get("Return Status")),
            "payment_status": payment_status,
            "refund": abs(float(payment_enrichment.get(oid,{}).get("refund",num(r.get("Refund"))) or 0.0)),
            "reimbursement": float(
                payment_enrichment.get(oid,{}).get("reimbursement",0.0) or 0.0
            ),
            "adjustment": 0.0,
            "invoice_no": txt(r.get("Invoice No")) or e.get("invoice_no",""),
            "invoice_date": iso_date(r.get("Invoice Date")) or e.get("invoice_date"),
            "invoice_qty": num(r.get("ERP Billed Qty")),
            "invoice_price": e.get("invoice_price", billing),
            "cn_no": txt(r.get("CN No")) or e.get("cn_no",""),
            "cn_date": iso_date(r.get("CN Date")) or e.get("cn_date"),
            "cn_qty": num(r.get("ERP CN Qty")),
            "cn_price": (
                abs(num(r.get("ERP CN Amount")))
                if "ERP CN Amount" in reco.columns
                else e.get("cn_price",0.0)
            ),
            "replacement_order_id": txt(r.get("Replacement Order")),
            # Core Flipkart engine uses the marketplace Payment Received
            # settlement field at PO level.
            "net_pay_received": received,
            "deferred_amount": 0.0,
            "total_deductions": payment_enrichment.get(oid,{}).get(
                "total_deductions", None
            ),
            "transaction_status": settlement_status_from_remark(remark,payment_status),
            "payment_date": payment_enrichment.get(oid,{}).get("payment_date"),
            "refund_date": payment_enrichment.get(oid,{}).get("refund_date") or iso_date(r.get("Refund Date")),
            "payment_reference": payment_enrichment.get(oid,{}).get("payment_reference",""),
            "reimbursement_reason": (
                "Flipkart Payments Column P — Reimbursement"
                if float(payment_enrichment.get(oid,{}).get("reimbursement",0.0) or 0.0) != 0
                else ""
            ),
            "pending_remarks": remark,
        })

    return pd.DataFrame(rows)




def flipkart_reimbursement_map(source_path):
    """
    Exact Flipkart reimbursement mapping:
      Sheet: Flipkart Payments
      Column G: Po Number
      Column P: Reimbursement
    Values are summed PO-wise.
    """
    result = {}
    try:
        payments = source_sheet(
            source_path, ("Flipkart Payments","Payments"), fuzzy=True
        )
        if payments.empty:
            return result
    except Exception:
        return result

    if payments.empty or payments.shape[1] < 16:
        return result

    # Exact supplied layout: G = PO Number, P = Reimbursement.
    order_series = payments.iloc[:,6].map(clean_id)
    reimbursement_series = pd.to_numeric(
        payments.iloc[:,15],errors="coerce"
    ).fillna(0.0)

    temp = pd.DataFrame({
        "order_no": order_series,
        "reimbursement": reimbursement_series,
    })
    temp = temp[temp["order_no"].astype(str).str.strip()!=""]

    if temp.empty:
        return result

    grouped = temp.groupby("order_no",dropna=False)["reimbursement"].sum()

    return {
        clean_id(order_no): float(value or 0.0)
        for order_no,value in grouped.items()
        if clean_id(order_no)
    }


def backfill_flipkart_reimbursement_from_source(source_path):
    """
    Persist Flipkart Payments Column-P reimbursement into reconciliation_master.
    """
    mapping = flipkart_reimbursement_map(source_path)

    with db() as c:
        c.execute(
            "UPDATE reconciliation_master SET reimbursement=0 "
            "WHERE portal='Flipkart'"
        )

        db_orders = {
            clean_id(row[0])
            for row in c.execute(
                "SELECT order_no FROM reconciliation_master "
                "WHERE portal='Flipkart'"
            ).fetchall()
            if clean_id(row[0])
        }

        matched_orders = 0
        matched_total = 0.0

        for order_no,value in mapping.items():
            if order_no not in db_orders:
                continue

            c.execute(
                """
                UPDATE reconciliation_master
                SET reimbursement=?,
                    reimbursement_reason=?
                WHERE portal='Flipkart' AND order_no=?
                """,
                (
                    float(value),
                    "Flipkart Payments Column P — Reimbursement",
                    order_no,
                )
            )

            if float(value) != 0:
                matched_orders += 1
                matched_total += float(value)

        c.commit()

    return {
        "source_total": float(sum(mapping.values())),
        "matched_total": float(matched_total),
        "matched_orders": int(matched_orders),
    }


def portal_database_totals(portal):
    """Read the values actually stored in SQLite for dashboard use."""
    with db() as c:
        row = c.execute("""
            SELECT
                COUNT(*) AS total_rows,
                SUM(CASE
                    WHEN COALESCE(TRIM(replacement_order_id),'')<>'' THEN 1
                    ELSE 0
                END) AS replacement_orders,
                COALESCE(SUM(repl_quantity),0) AS repl_quantity,
                COALESCE(SUM(replacement_item_price),0) AS replacement_item_price,
                COALESCE(SUM(courier_customer_return_qty),0) AS courier_returns,
                COALESCE(SUM(net_pay_received),0) AS rece_amount,
                COALESCE(SUM(deferred_amount),0) AS deferred_amount,
                COALESCE(SUM(adjustment),0) AS adjustment
            FROM reconciliation_master
            WHERE portal=?
        """,(portal,)).fetchone()

    return {
        "rows": int(row[0] or 0),
        "replacement_orders": int(row[1] or 0),
        "repl_quantity": float(row[2] or 0),
        "replacement_item_price": float(row[3] or 0),
        "courier_returns": float(row[4] or 0),
        "rece_amount": float(row[5] or 0),
        "deferred_amount": float(row[6] or 0),
        "adjustment": float(row[7] or 0),
    }


def normalized_dashboard_totals(frame):
    """Expected dashboard totals before SQLite save."""
    if frame is None or frame.empty:
        return {
            "rows":0,"replacement_orders":0,"repl_quantity":0.0,
            "replacement_item_price":0.0,"courier_returns":0.0,
            "rece_amount":0.0,"deferred_amount":0.0,"adjustment":0.0,
        }

    return {
        "rows": len(frame),
        "replacement_orders": int(
            frame["replacement_order_id"].fillna("").astype(str)
            .str.strip().ne("").sum()
        ),
        "repl_quantity": float(
            pd.to_numeric(frame["repl_quantity"],errors="coerce").fillna(0).sum()
        ),
        "replacement_item_price": float(
            pd.to_numeric(
                frame["replacement_item_price"],errors="coerce"
            ).fillna(0).sum()
        ),
        "courier_returns": float(
            pd.to_numeric(
                frame["courier_customer_return_qty"],errors="coerce"
            ).fillna(0).sum()
        ),
        "rece_amount": float(
            pd.to_numeric(
                frame["net_pay_received"],errors="coerce"
            ).fillna(0).sum()
        ),
        "deferred_amount": float(
            pd.to_numeric(
                frame["deferred_amount"],errors="coerce"
            ).fillna(0).sum()
        ),
        "adjustment": float(
            pd.to_numeric(
                frame["adjustment"],errors="coerce"
            ).fillna(0).sum()
        ),
    }


def verify_portal_database_sync(portal, expected):
    """
    Verify that the latest source snapshot is present in the database.

    Structural reconciliation must be exact. Large financial aggregates are
    compared with a small relative tolerance because source repair/backfill
    steps and floating-point aggregation can legitimately move totals by a
    tiny amount after the initial engine result.
    """
    actual = portal_database_totals(portal)

    # Structural integrity must be exact.
    for key in ["rows", "replacement_orders"]:
        if int(round(float(actual.get(key, 0)))) != int(round(float(expected.get(key, 0)))):
            raise RuntimeError(
                "Database structural verification failed after reconciliation save: "
                f"{key}: engine={expected.get(key)} database={actual.get(key)}"
            )

    warnings_list = []

    for key in [
        "repl_quantity",
        "replacement_item_price",
        "courier_returns",
        "rece_amount",
        "deferred_amount",
        "adjustment",
    ]:
        exp = float(expected.get(key, 0) or 0)
        act = float(actual.get(key, 0) or 0)

        # Minimum ₹/qty tolerance plus 0.01% of the expected aggregate.
        tolerance = max(1.0, abs(exp) * 0.0001)

        if abs(act - exp) > tolerance:
            warnings_list.append(
                f"{key}: engine={exp} database={act}"
            )

    actual["_verification_warnings"] = warnings_list
    return actual


def save_source_workbook(portal, filename, workbook_bytes):
    """Persist the latest uploaded source workbook until a newer one replaces it."""
    payload = bytes(workbook_bytes)
    now = datetime.now().isoformat(timespec="seconds")
    digest = hashlib.sha256(payload).hexdigest()

    with db() as c:
        c.execute("""
            INSERT INTO source_workbooks(
                portal,filename,uploaded_at,workbook_data,
                workbook_sha256,workbook_size
            ) VALUES(?,?,?,?,?,?)
            ON CONFLICT(portal) DO UPDATE SET
                filename=excluded.filename,
                uploaded_at=excluded.uploaded_at,
                workbook_data=excluded.workbook_data,
                workbook_sha256=excluded.workbook_sha256,
                workbook_size=excluded.workbook_size
        """,(portal, filename, now, payload, digest, len(payload)))
        c.commit()


def source_workbook_info(portal=None):
    """Return persistent source snapshot metadata without loading workbook blobs."""
    with db() as c:
        if portal:
            rows = c.execute("""
                SELECT portal,filename,uploaded_at,workbook_sha256,workbook_size
                FROM source_workbooks
                WHERE portal=?
            """,(portal,)).fetchall()
        else:
            rows = c.execute("""
                SELECT portal,filename,uploaded_at,workbook_sha256,workbook_size
                FROM source_workbooks
                ORDER BY portal
            """).fetchall()

    return [
        {
            "portal": txt(r[0]),
            "filename": txt(r[1]),
            "uploaded_at": txt(r[2]),
            "sha256": txt(r[3]),
            "size": int(r[4] or 0),
        }
        for r in rows
    ]


def has_saved_source(portal):
    """Check the persistent source registry first; local files are fallback only."""
    try:
        if source_workbook_info(portal):
            return True
    except Exception:
        pass
    return (SOURCE_DIR / f"{str(portal).lower()}_latest.xlsx").exists()


def saved_source_path(portal):
    """
    Return a usable local path for the latest persistent source workbook.

    PostgreSQL/Supabase is the source of truth in production. The exact blob is
    re-materialised into the Streamlit container when needed, so redeploying the
    application never requires the user to upload the same source workbook again.
    """
    path = SOURCE_DIR / f"{str(portal).lower()}_latest.xlsx"

    try:
        with db() as c:
            row = c.execute("""
                SELECT workbook_data
                FROM source_workbooks
                WHERE portal=?
            """,(portal,)).fetchone()

        if row and row[0] is not None:
            data = bytes(row[0])
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(data)
            return path
    except Exception:
        # Local Windows/offline mode may not have a database snapshot blob.
        pass

    return path

def rebuild_portal_from_saved_source(portal):
    source = saved_source_path(portal)
    if not source.exists():
        raise FileNotFoundError(
            f"No saved {portal} source snapshot is available yet. "
            "Upload the current source workbook once."
        )

    info = source_workbook_info(portal)
    original_name = info[0]["filename"] if info else source.name

    normalized = process_workbook(source,portal)
    expected = normalized_dashboard_totals(normalized)
    upsert_reconciliation(normalized,original_name,record_upload=False)
    actual = verify_portal_database_sync(portal,expected)

    log_activity(
        "SOURCE REBUILD",
        portal=portal,
        filename=original_name,
        details=json.dumps(actual)
    )
    return actual

def source_field_coverage(frame):
    if frame is None or frame.empty:
        return {
            "rows": 0,
            "payment_rows": 0,
            "replacement_rows": 0,
            "invoice_rows": 0,
            "cn_rows": 0,
            "return_date_rows": 0,
        }

    def nonblank(col):
        if col not in frame.columns:
            return 0
        s = frame[col].fillna("").astype(str).str.strip()
        return int((s != "").sum())

    payment = pd.to_numeric(
        frame["net_pay_received"]
        if "net_pay_received" in frame.columns
        else pd.Series([0] * len(frame), index=frame.index),
        errors="coerce"
    ).fillna(0)

    return {
        "rows": len(frame),
        "payment_rows": int(payment.abs().gt(0).sum()),
        "replacement_rows": nonblank("replacement_order_id"),
        "replacement_qty_rows": int(
            pd.to_numeric(
                frame["repl_quantity"]
                if "repl_quantity" in frame.columns
                else pd.Series([0]*len(frame), index=frame.index),
                errors="coerce"
            ).fillna(0).abs().gt(0).sum()
        ),
        "replacement_value_rows": int(
            pd.to_numeric(
                frame["replacement_item_price"]
                if "replacement_item_price" in frame.columns
                else pd.Series([0]*len(frame), index=frame.index),
                errors="coerce"
            ).fillna(0).abs().gt(0).sum()
        ),
        "courier_return_rows": int(
            pd.to_numeric(
                frame["courier_customer_return_qty"]
                if "courier_customer_return_qty" in frame.columns
                else pd.Series([0]*len(frame), index=frame.index),
                errors="coerce"
            ).fillna(0).abs().gt(0).sum()
        ),
        "invoice_rows": nonblank("invoice_no"),
        "cn_rows": nonblank("cn_no"),
        "return_date_rows": nonblank("return_completed_date"),
        "refund_rows": int(
            pd.to_numeric(
                frame["refund"] if "refund" in frame.columns
                else pd.Series([0]*len(frame), index=frame.index),
                errors="coerce"
            ).fillna(0).abs().gt(0).sum()
        ),
        "reimbursement_rows": int(
            pd.to_numeric(
                frame["reimbursement"] if "reimbursement" in frame.columns
                else pd.Series([0]*len(frame), index=frame.index),
                errors="coerce"
            ).fillna(0).abs().gt(0).sum()
        ),
    }


def reconciliation_integrity(frame, portal):
    """Coverage report for fields that must not silently disappear."""
    if frame is None or frame.empty:
        return {}

    def nonblank(col):
        if col not in frame.columns:
            return 0
        return int(
            frame[col].fillna("").astype(str).str.strip().ne("").sum()
        )

    def nonzero(col):
        if col not in frame.columns:
            return 0
        return int(
            pd.to_numeric(frame[col], errors="coerce")
            .fillna(0).abs().gt(0).sum()
        )

    report = {
        "rows": len(frame),
        "replacement_id": nonblank("replacement_order_id"),
        "repl_qty": nonzero("repl_quantity"),
        "replacement_price": nonzero("replacement_item_price"),
        "courier_return_qty": nonzero("courier_customer_return_qty"),
        "received": nonzero("net_pay_received"),
        "deferred": nonzero("deferred_amount"),
        "sale_invoice": nonblank("invoice_no"),
        "return_invoice": nonblank("cn_no"),
        "return_delivery_date": nonblank("return_completed_date"),
    }
    return report

def process_workbook(path, portal, progress=None):
    started = perf_counter()

    def note(message):
        if progress is not None:
            try:
                progress.write(message)
            except Exception:
                pass

    note(f"1/4 Reading and reconciling {portal} source workbook…")

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        if portal == "Amazon":
            out = tmp / "amazon.xlsx"
            amazon_engine.process_reconciliation(path, out)
        elif portal == "Flipkart":
            out = tmp / "flipkart.xlsx"
            flipkart_engine.generate_reconciliation(path, out)
        else:
            raise ValueError(f"Unsupported portal: {portal}")

        engine_seconds = perf_counter() - started
        note(f"2/4 {portal} core engine completed in {engine_seconds:,.1f}s. Enriching source fields…")

        reco = pd.read_excel(out, sheet_name="Reconciliation")

    enrich_started = perf_counter()
    normalized = (
        normalize_amazon(reco, path)
        if portal == "Amazon"
        else normalize_flipkart(reco, path)
    )
    enrich_seconds = perf_counter() - enrich_started

    if portal == "Amazon":
        with db() as c:
            c.execute("""
                INSERT INTO source_kpis(portal,adjustment_total,updated_at)
                VALUES(?,?,?)
                ON CONFLICT(portal) DO UPDATE SET
                    adjustment_total=excluded.adjustment_total,
                    updated_at=excluded.updated_at
            """,(
                "Amazon",
                amazon_adjustment_source_total(path),
                datetime.now().isoformat(timespec="seconds")
            ))
            c.commit()

    note(
        f"3/4 {portal} source enrichment completed in "
        f"{enrich_seconds:,.1f}s. Preparing cloud save…"
    )
    if portal == "Amazon":
        timing = st.session_state.get("_last_amazon_enrichment_timing", {})
        if timing:
            note(
                "Amazon enrichment breakdown — "
                + " | ".join(
                    f"{k}: {v:,.1f}s" for k, v in timing.items()
                )
            )
    return normalized


# ============================================================
# OWNER MASTER
# ============================================================
@st.cache_data(ttl=60, show_spinner=False)
def load_owner_rules():
    with db() as c:
        return pd.read_sql_query("""
            SELECT branch_code, task_type, owner_name, owner_email
            FROM branch_owner_rules
            ORDER BY branch_code, task_type, owner_name
        """, c)

def save_owner_rules(df):
    df = df.copy()
    df.columns = [txt(c) for c in df.columns]
    required = {"Branch","Task Type","Owner Name","Owner Email"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError("Owner sheet missing: " + ", ".join(sorted(missing)))

    count = 0
    with db() as c:
        for _, r in df.iterrows():
            branch = txt(r["Branch"])
            task_type = txt(r["Task Type"])
            owner = txt(r["Owner Name"])
            email = txt(r["Owner Email"])

            if not branch or not task_type or not email:
                continue

            c.execute("""
                INSERT INTO branch_owner_rules(
                    branch_code,task_type,owner_name,owner_email
                ) VALUES(?,?,?,?)
                ON CONFLICT(branch_code,task_type,owner_email)
                DO UPDATE SET owner_name=excluded.owner_name
            """,(branch,task_type,owner,email))
            count += 1
        c.commit()

    invalidate_read_cache()
    return count

def get_assignees(branch, task_type):
    rules = load_owner_rules()
    if rules.empty:
        return "", ""

    branch = txt(branch)
    task_type = txt(task_type)

    x = rules[
        (rules["branch_code"].astype(str) == branch)
        & (rules["task_type"].astype(str) == task_type)
    ]

    if x.empty:
        x = rules[
            (rules["branch_code"].astype(str).str.lower() == "all branch")
            & (rules["task_type"].astype(str) == task_type)
        ]

    if x.empty:
        return "", ""

    names = "; ".join(dict.fromkeys(
        [txt(v) for v in x["owner_name"] if txt(v)]
    ))
    emails = "; ".join(dict.fromkeys(
        [txt(v) for v in x["owner_email"] if txt(v)]
    ))
    return names, emails

# ============================================================
# SAVE RECONCILIATION + CREATE TASKS
# ============================================================
MASTER_FIELDS = [
    "portal","order_no","order_date","order_item","order_qty","repl_quantity",
    "order_price","replacement_item_price","courier_customer_return_qty",
    "branch_code","order_status","order_return_date","return_completed_date",
    "return_status","payment_status","refund","reimbursement","adjustment","invoice_no",
    "invoice_date","invoice_qty","invoice_price","cn_no","cn_date","cn_qty",
    "cn_price","pending_remarks","replacement_order_id",
    "net_pay_received","deferred_amount","total_deductions","transaction_status",
    "payment_date","refund_date","payment_reference","reimbursement_reason"
]


def snapshot_operational_history(portal):
    """
    Snapshot all non-source operational history for one marketplace before
    replacing the marketplace reconciliation snapshot.
    """
    with db() as c:
        task_rows = c.execute("""
            SELECT
                portal,order_no,task_type,branch_code,task_created_date,
                task_status,working_date,task_completed_date,team_remarks,
                ticket_raised,raised_date,last_update
            FROM pending_tasks
            WHERE portal=?
        """,(portal,)).fetchall()

        task_cols = [
            "portal","order_no","task_type","branch_code","task_created_date",
            "task_status","working_date","task_completed_date","team_remarks",
            "ticket_raised","raised_date","last_update"
        ]

        tasks = [dict(zip(task_cols,row)) for row in task_rows]

        # MIR is branch/order based and may be used across marketplaces.
        # Keep all rows; they are never deleted by source refresh.
        mir_rows = c.execute("""
            SELECT *
            FROM mir_details
        """).fetchall()

        mir_cols = [
            x[1] for x in c.execute(
                "PRAGMA table_info(mir_details)"
            ).fetchall()
        ]

        mir = [dict(zip(mir_cols,row)) for row in mir_rows]

    return {
        "tasks": tasks,
        "mir": mir,
    }


def restore_operational_history(history):
    """
    Re-apply team/task history after source refresh.
    This is idempotent and does not overwrite newer nonblank team values.
    """
    if not history:
        return

    now = datetime.now().isoformat(timespec="seconds")

    with db() as c:
        # Restore task history.
        for r in history.get("tasks",[]):
            c.execute("""
                INSERT INTO pending_tasks(
                    portal,order_no,task_type,branch_code,
                    task_created_date,task_status,working_date,
                    task_completed_date,team_remarks,ticket_raised,
                    raised_date,last_update
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(portal,order_no,task_type) DO UPDATE SET
                    branch_code=COALESCE(NULLIF(excluded.branch_code,''),pending_tasks.branch_code),
                    task_created_date=COALESCE(
                        pending_tasks.task_created_date,
                        excluded.task_created_date
                    ),
                    task_status=CASE
                        WHEN COALESCE(TRIM(pending_tasks.task_status),'')=''
                        THEN excluded.task_status
                        ELSE pending_tasks.task_status
                    END,
                    working_date=COALESCE(
                        pending_tasks.working_date,
                        excluded.working_date
                    ),
                    task_completed_date=COALESCE(
                        pending_tasks.task_completed_date,
                        excluded.task_completed_date
                    ),
                    team_remarks=CASE
                        WHEN COALESCE(TRIM(pending_tasks.team_remarks),'')=''
                        THEN excluded.team_remarks
                        ELSE pending_tasks.team_remarks
                    END,
                    ticket_raised=CASE
                        WHEN COALESCE(TRIM(pending_tasks.ticket_raised),'')=''
                        THEN excluded.ticket_raised
                        ELSE pending_tasks.ticket_raised
                    END,
                    raised_date=COALESCE(
                        pending_tasks.raised_date,
                        excluded.raised_date
                    ),
                    last_update=CASE
                        WHEN pending_tasks.last_update IS NULL
                        THEN excluded.last_update
                        ELSE pending_tasks.last_update
                    END
            """,(
                r.get("portal",""),
                r.get("order_no",""),
                r.get("task_type",""),
                r.get("branch_code",""),
                r.get("task_created_date"),
                r.get("task_status",""),
                r.get("working_date"),
                r.get("task_completed_date"),
                r.get("team_remarks",""),
                r.get("ticket_raised",""),
                r.get("raised_date"),
                r.get("last_update") or now,
            ))

        # MIR rows are already persistent and source refresh does not delete them.
        # No destructive action is performed here intentionally.

        c.commit()


def operational_history_counts(portal):
    """Return persistent history counts for upload confirmation."""
    with db() as c:
        tasks = c.execute("""
            SELECT COUNT(*)
            FROM pending_tasks
            WHERE portal=?
        """,(portal,)).fetchone()[0]

        completed = c.execute("""
            SELECT COUNT(*)
            FROM pending_tasks
            WHERE portal=? AND task_status='Completed'
        """,(portal,)).fetchone()[0]

        working = c.execute("""
            SELECT COUNT(*)
            FROM pending_tasks
            WHERE portal=? AND task_status='Working'
        """,(portal,)).fetchone()[0]

        tickets = c.execute("""
            SELECT COUNT(*)
            FROM pending_tasks
            WHERE portal=?
              AND COALESCE(TRIM(ticket_raised),'')<>''
        """,(portal,)).fetchone()[0]

    return {
        "tasks": int(tasks or 0),
        "completed": int(completed or 0),
        "working": int(working or 0),
        "tickets": int(tickets or 0),
    }


def upsert_reconciliation(frame, source_file, record_upload=True):
    """
    Bulk-save one complete portal snapshot while preserving operational history.

    Source-derived fields may refresh. Team-owned fields are not auto-closed or
    rewritten by a source refresh; they change only through an explicit team update,
    except for the narrow v15.11 no-billing/no-payment returned-order closure rule.
    """
    if frame.empty:
        return 0

    # v15.11 business precedence:
    # If a physical return has been received, but the order was never billed and
    # no payment/refund occurred, there is no billing/payment/TEI/CN recovery left.
    # Keep Payment Status as Pending (zero receipt), but financially close the order.
    frame = frame.copy()
    return_qty = pd.to_numeric(
        frame.get("courier_customer_return_qty", 0), errors="coerce"
    ).fillna(0).abs()
    return_delivered = pd.to_datetime(
        frame.get("return_completed_date"), errors="coerce"
    ).notna()
    invoice_qty = pd.to_numeric(
        frame.get("invoice_qty", 0), errors="coerce"
    ).fillna(0).abs()
    invoice_no = frame.get(
        "invoice_no", pd.Series("", index=frame.index)
    ).fillna("").astype(str).str.strip()
    received = pd.to_numeric(
        frame.get("net_pay_received", 0), errors="coerce"
    ).fillna(0)
    refund = pd.to_numeric(
        frame.get("refund", 0), errors="coerce"
    ).fillna(0)

    returned_unbilled_no_payment = (
        (return_qty > 0.000001)
        & return_delivered
        & (invoice_qty <= 0.000001)
        & invoice_no.eq("")
        & (received.abs() <= 0.000001)
        & (refund.abs() <= 0.000001)
    )

    frame.loc[returned_unbilled_no_payment, "pending_remarks"] = "Reconciled"
    frame.loc[returned_unbilled_no_payment, "transaction_status"] = "Reconciled"

    auto_reconciled_keys = [
        (txt(r.get("portal")), clean_id(r.get("order_no")))
        for _, r in frame.loc[returned_unbilled_no_payment].iterrows()
        if txt(r.get("portal")) and clean_id(r.get("order_no"))
    ]

    now = datetime.now().isoformat(timespec="seconds")
    today = date.today().isoformat()
    portal_snapshot = txt(frame.iloc[0]["portal"])

    # Snapshot history before the source refresh. This remains intentionally
    # separate from source-derived reconciliation data.
    operational_history = snapshot_operational_history(portal_snapshot)

    placeholders = ",".join("?" for _ in MASTER_FIELDS)
    update_sql = ",".join(
        f"{k}=excluded.{k}"
        for k in MASTER_FIELDS
        if k not in {"portal", "order_no"}
    )
    master_sql = f"""
        INSERT INTO reconciliation_master(
            {','.join(MASTER_FIELDS)},source_file,source_uploaded_at
        )
        VALUES({placeholders},?,?)
        ON CONFLICT(portal,order_no) DO UPDATE SET
            {update_sql},
            source_file=excluded.source_file,
            source_uploaded_at=excluded.source_uploaded_at
    """

    # Build all master rows once; executemany removes thousands of Python/DB
    # round trips on Supabase compared with row-wise execute().
    master_rows = []
    incoming = []
    for _, r in frame.iterrows():
        master_rows.append(
            [r.get(k) for k in MASTER_FIELDS] + [source_file, now]
        )
        incoming.append({
            "portal": txt(r.get("portal")),
            "order_no": txt(r.get("order_no")),
            "branch": txt(r.get("branch_code")),
            "current_task": task_type_from_remark(r.get("pending_remarks")),
            "remark": txt(r.get("pending_remarks")),
        })

    with db() as c:
        c.execute(
            "DELETE FROM reconciliation_master WHERE portal=?",
            (portal_snapshot,)
        )
        c.executemany(master_sql, master_rows)

        # Fetch task state once instead of SELECT-per-order.
        task_rows = c.execute("""
            SELECT order_no,task_type,task_status,task_created_date
            FROM pending_tasks
            WHERE portal=?
        """, (portal_snapshot,)).fetchall()

        existing = {}
        for order_no, task_type, status, created in task_rows:
            existing[(txt(order_no), txt(task_type))] = {
                "status": txt(status),
                "created": created,
            }

        insert_rows = []
        touch_rows = []
        audit_rows = []
        seen_insert = set()
        seen_touch = set()

        for item in incoming:
            portal = item["portal"]
            order_no = item["order_no"]
            branch = item["branch"]
            current_task = item["current_task"]

            # Existing team tasks are historical/operational records. A source
            # refresh must never auto-complete them or rewrite team remarks/dates.
            if not current_task:
                continue

            key = (order_no, current_task)
            if key not in existing:
                dedupe = (portal, order_no, current_task)
                if dedupe not in seen_insert:
                    insert_rows.append((
                        portal, order_no, current_task, branch,
                        today, "Pending", "", now
                    ))
                    audit_rows.append((
                        now, "TASK CREATED", portal, branch, order_no,
                        current_task, source_file, item["remark"]
                    ))
                    seen_insert.add(dedupe)
                    existing[key] = {"status": "Pending", "created": today}
            else:
                dedupe = (portal, order_no, current_task)
                if dedupe not in seen_touch:
                    touch_rows.append((branch, portal, order_no, current_task))
                    seen_touch.add(dedupe)

        if insert_rows:
            c.executemany("""
                INSERT INTO pending_tasks(
                    portal,order_no,task_type,branch_code,
                    task_created_date,task_status,team_remarks,last_update
                ) VALUES(?,?,?,?,?,?,?,?)
                ON CONFLICT(portal,order_no,task_type) DO NOTHING
            """, insert_rows)

        if touch_rows:
            c.executemany("""
                UPDATE pending_tasks
                SET branch_code=?
                WHERE portal=? AND order_no=? AND task_type=?
            """, touch_rows)

        if audit_rows:
            c.executemany("""
                INSERT INTO activity_audit(
                    activity_time,action_type,portal,branch_code,order_no,
                    task_type,filename,details
                ) VALUES(?,?,?,?,?,?,?,?)
            """, audit_rows)

        if record_upload:
            c.execute("""
                INSERT INTO upload_history(portal,filename,uploaded_at,rows_processed)
                VALUES(?,?,?,?)
            """, (portal_snapshot, source_file, now, len(frame)))
        c.commit()

    # Re-attach any historical nonblank team/task fields idempotently.
    restore_operational_history(operational_history)

    # v15.11 narrow exception: returned + physically received + never billed +
    # never paid is fully reconciled. Old pending/working tasks for these orders
    # must not remain in the active Pending Task dashboard.
    if auto_reconciled_keys:
        with db() as c:
            c.executemany("""
                UPDATE pending_tasks
                SET task_status='Completed',
                    task_completed_date=COALESCE(task_completed_date,?),
                    team_remarks=CASE
                        WHEN COALESCE(TRIM(team_remarks),'')=''
                        THEN 'Auto-closed: return received; no billing/payment was created'
                        ELSE team_remarks
                    END,
                    last_update=?
                WHERE portal=? AND order_no=?
                  AND COALESCE(task_status,'')<>'Completed'
            """, [
                (today, now, portal, order_no)
                for portal, order_no in auto_reconciled_keys
            ])
            c.commit()

    # v15.12: global Reconciled -> Completed invariant after each source save.
    reconciled_keys = [
        (txt(r.get("portal")), clean_id(r.get("order_no")))
        for _, r in frame.iterrows()
        if txt(r.get("pending_remarks")).strip().lower() == "reconciled"
        and txt(r.get("portal"))
        and clean_id(r.get("order_no"))
    ]
    if reconciled_keys:
        with db() as c:
            c.executemany("""
                UPDATE pending_tasks
                SET task_status='Completed',
                    task_completed_date=COALESCE(task_completed_date,?),
                    team_remarks=CASE
                        WHEN COALESCE(TRIM(team_remarks),'')=''
                        THEN 'Auto-completed because Pending Remarks changed to Reconciled'
                        ELSE team_remarks
                    END,
                    last_update=?
                WHERE portal=? AND order_no=?
                  AND COALESCE(task_status,'')<>'Completed'
            """, [
                (date.today().isoformat(), now, portal, order_no)
                for portal, order_no in reconciled_keys
            ])
            c.commit()

    # Verify that the cloud snapshot contains the expected row count before
    # reporting success to the UI.
    with db() as c:
        saved = c.execute(
            "SELECT COUNT(*) FROM reconciliation_master WHERE portal=?",
            (portal_snapshot,)
        ).fetchone()
    saved_count = int(saved[0] or 0) if saved else 0
    expected_unique = int(frame["order_no"].map(clean_id).replace("", pd.NA).dropna().nunique())
    if saved_count != expected_unique:
        raise RuntimeError(
            f"Database verification failed for {portal_snapshot}: "
            f"expected {expected_unique} orders, found {saved_count}."
        )

    invalidate_read_cache()
    return len(frame)

# ============================================================
# TASK / MIR DATA
# ============================================================
@st.cache_data(ttl=30, show_spinner=False)
def load_tasks():
    """
    v15.16 fast task loader.

    Pending tasks and branch-owner rules are fetched once each. Owner assignment
    is mapped in memory instead of calling Supabase once per task row.
    """
    with db() as c:
        t = pd.read_sql_query("SELECT * FROM pending_tasks", c)
        rules = pd.read_sql_query("""
            SELECT branch_code, task_type, owner_name, owner_email
            FROM branch_owner_rules
        """, c)

    if t.empty:
        return t

    # Dates + aging are fully vectorized.
    for col in ["task_created_date","working_date","task_completed_date","last_update"]:
        if col in t.columns:
            t[col] = pd.to_datetime(t[col], errors="coerce")

    today = pd.Timestamp(date.today()).normalize()
    start = t["task_created_date"].dt.normalize()
    completed_mask = (
        t["task_status"].fillna("").astype(str).str.strip().eq("Completed")
        & t["task_completed_date"].notna()
    )
    end = pd.Series(today, index=t.index, dtype="datetime64[ns]")
    end.loc[completed_mask] = t.loc[
        completed_mask, "task_completed_date"
    ].dt.normalize()
    t["aging_days"] = (end - start).dt.days.fillna(0).clip(lower=0).astype(int)

    # Owner assignment: one rules query + in-memory lookup.
    t["task_assign_to"] = ""
    t["task_assign_email"] = ""

    if not rules.empty:
        for col in ["branch_code","task_type","owner_name","owner_email"]:
            rules[col] = rules[col].fillna("").astype(str).str.strip()

        def join_unique_values(series):
            vals = [v for v in series.tolist() if v]
            return "; ".join(dict.fromkeys(vals))

        exact = (
            rules.groupby(["branch_code","task_type"], sort=False, as_index=False)
            .agg({
                "owner_name": join_unique_values,
                "owner_email": join_unique_values,
            })
        )
        exact_map = {
            (r.branch_code, r.task_type): (r.owner_name, r.owner_email)
            for r in exact.itertuples(index=False)
        }

        all_branch = exact[
            exact["branch_code"].str.lower().eq("all branch")
        ]
        fallback_map = {
            r.task_type: (r.owner_name, r.owner_email)
            for r in all_branch.itertuples(index=False)
        }

        task_branch = t["branch_code"].fillna("").astype(str).str.strip()
        task_type = t["task_type"].fillna("").astype(str).str.strip()

        assignments = [
            exact_map.get((branch, typ), fallback_map.get(typ, ("", "")))
            for branch, typ in zip(task_branch.tolist(), task_type.tolist())
        ]
        if assignments:
            t["task_assign_to"] = [x[0] for x in assignments]
            t["task_assign_email"] = [x[1] for x in assignments]

    return t

def update_task(portal, order_no, task_type, status, team_remarks="", working_date=None, completed_date=None):
    status = txt(status).title()
    if status not in {"Pending","Working","Completed"}:
        raise ValueError("Task Status must be Pending, Working or Completed.")

    today = date.today().isoformat()
    wd = iso_date(working_date)
    cd = iso_date(completed_date)

    if status == "Working" and not wd:
        wd = today
    if status == "Completed" and not cd:
        cd = today

    with db() as c:
        c.execute("""
            UPDATE pending_tasks
            SET task_status=?,
                working_date=COALESCE(?,working_date),
                task_completed_date=CASE
                    WHEN ?='Completed' THEN COALESCE(?,task_completed_date,?)
                    ELSE task_completed_date
                END,
                team_remarks=CASE WHEN ?<>'' THEN ? ELSE team_remarks END,
                last_update=?
            WHERE portal=? AND order_no=? AND task_type=?
        """,(
            status,wd,status,cd,today,
            txt(team_remarks),txt(team_remarks),
            datetime.now().isoformat(timespec="seconds"),
            portal,order_no,task_type
        ))

        c.execute("""
            INSERT INTO activity_audit(activity_time,action_type,portal,branch_code,order_no,task_type,filename,details)
            SELECT ?,?,?,?,?,?,?,?
        """,(datetime.now().isoformat(timespec="seconds"),"TEAM TASK UPDATE",portal,"",order_no,task_type,"",f"Status={status}; Remarks={txt(team_remarks)}"))
        c.commit()

    invalidate_read_cache()

@st.cache_data(ttl=15, show_spinner=False)
def load_mir():
    with db() as c:
        m = pd.read_sql_query("""
            SELECT * FROM mir_details
        """, c)

    if m.empty:
        return m

    for col in ["inv_date","mir_date","tei_date","sr_date","updated_at"]:
        m[col] = pd.to_datetime(m[col], errors="coerce")
    if "tei_qty" in m.columns:
        m["tei_qty"] = pd.to_numeric(m["tei_qty"], errors="coerce")
    return m

def apply_mir_upload(branch, df):
    df = df.copy()
    df.columns = [txt(c) for c in df.columns]
    missing = set(MIR_UPLOAD_COLUMNS) - set(df.columns)
    if missing:
        raise ValueError("MIR upload missing: " + ", ".join(sorted(missing)))

    now = datetime.now().isoformat(timespec="seconds")
    count = 0

    with db() as c:
        for _, r in df.iterrows():
            order_no = clean_id(r.get("Order No"))
            if not order_no:
                continue

            c.execute("""
                INSERT INTO mir_details(
                    branch_code,order_no,invoice_no,inv_date,inv_qty,price,
                    product_code,mir_no,mir_date,tei_no,tei_date,tei_qty,sr_no,sr_date,
                    remarks,responsible_persons,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(branch_code,order_no) DO UPDATE SET
                    invoice_no=excluded.invoice_no,
                    inv_date=excluded.inv_date,
                    inv_qty=excluded.inv_qty,
                    price=excluded.price,
                    product_code=excluded.product_code,
                    mir_no=excluded.mir_no,
                    mir_date=excluded.mir_date,
                    tei_no=excluded.tei_no,
                    tei_date=excluded.tei_date,
                    tei_qty=excluded.tei_qty,
                    sr_no=excluded.sr_no,
                    sr_date=excluded.sr_date,
                    remarks=excluded.remarks,
                    responsible_persons=excluded.responsible_persons,
                    updated_at=excluded.updated_at
            """,(
                str(branch),order_no,txt(r.get("Invoice No")),
                iso_date(r.get("Inv Date")),num(r.get("Inv Qty")),
                num(r.get("Price")),txt(r.get("Product Code")),
                txt(r.get("MIR No")),iso_date(r.get("MIR Date")),
                txt(r.get("TEI No")),iso_date(r.get("TEI Date")),
                (
                    max(num(r.get("TEI Qty")), 1.0)
                    if txt(r.get("TEI No")) else 0.0
                ),
                txt(r.get("SR No")),iso_date(r.get("SR Date")),
                txt(r.get("Remarks")),MIR_BRANCH_PEOPLE[str(branch)],now
            ))

            # MIR / TEI sheets can close corresponding pending tasks automatically.
            if txt(r.get("MIR No")) and iso_date(r.get("MIR Date")):
                c.execute("""
                    UPDATE pending_tasks
                    SET task_status='Completed',
                        task_completed_date=COALESCE(task_completed_date,?),
                        team_remarks=CASE
                            WHEN COALESCE(team_remarks,'')='' THEN 'MIR updated by branch team'
                            ELSE team_remarks
                        END,
                        last_update=?
                    WHERE branch_code=? AND order_no=? AND task_type='MIR Pending'
                """,(date.today().isoformat(),now,str(branch),order_no))

            if txt(r.get("TEI No")) and iso_date(r.get("TEI Date")):
                tei_qty_now = max(num(r.get("TEI Qty")), 1.0)
                required_row = c.execute("""
                    SELECT COALESCE(MAX(ABS(courier_customer_return_qty)),0),
                           COALESCE(MAX(ABS(cn_qty)),0)
                    FROM reconciliation_master
                    WHERE branch_code=? AND order_no=?
                """,(str(branch),order_no)).fetchone()
                required_return_qty = max(
                    float(required_row[0] or 0),
                    float(required_row[1] or 0),
                    1.0,
                )
                if tei_qty_now + 0.000001 >= required_return_qty:
                    c.execute("""
                        UPDATE pending_tasks
                        SET task_status='Completed',
                            task_completed_date=COALESCE(task_completed_date,?),
                            team_remarks=CASE
                                WHEN COALESCE(team_remarks,'')='' THEN 'TEI quantity fully updated by branch team'
                                ELSE team_remarks
                            END,
                            last_update=?
                        WHERE branch_code=? AND order_no=? AND task_type='TEI Pending'
                    """,(date.today().isoformat(),now,str(branch),order_no))
                else:
                    c.execute("""
                        UPDATE pending_tasks
                        SET task_status=CASE
                                WHEN task_status='Completed' THEN 'Working'
                                WHEN COALESCE(TRIM(task_status),'')='' THEN 'Working'
                                ELSE task_status
                            END,
                            task_completed_date=NULL,
                            team_remarks=CASE
                                WHEN COALESCE(team_remarks,'')='' THEN 'Partial TEI quantity updated - balance TEI pending'
                                ELSE team_remarks
                            END,
                            last_update=?
                        WHERE branch_code=? AND order_no=? AND task_type='TEI Pending'
                    """,(now,str(branch),order_no))

            c.execute("""
                INSERT INTO activity_audit(activity_time,action_type,portal,branch_code,order_no,task_type,filename,details)
                VALUES(?,?,?,?,?,?,?,?)
            """,(now,"MIR UPDATE","",str(branch),order_no,"MIR/TEI","",f"MIR={txt(r.get('MIR No'))}; TEI={txt(r.get('TEI No'))}; TEI Qty={max(num(r.get('TEI Qty')), 1.0) if txt(r.get('TEI No')) else 0}; SR={txt(r.get('SR No'))}"))
            count += 1

        c.commit()

    invalidate_read_cache()
    return count

# ============================================================
# MAIN PERSISTENT VIEW
# ============================================================
@st.cache_data(ttl=15, show_spinner=False)
def load_master():
    """Load the persistent source reconciliation only.

    Task workflow and MIR/TEI are overlaid exactly once by
    _master_with_operational_overlay().
    """
    with db() as c:
        return pd.read_sql_query("""
            SELECT * FROM reconciliation_master
        """, c)


def _master_with_operational_overlay(master):
    if master.empty:
        return master

    tasks = load_tasks()
    mir = load_mir()

    if not tasks.empty:
        def join_unique(series):
            vals = []
            for v in series:
                s = txt(v)
                if s and s not in vals:
                    vals.append(s)
            return " | ".join(vals)

        def task_status_summary(series):
            vals = [txt(v) for v in series if txt(v)]
            if not vals:
                return ""
            if "Pending" in vals:
                return "Pending"
            if "Working" in vals:
                return "Working"
            if all(v == "Completed" for v in vals):
                return "Completed"
            return " | ".join(dict.fromkeys(vals))

        task_agg = tasks.groupby(["portal","order_no"],as_index=False).agg({
            "task_type":join_unique,
            "task_created_date":"min",
            "task_completed_date":"max",
            "task_status":task_status_summary,
            "team_remarks":join_unique,
            "ticket_raised":join_unique,
            "raised_date":"max",
            "aging_days":"max",
            "task_assign_to":join_unique,
            "task_assign_email":join_unique,
        })
        master = master.merge(task_agg,on=["portal","order_no"],how="left")
    else:
        for c in [
            "task_type","task_created_date","task_completed_date",
            "task_status","team_remarks","ticket_raised","raised_date","aging_days",
            "task_assign_to","task_assign_email"
        ]:
            master[c] = ""

    if not mir.empty:
        mir_keep = mir[[
            "branch_code","order_no","mir_no","mir_date","tei_no","tei_date","tei_qty",
            "sr_no","sr_date","remarks"
        ]].rename(columns={"remarks":"mir_remarks"})

        # Primary MIR match is Branch + Order No. If a source order has a blank
        # or differently formatted branch, fall back to unique Order No so the
        # branch-team MIR upload still updates the main reconciliation.
        master = master.merge(
            mir_keep,on=["branch_code","order_no"],how="left"
        )
        missing_mir = master["mir_no"].fillna("").astype(str).str.strip() == ""
        if missing_mir.any():
            unique_mir = mir_keep.drop_duplicates("order_no", keep=False).drop(
                columns=["branch_code"]
            )
            fallback = master.loc[missing_mir,["order_no"]].merge(
                unique_mir,on="order_no",how="left"
            )
            for col in [
                "mir_no","mir_date","tei_no","tei_date","tei_qty","sr_no","sr_date","mir_remarks"
            ]:
                if col in fallback.columns:
                    master.loc[missing_mir,col] = fallback[col].values
    else:
        for c in [
            "mir_no","mir_date","tei_no","tei_date","tei_qty",
            "sr_no","sr_date","mir_remarks"
        ]:
            master[c] = ""

    # v15.13: vectorized team/MIR/SR remark assembly for fast dashboard render.
    team_s = master.get(
        "team_remarks", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    mir_s = master.get(
        "mir_remarks", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    sr_s = master.get(
        "sr_no", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    sr_date_s = pd.to_datetime(
        master.get("sr_date", pd.Series(pd.NaT, index=master.index)),
        errors="coerce",
    ).dt.strftime("%Y-%m-%d").fillna("")

    combined = team_s.copy()
    add_mir = mir_s.ne("") & mir_s.ne(team_s)
    combined = combined.where(~add_mir, combined.where(combined.ne(""), "") + mir_s.where(combined.eq(""), " | " + mir_s))
    sr_piece = pd.Series("", index=master.index, dtype="object")
    has_sr = sr_s.ne("")
    sr_piece.loc[has_sr] = "SR No: " + sr_s.loc[has_sr]
    has_sr_date = has_sr & sr_date_s.ne("")
    sr_piece.loc[has_sr_date] = (
        sr_piece.loc[has_sr_date] + " | SR Date: " + sr_date_s.loc[has_sr_date]
    )
    add_sr = sr_piece.ne("")
    combined = combined.where(
        ~add_sr,
        combined.where(combined.ne(""), "") + sr_piece.where(combined.eq(""), " | " + sr_piece),
    )
    master["combined_team_remarks"] = combined

    # v15.13: vectorized Quantity-aware Return -> TEI -> CN workflow.
    current = master.get(
        "pending_remarks", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    return_date = pd.to_datetime(
        master.get("return_delivery_date", pd.Series(pd.NaT, index=master.index)),
        errors="coerce",
    )
    tei_no = master.get(
        "tei_no", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    tei_date = pd.to_datetime(
        master.get("tei_date", pd.Series(pd.NaT, index=master.index)),
        errors="coerce",
    )
    cn_no = master.get(
        "cn_no", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    cn_date = pd.to_datetime(
        master.get("cn_date", pd.Series(pd.NaT, index=master.index)),
        errors="coerce",
    )

    return_qty = pd.to_numeric(
        master.get("courier_customer_return_qty", 0), errors="coerce"
    ).fillna(0).abs()
    cn_qty = pd.to_numeric(
        master.get("cn_qty", 0), errors="coerce"
    ).fillna(0).abs()
    order_qty = pd.to_numeric(
        master.get("order_qty", 0), errors="coerce"
    ).fillna(0).abs()
    order_status = master.get(
        "order_status", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip().str.lower()

    return_qty = return_qty.where(return_qty.gt(0), cn_qty)
    returned_without_qty = return_qty.le(0) & order_status.eq("returned")
    return_qty = return_qty.where(
        ~returned_without_qty,
        order_qty.where(order_qty.gt(1), 1.0),
    )

    invoice_qty = pd.to_numeric(
        master.get("invoice_qty", 0), errors="coerce"
    ).fillna(0).abs()
    invoice_no = master.get(
        "invoice_no", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    received = pd.to_numeric(
        master.get("net_pay_received", 0), errors="coerce"
    ).fillna(0)
    refund = pd.to_numeric(
        master.get("refund", 0), errors="coerce"
    ).fillna(0)

    no_billing = invoice_qty.le(0.000001) & invoice_no.eq("")
    no_payment = received.abs().le(0.000001) & refund.abs().le(0.000001)

    tei_qty = pd.to_numeric(
        master.get("tei_qty", 0), errors="coerce"
    ).fillna(0)
    tei_covered = tei_qty.where(~(tei_no.ne("") & tei_qty.le(0)), 1.0)

    today_ts = pd.Timestamp(date.today())
    return_age = (today_ts - return_date.dt.normalize()).dt.days
    tei_age = (today_ts - tei_date.dt.normalize()).dt.days

    result = current.copy()
    locked = pd.Series(False, index=master.index)

    m = return_qty.gt(0) & return_date.notna() & no_billing & no_payment
    result.loc[m] = "Reconciled"
    locked |= m

    m = (
        ~locked & return_qty.gt(0) & return_date.notna()
        & tei_no.eq("") & tei_date.isna() & return_age.ge(2)
    )
    result.loc[m] = "Return Received TEI Pending"
    locked |= m

    m = (
        ~locked & return_qty.gt(0) & tei_no.ne("")
        & (tei_covered + 0.000001 < return_qty)
    )
    result.loc[m] = "Partial TEI Generated - Balance TEI Pending"
    locked |= m

    full_tei = (
        ~locked & return_qty.gt(0) & tei_no.ne("")
        & (tei_covered + 0.000001 >= return_qty)
        & cn_no.eq("") & cn_date.isna()
    )
    aged_full_tei = full_tei & tei_date.notna() & tei_age.ge(5)
    result.loc[aged_full_tei] = "TEI Generated but CN Pending"
    result.loc[full_tei & ~aged_full_tei] = "CN Pending"
    locked |= full_tei

    legacy_cn = (
        ~locked
        & current.str.lower().str.contains("cn pending", regex=False)
        & tei_no.ne("") & tei_date.notna()
        & cn_no.eq("") & cn_date.isna()
        & tei_age.ge(5)
    )
    result.loc[legacy_cn] = "TEI Generated but CN Pending"

    master["pending_remarks"] = result

    # Pending Remarks is the operational source of truth for this closure class.
    # Do not change Payment Status; zero receipt correctly remains Pending.
    reconciled_mask = (
        master["pending_remarks"].fillna("").astype(str).str.strip().str.lower()
        == "reconciled"
    )
    master.loc[reconciled_mask, "transaction_status"] = "Reconciled"
    return master

def _fast_dashboard_operational_overlay(master):
    """
    v15.15 read-only dashboard overlay.

    The reconciliation/business result (Pending Remarks, Transaction Status,
    Payment Status, return/TEI/CN state) is persisted when source/team data is
    saved. Normal dashboard viewing must therefore NOT rerun those business
    rules over every order.

    This function only attaches the small operational task + MIR/TEI tables.
    """
    if master.empty:
        return master

    master = master.copy()
    tasks = load_tasks()
    mir = load_mir()

    # ---- Task overlay: aggregate only the ~task-table rows, not the master.
    if not tasks.empty:
        def _join_unique_fast(series):
            vals = (
                series.fillna("")
                .astype(str)
                .str.strip()
            )
            vals = vals[(vals != "") & (vals.str.lower() != "nan")]
            return " | ".join(dict.fromkeys(vals.tolist()))

        def _task_status_fast(series):
            vals = (
                series.fillna("")
                .astype(str)
                .str.strip()
            )
            vals = vals[vals != ""].tolist()
            if not vals:
                return ""
            if "Pending" in vals:
                return "Pending"
            if "Working" in vals:
                return "Working"
            if all(v == "Completed" for v in vals):
                return "Completed"
            return " | ".join(dict.fromkeys(vals))

        task_agg = tasks.groupby(
            ["portal", "order_no"], as_index=False, sort=False
        ).agg({
            "task_type": _join_unique_fast,
            "task_created_date": "min",
            "task_completed_date": "max",
            "task_status": _task_status_fast,
            "team_remarks": _join_unique_fast,
            "ticket_raised": _join_unique_fast,
            "raised_date": "max",
            "aging_days": "max",
            "task_assign_to": _join_unique_fast,
            "task_assign_email": _join_unique_fast,
        })
        master = master.merge(
            task_agg, on=["portal", "order_no"], how="left", sort=False
        )
    else:
        for col in [
            "task_type","task_created_date","task_completed_date",
            "task_status","team_remarks","ticket_raised","raised_date",
            "aging_days","task_assign_to","task_assign_email",
        ]:
            master[col] = ""

    # ---- MIR overlay.
    if not mir.empty:
        mir_keep = mir[[
            "branch_code","order_no","mir_no","mir_date","tei_no","tei_date",
            "tei_qty","sr_no","sr_date","remarks"
        ]].rename(columns={"remarks": "mir_remarks"})

        master = master.merge(
            mir_keep,
            on=["branch_code", "order_no"],
            how="left",
            sort=False,
        )

        # Unique-order fallback only for still-unmatched rows.
        missing = master["mir_no"].fillna("").astype(str).str.strip().eq("")
        if missing.any():
            unique_mir = (
                mir_keep.drop_duplicates("order_no", keep=False)
                .drop(columns=["branch_code"])
                .set_index("order_no")
            )
            order_index = master.loc[missing, "order_no"]
            for col in [
                "mir_no","mir_date","tei_no","tei_date","tei_qty",
                "sr_no","sr_date","mir_remarks",
            ]:
                if col in unique_mir.columns:
                    mapped = order_index.map(unique_mir[col])
                    master.loc[missing, col] = mapped.values
    else:
        for col in [
            "mir_no","mir_date","tei_no","tei_date","tei_qty",
            "sr_no","sr_date","mir_remarks",
        ]:
            master[col] = ""

    # ---- Team remarks, vectorized.
    team_s = master.get(
        "team_remarks", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    mir_s = master.get(
        "mir_remarks", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()

    combined = team_s.copy()
    mir_only = mir_s.ne("") & mir_s.ne(team_s)
    both = mir_only & combined.ne("")
    combined.loc[mir_only & combined.eq("")] = mir_s.loc[
        mir_only & combined.eq("")
    ]
    combined.loc[both] = combined.loc[both] + " | " + mir_s.loc[both]

    sr_s = master.get(
        "sr_no", pd.Series("", index=master.index)
    ).fillna("").astype(str).str.strip()
    sr_date_s = pd.to_datetime(
        master.get("sr_date", pd.Series(pd.NaT, index=master.index)),
        errors="coerce",
    ).dt.strftime("%Y-%m-%d").fillna("")

    has_sr = sr_s.ne("")
    sr_piece = pd.Series("", index=master.index, dtype="object")
    sr_piece.loc[has_sr] = "SR No: " + sr_s.loc[has_sr]
    with_date = has_sr & sr_date_s.ne("")
    sr_piece.loc[with_date] = (
        sr_piece.loc[with_date] + " | SR Date: " + sr_date_s.loc[with_date]
    )

    sr_only = sr_piece.ne("")
    both_sr = sr_only & combined.ne("")
    combined.loc[sr_only & combined.eq("")] = sr_piece.loc[
        sr_only & combined.eq("")
    ]
    combined.loc[both_sr] = combined.loc[both_sr] + " | " + sr_piece.loc[both_sr]
    master["combined_team_remarks"] = combined

    return master


def ecom_process_display(master):
    """Fast attached reconciliation output structure + hidden filter columns."""
    if master.empty:
        return pd.DataFrame()

    master = _fast_dashboard_operational_overlay(master)
    out = pd.DataFrame(index=master.index)

    out["Market Place"] = master["portal"]
    out["Task Assign To"] = master.get("task_assign_to", "")

    task_type_series = (
        master["task_type"].fillna("").astype(str)
        if "task_type" in master.columns
        else pd.Series("", index=master.index)
    )
    remark_series = master["pending_remarks"].fillna("").astype(str)
    task_map = {
        value: task_type_from_remark(value)
        for value in remark_series.drop_duplicates().tolist()
    }
    derived_task_type = remark_series.map(task_map).fillna("")
    out["Task Type"] = task_type_series.where(
        task_type_series.str.strip().ne(""),
        derived_task_type,
    )

    out["Branch Code"] = master["branch_code"]
    out["Order Date/Shipment Date"] = master["order_date"]
    out["Order Item"] = master["order_item"]
    out["Order No"] = master["order_no"]
    out["Replacement Order Id"] = master.get("replacement_order_id", "")
    out["Order Qty"] = master["order_qty"]
    out["Repl Quantity"] = master.get("repl_quantity", 0)
    out["Order Price"] = master["order_price"]
    out["Replacement Item Price"] = master.get("replacement_item_price", 0)
    out["Courier & Customer Return Qty"] = master.get(
        "courier_customer_return_qty", 0
    )
    out["Order Status"] = master["order_status"]
    out["Return Created Date"] = master["order_return_date"]
    out["Return Delivery Date"] = master["return_completed_date"]
    out["Return Type"] = master["return_status"]
    out["Payment Status"] = master["payment_status"]
    out["Rece Amount"] = master.get("net_pay_received", 0)
    out["Rece Amount Date"] = master.get("payment_date", "")
    out["Deferred Amount"] = master.get("deferred_amount", 0)
    out["Refund"] = master["refund"]
    out["Refund Date"] = master.get("refund_date", "")
    out["Reimbursement"] = master["reimbursement"]
    out["Adjustment"] = master.get("adjustment", 0)
    out["Sale Invoice No"] = master["invoice_no"]
    out["Sale Invoice Date"] = master["invoice_date"]
    out["Invoice Qty"] = master["invoice_qty"]
    out["Return Invoice No"] = master["cn_no"]
    out["Return Invoice Date"] = master["cn_date"]
    out["Return Qty"] = master["cn_qty"]
    out["Sale Price"] = master["invoice_price"]
    out["Pending Remarks"] = master["pending_remarks"]
    out["Pending Task Created Date"] = master.get("task_created_date", "")
    out["MIR No"] = master.get("mir_no", "")
    out["MIR Date"] = master.get("mir_date", "")
    out["TEI No"] = master.get("tei_no", "")
    out["TEI Date"] = master.get("tei_date", "")
    out["TEI Qty"] = master.get("tei_qty", 0)
    out["Ticket Raised If Any"] = master.get("ticket_raised", "")
    out["Ticket Raised Date"] = master.get("raised_date", "")
    out["Team Remarks"] = master.get("combined_team_remarks", "")
    out["Task Completed Date"] = master.get("task_completed_date", "")

    raw_task_status = (
        master["task_status"].fillna("").astype(str)
        if "task_status" in master.columns
        else pd.Series("", index=master.index)
    )
    active_issue = out["Task Type"].fillna("").astype(str).str.strip().ne("")
    out["Task Status"] = raw_task_status.where(
        raw_task_status.str.strip().ne(""),
        active_issue.map({True: "Pending", False: ""}),
    )
    out["Aging Days of Task Pending/Completed"] = master.get("aging_days", "")
    return out


def settlement_process_display(master):
    """Exact attached Settlement Reco structure + hidden filter columns."""
    if master.empty:
        return pd.DataFrame()

    master = _master_with_operational_overlay(master)
    out = pd.DataFrame()

    # Hidden/filter fields.
    out["Market Place"] = master["portal"]
    out["Branch Code"] = master["branch_code"]
    out["Task Type"] = master.get("task_type","")
    out["Task Status"] = master.get("task_status","")
    out["Task Assign To"] = master.get("task_assign_to","")

    # Exact attached visible structure.
    out["Order No"] = master["order_no"]
    out["Replacement Order"] = master.get("replacement_order_id","")
    out["Order Date/Shipment Date"] = master["order_date"]
    out["Order Qty"] = master["order_qty"]
    out["Order Price"] = master["order_price"]
    out["Sale (Invoice No)"] = master["invoice_no"]
    out["Sale Date (Invoice Date)"] = master["invoice_date"]
    out["Sale Qty"] = master["invoice_qty"]
    out["Sale Price"] = master["invoice_price"]
    out["Return (CN No)"] = master["cn_no"]
    out["Return Date (CN Date)"] = master["cn_date"]
    out["Return Qty"] = master["cn_qty"]
    out["Return Price (CN Price)"] = master["cn_price"]
    out["Payment Received"] = master.get("net_pay_received",0)
    out["Payment Received Date"] = master.get("payment_date","")
    out["Deferred Amount"] = master.get("deferred_amount",0)
    out["Refund"] = master.get("refund",0)
    out["Refund Date"] = master.get("refund_date","")
    out["Total Deductions"] = master.get("total_deductions",None)
    out["Transaction Status"] = master.get("transaction_status","")
    return out


def apply_settlement_date_filters(df):
    """
    Settlement dashboard filters:
      1. Market Place
      2. Order Date Range
      3. Invoice Date Range
      4. Quarter
      5. Year

    Quarter and Year are based on Order Date/Shipment Date.
    """
    if df.empty:
        return df

    view = df.copy()

    order_dates = pd.to_datetime(
        safe_series(view, "Order Date/Shipment Date", ""),
        errors="coerce"
    )

    # Settlement display uses Sale Date (Invoice Date).
    invoice_col = (
        "Sale Date (Invoice Date)"
        if "Sale Date (Invoice Date)" in view.columns
        else "Sale Invoice Date"
    )
    invoice_dates = pd.to_datetime(
        safe_series(view, invoice_col, ""),
        errors="coerce"
    )

    c0, c1, c2, c3, c4 = st.columns(5)

    marketplace_col = (
        "Market Place"
        if "Market Place" in view.columns
        else "portal"
    )
    marketplace_values = sorted(
        {
            str(v).strip()
            for v in safe_series(view, marketplace_col, "").tolist()
            if str(v).strip()
        }
    )
    marketplace = c0.selectbox(
        "Market Place",
        ["All"] + marketplace_values,
        index=0,
        key="settlement_marketplace",
    )

    valid_order_dates = order_dates.dropna()
    if not valid_order_dates.empty:
        order_min = valid_order_dates.min().date()
        order_max = valid_order_dates.max().date()
        order_range = c1.date_input(
            "Order Date Range",
            value=(order_min, order_max),
            min_value=order_min,
            max_value=order_max,
            key="settlement_order_date_range",
        )
    else:
        order_range = None
        c1.caption("Order Date Range: No dates")

    valid_invoice_dates = invoice_dates.dropna()
    if not valid_invoice_dates.empty:
        invoice_min = valid_invoice_dates.min().date()
        invoice_max = valid_invoice_dates.max().date()
        invoice_range = c2.date_input(
            "Invoice Date Range",
            value=(invoice_min, invoice_max),
            min_value=invoice_min,
            max_value=invoice_max,
            key="settlement_invoice_date_range",
        )
    else:
        invoice_range = None
        c2.caption("Invoice Date Range: No dates")

    quarter_options = ["All", "Q1", "Q2", "Q3", "Q4"]
    quarter = c3.selectbox(
        "Quarter",
        quarter_options,
        index=0,
        key="settlement_quarter",
    )

    years = sorted(
        {
            int(y)
            for y in order_dates.dt.year.dropna().tolist()
        },
        reverse=True,
    )
    year_options = ["All"] + years
    year = c4.selectbox(
        "Year",
        year_options,
        index=0,
        key="settlement_year",
    )

    mask = pd.Series(True, index=view.index)

    if marketplace != "All":
        mask &= (
            safe_series(view, marketplace_col, "")
            .astype(str)
            .str.strip()
            .eq(marketplace)
        )

    if isinstance(order_range, (tuple, list)) and len(order_range) == 2:
        start_date, end_date = order_range
        mask &= order_dates.dt.date.between(start_date, end_date)

    if isinstance(invoice_range, (tuple, list)) and len(invoice_range) == 2:
        start_date, end_date = invoice_range
        mask &= invoice_dates.dt.date.between(start_date, end_date)

    if quarter != "All":
        quarter_number = int(quarter[-1])
        mask &= order_dates.dt.quarter.eq(quarter_number)

    if year != "All":
        mask &= order_dates.dt.year.eq(int(year))

    return view[mask].copy()


ECOM_VISIBLE_COLUMNS = [
    "Branch Code",
    "Order Date/Shipment Date",
    "Order Item",
    "Order No",
    "Replacement Order Id",
    "Order Qty",
    "Repl Quantity",
    "Order Price",
    "Replacement Item Price",
    "Courier & Customer Return Qty",
    "Order Status",
    "Return Created Date",
    "Return Delivery Date",
    "Return Type",
    "Payment Status",
    "Rece Amount",
    "Rece Amount Date",
    "Deferred Amount",
    "Refund",
    "Refund Date",
    "Reimbursement",
    "Adjustment",
    "Sale Invoice No",
    "Sale Invoice Date",
    "Invoice Qty",
    "Return Invoice No",
    "Return Invoice Date",
    "Return Qty",
    "Sale Price",
    "Pending Remarks",
    "Pending Task Created Date",
    "MIR No",
    "MIR Date",
    "TEI No",
    "TEI Date",
    "TEI Qty",
    "Ticket Raised If Any",
    "Ticket Raised Date",
    "Team Remarks",
    "Task Completed Date",
    "Task Status",
    "Aging Days of Task Pending/Completed",
]

SETTLEMENT_VISIBLE_COLUMNS = [
    "Order No","Replacement Order","Order Date/Shipment Date","Order Qty",
    "Order Price","Sale (Invoice No)","Sale Date (Invoice Date)","Sale Qty",
    "Sale Price","Return (CN No)","Return Date (CN Date)","Return Qty",
    "Return Price (CN Price)","Payment Received","Payment Received Date",
    "Deferred Amount","Refund","Refund Date","Total Deductions",
    "Transaction Status",
]

# Exact Excel headers from attached workbook (duplicate Date/Qty/Price labels).
ECOM_EXPORT_HEADERS = [
    "Branch Code",
    "Order Date/Shipment Date",
    "Order Item",
    "Order No",
    "Replacement Order Id",
    "Order Qty",
    "Repl Quantity",
    "Order Price",
    "Replacement Item Price",
    "Courier & Customer Return Qty",
    "Order Status",
    "Return Created Date",
    "Return Delivery Date",
    "Return Type",
    "Payment Status",
    "Rece Amount",
    "Rece Amount Date",
    "Deferred Amount",
    "Refund",
    "Refund Date",
    "Reimbursement",
    "Adjustment",
    "Sale Invoice No",
    "Sale Invoice Date",
    "Invoice Qty",
    "Return Invoice No",
    "Return Invoice Date",
    "Return Qty",
    "Sale Price",
    "Pending Remarks",
    "Pending Task Created Date",
    "MIR No",
    "MIR Date",
    "TEI No",
    "TEI Date",
    "TEI Qty",
    "Ticket Raised If Any",
    "Ticket Raised Date",
    "Team Remarks",
    "Task Completed Date",
    "Task Status",
    "Aging Days of Task Pending/Completed",
]

SETTLEMENT_EXPORT_HEADERS = [
    "Order No","Replacement Order","Order Date/Shipment Date","Order Qty",
    "Order Price","Sale (Invoice No)","Sale Date (Invoice Date)","Sale Qty",
    "Sale Price","Return (CN No)","Return Date (CN Date)","Return Qty",
    "Return Price (CN Price)","Payment Received","Payment Received Date",
    "Deferred Amount","Refund","Refund Date","Total Deductions",
    "Transaction Status",
]

def exact_dashboard_export(view, visible_columns, export_headers):
    if view.empty:
        return pd.DataFrame(columns=export_headers)
    out = pd.DataFrame(index=view.index)
    for col in visible_columns:
        out[col] = safe_series(view, col, "")
    out.columns = export_headers
    return out.reset_index(drop=True)

# ============================================================
# FILTER BAR — SAME ON ALL DASHBOARDS
# ============================================================

def apply_global_filters(df, prefix):
    if df.empty:
        return df

    c1,c2,c3,c4,c5 = st.columns(5)

    def options(col):
        if col not in df.columns:
            return []
        vals = []
        for x in df[col].fillna("").astype(str).unique():
            x = x.strip()
            if x and x.lower() != "nan":
                vals.append(x)
        return sorted(dict.fromkeys(vals))

    # v15.6: Pending Task filter must reflect only the actual Pending Remarks
    # currently present in the filtered dataset. Do not mix in Task Type values
    # or hard-coded/canonical statuses that are not present as Pending Remarks.
    task_options = options("Pending Remarks")

    mp = c1.multiselect("Market Place", options("Market Place"), key=f"{prefix}_mp")
    br = c2.multiselect("Branch", options("Branch Code"), key=f"{prefix}_branch")
    pt = c3.multiselect("Pending Task", task_options, key=f"{prefix}_task")
    ts = c4.multiselect("Task Status", ["Pending","Working","Completed"], key=f"{prefix}_status")
    oid = c5.text_input("Individual Order ID", key=f"{prefix}_oid")

    view = df.copy()

    if mp and "Market Place" in view.columns:
        view = view[view["Market Place"].astype(str).isin(mp)]
    if br and "Branch Code" in view.columns:
        view = view[view["Branch Code"].astype(str).isin(br)]
    if pt and "Pending Remarks" in view.columns:
        view = view[
            view["Pending Remarks"]
            .fillna("")
            .astype(str)
            .str.strip()
            .isin(pt)
        ]
    if ts and "Task Status" in view.columns:
        view = view[view["Task Status"].fillna("").astype(str).apply(
            lambda s: any(x == part.strip() for x in ts for part in s.split("|"))
        )]
    if oid and "Order No" in view.columns:
        view = view[view["Order No"].astype(str).str.contains(oid,case=False,na=False)]

    return view

# ============================================================
# EXCEL DOWNLOADS / TEAM UPDATE IMPORT
# ============================================================
def exact_attached_export(view):
    # Build 31 columns in user's exact attached order and header wording.
    data = [
        view.get("Order No",pd.Series(dtype=object)),
        view.get("Order Date/Shipment Date",pd.Series(dtype=object)),
        view.get("Order Item",pd.Series(dtype=object)),
        view.get("Order Qty",pd.Series(dtype=object)),
        view.get("Order Price",pd.Series(dtype=object)),
        view.get("Branch Code",pd.Series(dtype=object)),
        view.get("Order Status",pd.Series(dtype=object)),
        view.get("Order Return Date",pd.Series(dtype=object)),
        view.get("Return Completed Date",pd.Series(dtype=object)),
        view.get("Return Status",pd.Series(dtype=object)),
        view.get("Payment Status",pd.Series(dtype=object)),
        view.get("Refund",pd.Series(dtype=object)),
        view.get("Reimbursement",pd.Series(dtype=object)),
        view.get("Invoice",pd.Series(dtype=object)),
        view.get("Invoice Date",pd.Series(dtype=object)),
        view.get("Invoice Qty",pd.Series(dtype=object)),
        view.get("Invoice Price",pd.Series(dtype=object)),
        view.get("Credit Note",pd.Series(dtype=object)),
        view.get("CN Date",pd.Series(dtype=object)),
        view.get("CN Qty",pd.Series(dtype=object)),
        view.get("CN Price",pd.Series(dtype=object)),
        view.get("Pending Remarks",pd.Series(dtype=object)),
        view.get("Pending Task Created Date",pd.Series(dtype=object)),
        view.get("MIR No",pd.Series(dtype=object)),
        view.get("MIR Date",pd.Series(dtype=object)),
        view.get("TEI No",pd.Series(dtype=object)),
        view.get("TEI Date",pd.Series(dtype=object)),
        view.get("Team Remarks",pd.Series(dtype=object)),
        view.get("Task Completed Date",pd.Series(dtype=object)),
        view.get("Task Status",pd.Series(dtype=object)),
        view.get("Aging Days of Task Pending/Completed",pd.Series(dtype=object)),
    ]

    if view.empty:
        return pd.DataFrame(columns=ATTACHED_COLUMNS_EXPORT)

    matrix = pd.concat(data,axis=1)
    matrix.columns = ATTACHED_COLUMNS_EXPORT
    return matrix

def workbook_bytes(main_sheet, system_mapping=None, main_name="Reconciliation"):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio,engine="openpyxl") as writer:
        main_sheet.to_excel(writer,index=False,sheet_name=main_name[:31])
        if system_mapping is not None and not system_mapping.empty:
            system_mapping.to_excel(writer,index=False,sheet_name="System Mapping")
    return bio.getvalue()

def pending_download(view):
    """
    Main Pending Task Dashboard download.
    Uses the exact main reconciliation format and protects all source fields.
    Only the seven approved team-working columns are editable.
    """
    pending = view[
        view["Task Type"].fillna("").astype(str).str.strip() != ""
    ].copy()

    exact = exact_dashboard_export(
        pending,ECOM_VISIBLE_COLUMNS,ECOM_EXPORT_HEADERS
    )
    mapping = pending[[
        "Market Place","Branch Code","Order No","Task Type"
    ]].copy()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio,engine="openpyxl") as writer:
        exact.to_excel(writer,index=False,sheet_name="Pending Tasks")
        mapping.to_excel(writer,index=False,sheet_name="System Mapping")

    bio.seek(0)
    book = load_workbook(bio)
    ws = book["Pending Tasks"]
    book["System Mapping"].sheet_state = "veryHidden"

    header_map = {
        str(cell.value).strip(): cell.column
        for cell in ws[1] if cell.value is not None
    }

    editable_fill = PatternFill(fill_type="solid",fgColor="FFF2CC")
    editable_header_fill = PatternFill(fill_type="solid",fgColor="C65911")
    normal_header_fill = PatternFill(fill_type="solid",fgColor="1F4E78")
    white_bold = Font(color="FFFFFF",bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    for header in TEAM_EDITABLE_COLUMNS:
        col_idx = header_map.get(header)
        if not col_idx:
            continue
        ws.cell(1,col_idx).fill = editable_header_fill
        ws.cell(1,col_idx).font = white_bold
        for row_idx in range(2,ws.max_row+1):
            cell = ws.cell(row_idx,col_idx)
            cell.protection = Protection(locked=False)
            cell.fill = editable_fill

    for cell in ws[1]:
        if str(cell.value).strip() not in TEAM_EDITABLE_COLUMNS:
            cell.fill = normal_header_fill
            cell.font = white_bold

    for header in ["MIR Date","TEI Date","Ticket Raised Date","Task Completed Date"]:
        col_idx = header_map.get(header)
        if col_idx:
            for row_idx in range(2,ws.max_row+1):
                ws.cell(row_idx,col_idx).number_format = "dd-mmm-yyyy"

    tei_qty_col = header_map.get("TEI Qty")
    if tei_qty_col:
        for row_idx in range(2,ws.max_row+1):
            ws.cell(row_idx,tei_qty_col).number_format = "0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.protection.sheet = True
    ws.protection.password = "GlenBranchTeam"
    ws.protection.enable()

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()




TEAM_EDITABLE_COLUMNS = [
    "MIR No",
    "MIR Date",
    "TEI No",
    "TEI Date",
    "TEI Qty",
    "Ticket Raised If Any",
    "Ticket Raised Date",
    "Team Remarks",
    "Task Completed Date",
]

def protected_branch_pending_download(view, branch):
    """
    Download the exact E-Com reconciliation format.
    Only TEAM_EDITABLE_COLUMNS are unlocked for team editing.
    All reconciliation/source columns are locked.
    """
    pending = view[
        view["Task Type"].fillna("").astype(str).str.strip() != ""
    ].copy()

    exact = exact_dashboard_export(
        pending, ECOM_VISIBLE_COLUMNS, ECOM_EXPORT_HEADERS
    )

    mapping = pending[[
        "Market Place","Branch Code","Order No","Task Type"
    ]].copy()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        exact.to_excel(writer,index=False,sheet_name="Pending Tasks")
        mapping.to_excel(writer,index=False,sheet_name="System Mapping")

    bio.seek(0)
    book = load_workbook(bio)
    ws = book["Pending Tasks"]
    mapping_ws = book["System Mapping"]

    # Keep internal mapping inaccessible during normal team working.
    mapping_ws.sheet_state = "veryHidden"

    # Header map.
    header_map = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    editable_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC"
    )
    editable_header_fill = PatternFill(
        fill_type="solid",
        fgColor="C65911"
    )
    normal_header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )
    white_bold = Font(color="FFFFFF",bold=True)

    # Lock every cell first.
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # Unlock only the approved team columns.
    for header in TEAM_EDITABLE_COLUMNS:
        col_idx = header_map.get(header)
        if not col_idx:
            continue

        ws.cell(1,col_idx).fill = editable_header_fill
        ws.cell(1,col_idx).font = white_bold

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx,col_idx)
            cell.protection = Protection(locked=False)
            cell.fill = editable_fill

    # Style all other headers.
    for cell in ws[1]:
        if str(cell.value).strip() not in TEAM_EDITABLE_COLUMNS:
            cell.fill = normal_header_fill
            cell.font = white_bold

    # Date formatting for editable date columns.
    for header in ["MIR Date","TEI Date","Ticket Raised Date","Task Completed Date"]:
        col_idx = header_map.get(header)
        if col_idx:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx,col_idx).number_format = "dd-mmm-yyyy"

    tei_qty_col = header_map.get("TEI Qty")
    if tei_qty_col:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx,tei_qty_col).number_format = "0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Protect source/reconciliation cells. Users can edit only unlocked cells.
    ws.protection.sheet = True
    ws.protection.password = "GlenBranchTeam"
    ws.protection.enable()

    # Add a note on top-level properties without changing the attached format.
    ws.sheet_view.showGridLines = True

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def apply_branch_team_workbook(uploaded, branch):
    """
    Apply permitted team fields with batch database operations.

    v15.8 performance changes:
    - Excel is parsed once.
    - System Mapping is indexed once instead of filtered row-by-row.
    - Existing MIR/task/return-qty data is prefetched in bulk.
    - MIR updates, task updates and audit rows are written with executemany().
    - One database commit is used for the full workbook.

    This keeps all v15.7 quantity-level TEI protections intact.
    """
    if uploaded.name.lower().endswith(".csv"):
        raise ValueError(
            "Please upload the protected XLSX downloaded from this branch. "
            "CSV cannot preserve locked columns/System Mapping."
        )

    with pd.ExcelFile(uploaded) as xl:
        first = xl.sheet_names[0]
        df = normalize_uploaded_headers(pd.read_excel(xl, sheet_name=first))
        mapping = (
            pd.read_excel(xl, sheet_name="System Mapping")
            if "System Mapping" in xl.sheet_names
            else pd.DataFrame()
        )

    required = {"Order No", "Branch Code"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError("Team workbook missing: " + ", ".join(sorted(missing)))

    all_branches = str(branch).strip().upper() == "ALL"

    # Normalize only usable rows once.
    df = df.copy()
    df["_order_no"] = df["Order No"].map(clean_id)
    df["_row_branch"] = (
        df["Branch Code"].map(clean_id)
        if all_branches
        else clean_id(branch)
    )
    df = df[df["_order_no"].astype(str).str.strip().ne("")].copy()

    if not all_branches:
        wrong_branch = df[
            df["Branch Code"].map(clean_id).ne(clean_id(branch))
        ]
        if not wrong_branch.empty:
            raise ValueError(
                f"Uploaded workbook contains {len(wrong_branch)} row(s) outside branch {branch}. "
                "Please use only the branch download."
            )

    if df.empty:
        return 0, []

    now = datetime.now().isoformat(timespec="seconds")
    today_iso = date.today().isoformat()

    def changed_text(new_value, old_value):
        return bool(new_value) and txt(new_value) != txt(old_value)

    def changed_date(new_value, old_value):
        return bool(new_value) and iso_date(new_value) != iso_date(old_value)

    # Build exact mapping lookup once.
    mapping_index = {}
    if not mapping.empty and {"Order No", "Branch Code"}.issubset(mapping.columns):
        m = mapping.copy()
        m["_order_no"] = m["Order No"].map(clean_id)
        m["_branch"] = m["Branch Code"].map(clean_id)
        for _, mr in m.iterrows():
            key = (clean_id(mr.get("_branch")), clean_id(mr.get("_order_no")))
            mapping_index.setdefault(key, []).append((
                txt(mr.get("Market Place")),
                txt(mr.get("_order_no")),
                txt(mr.get("Task Type")),
            ))

    row_keys = list(dict.fromkeys(
        (clean_id(r["_row_branch"]), clean_id(r["_order_no"]))
        for _, r in df.iterrows()
        if txt(r["_row_branch"]) and txt(r["_order_no"])
    ))
    key_set = set(row_keys)

    # Prefetch all DB state needed for this workbook in a few queries.
    with db() as c:
        branch_values = sorted({k[0] for k in row_keys})
        order_values = sorted({k[1] for k in row_keys})

        if branch_values and order_values:
            bq = ",".join("?" for _ in branch_values)
            oq = ",".join("?" for _ in order_values)

            task_rows = c.execute(f"""
                SELECT portal,order_no,task_type,branch_code,
                       task_status,working_date,task_completed_date,
                       team_remarks,ticket_raised,raised_date
                FROM pending_tasks
                WHERE branch_code IN ({bq}) AND order_no IN ({oq})
            """, tuple(branch_values + order_values)).fetchall()

            mir_rows = c.execute(f"""
                SELECT branch_code,order_no,mir_no,mir_date,tei_no,tei_date,tei_qty
                FROM mir_details
                WHERE branch_code IN ({bq}) AND order_no IN ({oq})
            """, tuple(branch_values + order_values)).fetchall()

            rq_rows = c.execute(f"""
                SELECT branch_code,order_no,
                       COALESCE(MAX(ABS(courier_customer_return_qty)),0),
                       COALESCE(MAX(ABS(cn_qty)),0),
                       COALESCE(MAX(ABS(order_qty)),0)
                FROM reconciliation_master
                WHERE branch_code IN ({bq}) AND order_no IN ({oq})
                GROUP BY branch_code,order_no
            """, tuple(branch_values + order_values)).fetchall()
        else:
            task_rows, mir_rows, rq_rows = [], [], []

    tasks_by_key = {}
    for row in task_rows:
        portal = txt(row[0])
        order_no = clean_id(row[1])
        task_type = txt(row[2])
        branch_code = clean_id(row[3])
        if (branch_code, order_no) not in key_set:
            continue
        tasks_by_key.setdefault((branch_code, order_no), []).append({
            "portal": portal,
            "order_no": order_no,
            "task_type": task_type,
            "task_status": txt(row[4]),
            "working_date": row[5],
            "task_completed_date": row[6],
            "team_remarks": txt(row[7]),
            "ticket_raised": txt(row[8]),
            "raised_date": row[9],
        })

    mir_by_key = {}
    for row in mir_rows:
        key = (clean_id(row[0]), clean_id(row[1]))
        if key in key_set:
            mir_by_key[key] = {
                "mir_no": txt(row[2]),
                "mir_date": row[3],
                "tei_no": txt(row[4]),
                "tei_date": row[5],
                "tei_qty": float(row[6] or 0),
            }

    rq_by_key = {}
    for row in rq_rows:
        key = (clean_id(row[0]), clean_id(row[1]))
        rq_by_key[key] = max(
            float(row[2] or 0),
            float(row[3] or 0),
            1.0 if float(row[4] or 0) > 0 else 0.0,
        )

    mir_upserts = []
    task_updates = []
    audit_rows = []
    updated = 0
    unmatched = []

    for _, r in df.iterrows():
        order_no = clean_id(r.get("_order_no"))
        row_branch = clean_id(r.get("_row_branch"))
        if not row_branch:
            unmatched.append(f"{row_branch}:{order_no}" if row_branch else order_no)
            continue

        key = (row_branch, order_no)
        targets = mapping_index.get(key, [])
        if not targets:
            targets = [
                (t["portal"], order_no, t["task_type"])
                for t in tasks_by_key.get(key, [])
            ]
        if not targets:
            unmatched.append(f"{row_branch}:{order_no}")
            continue

        mir_no = txt(r.get("MIR No"))
        mir_date = iso_date(r.get("MIR Date"))
        tei_no = txt(r.get("TEI No"))
        tei_date = iso_date(r.get("TEI Date"))

        tei_qty_raw = r.get("TEI Qty")
        tei_qty = None
        if not pd.isna(tei_qty_raw) and txt(tei_qty_raw):
            tei_qty = num(tei_qty_raw)
            if tei_qty < 0:
                raise ValueError(f"TEI Qty cannot be negative for Order No {order_no}.")

        ticket = txt(r.get("Ticket Raised If Any"))
        ticket_raised_date = iso_date(r.get("Ticket Raised Date"))
        team_remarks = txt(r.get("Team Remarks"))
        completed_date = iso_date(r.get("Task Completed Date"))

        old_mir = mir_by_key.get(key, {})
        old_mir_no = txt(old_mir.get("mir_no"))
        old_mir_date = old_mir.get("mir_date")
        old_tei_no = txt(old_mir.get("tei_no"))
        old_tei_date = old_mir.get("tei_date")
        old_tei_qty = float(old_mir.get("tei_qty") or 0)

        effective_tei_qty = old_tei_qty
        if tei_qty is not None:
            effective_tei_qty = tei_qty
        elif tei_no and not old_tei_no:
            effective_tei_qty = 1.0

        mir_changed = any([
            changed_text(mir_no, old_mir_no),
            changed_date(mir_date, old_mir_date),
            changed_text(tei_no, old_tei_no),
            changed_date(tei_date, old_tei_date),
            abs(effective_tei_qty - old_tei_qty) > 0.000001,
        ])

        final_mir_no = mir_no or old_mir_no
        final_mir_date = mir_date or old_mir_date
        final_tei_no = tei_no or old_tei_no
        final_tei_date = tei_date or old_tei_date

        if mir_changed:
            mir_upserts.append((
                row_branch, order_no,
                final_mir_no, final_mir_date,
                final_tei_no, final_tei_date,
                effective_tei_qty,
                MIR_BRANCH_PEOPLE.get(row_branch, ""),
                now,
            ))
            # Update local state so duplicate rows in the same workbook see latest values.
            mir_by_key[key] = {
                "mir_no": final_mir_no,
                "mir_date": final_mir_date,
                "tei_no": final_tei_no,
                "tei_date": final_tei_date,
                "tei_qty": effective_tei_qty,
            }

        task_updates_this_row = 0
        available_tasks = {
            (t["portal"], t["task_type"]): t
            for t in tasks_by_key.get(key, [])
        }

        for portal, oid, task_type in targets:
            current = available_tasks.get((txt(portal), txt(task_type)))
            if not current:
                continue

            task_changed = any([
                changed_text(team_remarks, current["team_remarks"]),
                changed_text(ticket, current["ticket_raised"]),
                changed_date(ticket_raised_date, current["raised_date"]),
                changed_date(completed_date, current["task_completed_date"]),
            ])
            if not task_changed and not mir_changed:
                continue

            if txt(task_type) == "TEI Pending":
                required_return_qty = rq_by_key.get(key, 0.0)
                covered_qty = effective_tei_qty
                if final_tei_no and covered_qty <= 0:
                    covered_qty = 1.0
                tei_complete = (
                    bool(final_tei_no)
                    and required_return_qty > 0
                    and covered_qty + 0.000001 >= required_return_qty
                )
                new_status = "Completed" if tei_complete else "Working"
                completed_for_write = completed_date if tei_complete else None
            elif completed_date or current["task_status"] == "Completed":
                new_status = "Completed"
                completed_for_write = completed_date
            else:
                new_status = "Working"
                completed_for_write = completed_date

            task_updates.append((
                new_status,
                new_status, today_iso,
                new_status, completed_for_write, today_iso,
                team_remarks, team_remarks,
                ticket, ticket,
                ticket_raised_date, ticket_raised_date,
                now, txt(portal), oid, txt(task_type),
            ))

            audit_rows.append((
                now, "BRANCH TEAM UPDATE", txt(portal), row_branch,
                oid, txt(task_type), uploaded.name,
                (
                    f"MIR={mir_no}; MIR Date={mir_date}; "
                    f"TEI={tei_no}; TEI Date={tei_date}; TEI Qty={effective_tei_qty}; "
                    f"Ticket={ticket}; Ticket Date={ticket_raised_date}; "
                    f"Remarks={team_remarks}; Completed={completed_for_write}; "
                    f"Status={new_status}"
                ),
            ))

            # Maintain local state for duplicate rows within the same file.
            current["task_status"] = new_status
            current["team_remarks"] = team_remarks or current["team_remarks"]
            current["ticket_raised"] = ticket or current["ticket_raised"]
            current["raised_date"] = ticket_raised_date or current["raised_date"]
            if new_status == "Completed":
                current["task_completed_date"] = (
                    completed_for_write or current["task_completed_date"] or today_iso
                )
            task_updates_this_row += 1
            updated += 1

        if mir_changed and task_updates_this_row == 0:
            updated += 1

    # One transaction / commit for the entire uploaded workbook.
    with db() as c:
        if mir_upserts:
            c.executemany("""
                INSERT INTO mir_details(
                    branch_code,order_no,mir_no,mir_date,tei_no,tei_date,tei_qty,
                    responsible_persons,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?)
                ON CONFLICT(branch_code,order_no) DO UPDATE SET
                    mir_no=excluded.mir_no,
                    mir_date=excluded.mir_date,
                    tei_no=excluded.tei_no,
                    tei_date=excluded.tei_date,
                    tei_qty=excluded.tei_qty,
                    responsible_persons=excluded.responsible_persons,
                    updated_at=excluded.updated_at
            """, mir_upserts)

        if task_updates:
            c.executemany("""
                UPDATE pending_tasks
                SET
                    task_status=?,
                    working_date=CASE
                        WHEN ?='Working' THEN COALESCE(working_date,?)
                        ELSE working_date
                    END,
                    task_completed_date=CASE
                        WHEN ?='Completed' THEN COALESCE(?,task_completed_date,?)
                        ELSE task_completed_date
                    END,
                    team_remarks=CASE WHEN ?<>'' THEN ? ELSE team_remarks END,
                    ticket_raised=CASE WHEN ?<>'' THEN ? ELSE ticket_raised END,
                    raised_date=CASE WHEN ? IS NOT NULL THEN ? ELSE raised_date END,
                    last_update=?
                WHERE portal=? AND order_no=? AND task_type=?
            """, task_updates)

        if audit_rows:
            c.executemany("""
                INSERT INTO activity_audit(
                    activity_time,action_type,portal,branch_code,
                    order_no,task_type,filename,details
                ) VALUES(?,?,?,?,?,?,?,?)
            """, audit_rows)

        c.commit()

    invalidate_read_cache()
    
    # Return unique unmatched orders only.
    return updated, list(dict.fromkeys(x for x in unmatched if x))

def normalize_uploaded_headers(df):
    df = df.copy()
    aliases = {
        "Completion Date": "Task Completed Date",
        "Remarks": "Team Remarks",
        "Status": "Task Status",
        "Order ID": "Order No",
        "Branch": "Branch Code",
    }
    cleaned = []
    for c in df.columns:
        name = str(c).strip()
        # Backward compatibility with old misspelled templates.
        if name == "Net Pay Rece" + "vied":
            name = "Received"
        elif name == "Reimburs" + "ment":
            name = "Reimbursement"
        elif name == "Short Payment Rece" + "vied":
            name = "Short Payment Received"
        cleaned.append(aliases.get(name, name))
    df.columns = cleaned
    return df

def apply_team_update_workbook(uploaded):
    if uploaded.name.lower().endswith(".csv"):
        df = normalize_uploaded_headers(pd.read_csv(uploaded))
        mapping = pd.DataFrame()
    else:
        with pd.ExcelFile(uploaded) as xl:
            first = xl.sheet_names[0]
            df = normalize_uploaded_headers(pd.read_excel(xl,sheet_name=first))
            mapping = (
                pd.read_excel(xl,sheet_name="System Mapping")
                if "System Mapping" in xl.sheet_names
                else pd.DataFrame()
            )

    # Format A: user's attached reconciliation/pending format.
    if "Order No" in df.columns and "Task Status" in df.columns:
        updated = 0
        unmatched = []

        for _, r in df.iterrows():
            order_no = clean_id(r.get("Order No"))
            if not order_no:
                continue

            task_status = txt(r.get("Task Status"))
            team_remarks = txt(r.get("Team Remarks"))
            completed_date = r.get("Task Completed Date")
            pending_remark = txt(r.get("Pending Remarks"))
            task_type = task_type_from_remark(pending_remark)

            # If System Mapping is present, portal/task are exact.
            matched_map = pd.DataFrame()
            if not mapping.empty:
                matched_map = mapping[
                    mapping["Order No"].astype(str).map(clean_id) == order_no
                ]
                if task_type and "Task Type" in matched_map.columns:
                    matched_map = matched_map[
                        matched_map["Task Type"].astype(str) == task_type
                    ]

            targets = []

            if not matched_map.empty:
                for _, mr in matched_map.iterrows():
                    targets.append((
                        txt(mr.get("Market Place")),
                        order_no,
                        txt(mr.get("Task Type")) or task_type
                    ))
            else:
                branch = txt(r.get("Branch Code"))
                with db() as c:
                    if task_type:
                        rows = c.execute("""
                            SELECT portal,order_no,task_type
                            FROM pending_tasks
                            WHERE order_no=? AND task_type=?
                              AND (?='' OR branch_code=?)
                        """,(order_no,task_type,branch,branch)).fetchall()
                    else:
                        rows = c.execute("""
                            SELECT portal,order_no,task_type
                            FROM pending_tasks
                            WHERE order_no=? AND (?='' OR branch_code=?)
                        """,(order_no,branch,branch)).fetchall()
                targets = rows

            if not targets:
                unmatched.append(order_no)
                continue

            for portal, oid, tt in targets:
                if not task_status:
                    continue
                update_task(
                    portal,oid,tt,task_status,
                    team_remarks=team_remarks,
                    completed_date=completed_date
                )
                updated += 1

        return updated, unmatched

    # Format B: compact legacy team template.
    required = {"Branch","Order ID","Task Type","Status"}
    if required.issubset(set(df.columns)):
        updated = 0
        unmatched = []

        for _, r in df.iterrows():
            branch = txt(r.get("Branch"))
            order_no = clean_id(r.get("Order ID"))
            task_type = txt(r.get("Task Type"))
            status = txt(r.get("Status"))

            with db() as c:
                targets = c.execute("""
                    SELECT portal,order_no,task_type
                    FROM pending_tasks
                    WHERE branch_code=? AND order_no=? AND task_type=?
                """,(branch,order_no,task_type)).fetchall()

            if not targets:
                unmatched.append(order_no)
                continue

            for portal,oid,tt in targets:
                update_task(
                    portal,oid,tt,status,
                    team_remarks=txt(r.get("Remarks")),
                    working_date=r.get("Working Date"),
                    completed_date=r.get("Completion Date")
                )
                updated += 1

        return updated, unmatched

    raise ValueError(
        "Unsupported Team Update format. Use the downloaded Pending Task workbook "
        "or the compact Branch/Order ID/Task Type/Status format."
    )

# ============================================================
# MICROSOFT 365 EMAIL (OPTIONAL, PERSISTENT CONFIG)
# ============================================================
GRAPH_SCOPES = ["Mail.Send","User.Read","offline_access"]

def load_m365_config():
    # Cloud-first so configuration survives Streamlit restart/redeploy.
    try:
        with db() as c:
            row = c.execute("""
                SELECT setting_value
                FROM app_settings
                WHERE setting_key=?
            """,("m365_config",)).fetchone()
        if row and txt(row[0]):
            return json.loads(row[0])
    except Exception:
        pass

    if not M365_CONFIG.exists():
        return {}
    try:
        return json.loads(M365_CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_m365_config(tenant_id,client_id,sender_email,auto_send):
    cfg = {
        "tenant_id":txt(tenant_id),
        "client_id":txt(client_id),
        "sender_email":txt(sender_email),
        "auto_send":bool(auto_send),
    }
    payload = json.dumps(cfg)
    now = datetime.now().isoformat(timespec="seconds")
    with db() as c:
        c.execute("""
            INSERT INTO app_settings(setting_key,setting_value,updated_at)
            VALUES(?,?,?)
            ON CONFLICT(setting_key) DO UPDATE SET
                setting_value=excluded.setting_value,
                updated_at=excluded.updated_at
        """,("m365_config",payload,now))
        c.commit()

    try:
        M365_CONFIG.write_text(json.dumps(cfg,indent=2),encoding="utf-8")
    except Exception:
        pass
    return cfg

def token_cache():
    cache = msal.SerializableTokenCache()
    if M365_TOKEN_CACHE.exists():
        try:
            cache.deserialize(M365_TOKEN_CACHE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return cache

def persist_cache(cache):
    if cache.has_state_changed:
        M365_TOKEN_CACHE.write_text(cache.serialize(),encoding="utf-8")

def msal_app(cfg, cache):
    return msal.PublicClientApplication(
        cfg["client_id"],
        authority=f"https://login.microsoftonline.com/{cfg['tenant_id']}",
        token_cache=cache
    )

def begin_m365_login():
    cfg = load_m365_config()
    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        raise ValueError("Save Tenant ID and Client ID first.")
    cache = token_cache()
    app = msal_app(cfg,cache)
    flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(str(flow))
    return flow

def complete_m365_login(flow):
    cfg = load_m365_config()
    cache = token_cache()
    app = msal_app(cfg,cache)
    result = app.acquire_token_by_device_flow(flow)
    persist_cache(cache)
    if "access_token" not in result:
        raise RuntimeError(result.get("error_description","Microsoft sign-in failed."))
    return result

def graph_token():
    cfg = load_m365_config()
    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        return None
    cache = token_cache()
    app = msal_app(cfg,cache)
    accounts = app.get_accounts()
    if not accounts:
        return None
    result = app.acquire_token_silent(GRAPH_SCOPES,account=accounts[0])
    persist_cache(cache)
    return result.get("access_token") if result else None

def send_graph_mail(token,to_email,subject,body):
    r = requests.post(
        "https://graph.microsoft.com/v1.0/me/sendMail",
        headers={"Authorization":f"Bearer {token}","Content-Type":"application/json"},
        json={
            "message":{
                "subject":subject,
                "body":{"contentType":"Text","content":body},
                "toRecipients":[{"emailAddress":{"address":to_email}}],
            },
            "saveToSentItems":True,
        },
        timeout=30
    )
    if r.status_code not in (200,202):
        raise RuntimeError(f"Microsoft Graph {r.status_code}: {r.text[:300]}")

def send_pending_emails():
    token = graph_token()
    if not token:
        raise RuntimeError("Microsoft 365 sign-in required.")

    tasks = load_tasks()
    if tasks.empty:
        return 0

    tasks = tasks[
        tasks["task_status"].isin(["Pending","Working"])
        & tasks["task_assign_email"].fillna("").astype(str).str.strip().ne("")
    ]

    expanded = []
    for _, r in tasks.iterrows():
        emails = [x.strip() for x in r["task_assign_email"].split(";") if x.strip()]
        names = [x.strip() for x in r["task_assign_to"].split(";") if x.strip()]
        for i,email in enumerate(emails):
            rec = r.to_dict()
            rec["recipient_email"] = email
            rec["recipient_name"] = names[i] if i < len(names) else ""
            expanded.append(rec)

    if not expanded:
        return 0

    x = pd.DataFrame(expanded)
    count = 0

    for email,g in x.groupby("recipient_email"):
        owner = txt(g["recipient_name"].iloc[0]) or "Team"
        lines = [
            f"Dear {owner},","",
            "The following e-commerce reconciliation tasks require action:",""
        ]
        for _, r in g.iterrows():
            lines.append(
                f"- {r['portal']} | Branch {r['branch_code']} | Order {r['order_no']} | "
                f"{r['task_type']} | {r['task_status']} | Ageing {int(r['aging_days'])} day(s)"
            )
        lines += ["","Please update the task through the Reconciliation Tower."]
        send_graph_mail(
            token,email,
            f"E-Commerce Pending Tasks - {len(g)} item(s)",
            "\n".join(lines)
        )
        count += 1

    return count


@st.cache_resource(show_spinner=False)
def repair_payment_status_consistency():
    """Repair legacy rows created by the old default-Received logic.

    This is intentionally idempotent and runs once per app process. It fixes
    Payment Status for every saved order from the actual received amount. For
    delivered orders that were incorrectly marked Reconciled despite having
    zero payment and no refund/reimbursement, it also restores the correct
    pending remark/transaction status so dashboards, tasks and exports agree.
    """
    updates = []
    with db() as c:
        rows = c.execute("""
            SELECT portal,order_no,order_status,net_pay_received,invoice_qty,
                   refund,reimbursement,pending_remarks,payment_status,
                   transaction_status
            FROM reconciliation_master
        """).fetchall()

        for (portal, order_no, order_status, received, invoice_qty, refund,
             reimbursement, remark, old_payment_status, old_transaction_status) in rows:
            received_num = num(received)
            remark_txt = txt(remark)
            new_payment_status = payment_status_from_amount(received_num, remark_txt)
            new_remark = remark_txt
            new_transaction_status = txt(old_transaction_status)

            status_l = txt(order_status).lower()
            is_delivered = "delivered" in status_l and not any(
                word in status_l for word in ("return", "cancel")
            )
            no_financial_closure = (
                received_num <= 0.000001
                and abs(num(refund)) <= 0.000001
                and abs(num(reimbursement)) <= 0.000001
            )

            # Legacy builds could leave an open delivered order as Reconciled
            # solely because payment_status defaulted to Received or because
            # an age-based fallback fired. Repair only that narrow unsafe case.
            if (
                is_delivered
                and no_financial_closure
                and remark_txt.lower().startswith("reconciled")
            ):
                if num(invoice_qty) > 0.000001:
                    new_remark = "Payment Pending"
                else:
                    new_remark = "Billing & Payment Pending"
                new_transaction_status = new_remark

            if not new_transaction_status:
                new_transaction_status = settlement_status_from_remark(
                    new_remark, new_payment_status
                )

            if (
                new_payment_status != txt(old_payment_status)
                or new_remark != remark_txt
                or new_transaction_status != txt(old_transaction_status)
            ):
                updates.append((
                    new_payment_status, new_remark, new_transaction_status,
                    txt(portal), clean_id(order_no)
                ))

        if updates:
            c.executemany("""
                UPDATE reconciliation_master
                SET payment_status=?, pending_remarks=?, transaction_status=?
                WHERE portal=? AND order_no=?
            """, updates)
            c.commit()

    if updates:
        invalidate_read_cache()
    return len(updates)


# Repair persisted legacy rows before creating any missing pending tasks.


@st.cache_resource(show_spinner=False)
def repair_returned_unbilled_no_payment():
    """v15.11 repair for already-saved orders.

    When return quantity exists and Return Delivery Date is available, but there
    is no sale invoice/billed quantity and no payment/refund, set Pending Remarks
    and Transaction Status to Reconciled. Existing active tasks are completed so
    they no longer appear in the default Pending Task dashboard.
    """
    today = date.today().isoformat()
    now = datetime.now().isoformat(timespec="seconds")
    affected = []

    with db() as c:
        rows = c.execute("""
            SELECT portal,order_no,courier_customer_return_qty,
                   return_completed_date,invoice_no,invoice_qty,
                   net_pay_received,refund,pending_remarks,transaction_status
            FROM reconciliation_master
        """).fetchall()

        for (
            portal, order_no, return_qty, return_date, invoice_no, invoice_qty,
            received, refund, remark, transaction_status
        ) in rows:
            physically_returned = abs(num(return_qty)) > 0.000001
            return_received = pd.notna(
                pd.to_datetime(return_date, errors="coerce")
            )
            no_billing = (
                abs(num(invoice_qty)) <= 0.000001
                and not txt(invoice_no)
            )
            no_payment = (
                abs(num(received)) <= 0.000001
                and abs(num(refund)) <= 0.000001
            )

            if physically_returned and return_received and no_billing and no_payment:
                key = (txt(portal), clean_id(order_no))
                affected.append(key)

        if affected:
            c.executemany("""
                UPDATE reconciliation_master
                SET pending_remarks='Reconciled',
                    transaction_status='Reconciled'
                WHERE portal=? AND order_no=?
            """, affected)

            c.executemany("""
                UPDATE pending_tasks
                SET task_status='Completed',
                    task_completed_date=COALESCE(task_completed_date,?),
                    team_remarks=CASE
                        WHEN COALESCE(TRIM(team_remarks),'')=''
                        THEN 'Auto-closed: return received; no billing/payment was created'
                        ELSE team_remarks
                    END,
                    last_update=?
                WHERE portal=? AND order_no=?
                  AND COALESCE(task_status,'')<>'Completed'
            """, [
                (today, now, portal, order_no)
                for portal, order_no in affected
            ])
            c.commit()

    if affected:
        invalidate_read_cache()
    return len(affected)




@st.cache_resource(show_spinner=False)
def complete_tasks_for_reconciled_orders():
    """Globally complete active tasks whenever Pending Remarks is Reconciled."""
    today = date.today().isoformat()
    now = datetime.now().isoformat(timespec="seconds")

    with db() as c:
        rows = c.execute("""
            SELECT DISTINCT portal,order_no
            FROM reconciliation_master
            WHERE LOWER(TRIM(COALESCE(pending_remarks,'')))='reconciled'
        """).fetchall()

        keys = [
            (txt(portal), clean_id(order_no))
            for portal, order_no in rows
            if txt(portal) and clean_id(order_no)
        ]
        if not keys:
            return 0

        active_before = c.execute("""
            SELECT COUNT(*)
            FROM pending_tasks pt
            WHERE COALESCE(pt.task_status,'')<>'Completed'
              AND EXISTS (
                  SELECT 1
                  FROM reconciliation_master rm
                  WHERE rm.portal=pt.portal
                    AND rm.order_no=pt.order_no
                    AND LOWER(TRIM(COALESCE(rm.pending_remarks,'')))='reconciled'
              )
        """).fetchone()[0]

        c.executemany("""
            UPDATE pending_tasks
            SET task_status='Completed',
                task_completed_date=COALESCE(task_completed_date,?),
                team_remarks=CASE
                    WHEN COALESCE(TRIM(team_remarks),'')=''
                    THEN 'Auto-completed because Pending Remarks changed to Reconciled'
                    ELSE team_remarks
                END,
                last_update=?
            WHERE portal=? AND order_no=?
              AND COALESCE(task_status,'')<>'Completed'
        """, [
            (today, now, portal, order_no)
            for portal, order_no in keys
        ])
        c.commit()

    if active_before:
        invalidate_read_cache()
    return int(active_before or 0)


# Do not mutate task workflow merely because the app was opened/restarted.
# Completion/backfill maintenance runs only after an explicit source/team update.

@st.cache_resource(show_spinner=False)
def backfill_missing_tasks_from_master():
    """
    One-time-per-process repair of missing task rows.

    Previous code performed SELECT-per-order against Supabase on every Streamlit
    rerun. That made even an Admin PIN click rerun thousands of remote queries.
    This version loads master/task keys once, computes missing tasks in memory,
    and inserts them in a single batched operation.
    """
    now = datetime.now().isoformat(timespec="seconds")

    with db() as c:
        master_rows = c.execute("""
            SELECT portal,order_no,branch_code,pending_remarks,source_uploaded_at
            FROM reconciliation_master
        """).fetchall()

        existing_rows = c.execute("""
            SELECT portal,order_no,task_type
            FROM pending_tasks
        """).fetchall()

        existing = {
            (txt(portal), txt(order_no), txt(task_type))
            for portal, order_no, task_type in existing_rows
        }

        inserts = []
        for portal, order_no, branch, remark, uploaded_at in master_rows:
            portal = txt(portal)
            order_no = txt(order_no)
            task_type = task_type_from_remark(remark)

            if not portal or not order_no or not task_type:
                continue

            key = (portal, order_no, task_type)
            if key in existing:
                continue

            created = (
                str(uploaded_at)[:10]
                if uploaded_at and len(str(uploaded_at)) >= 10
                else date.today().isoformat()
            )

            inserts.append((
                portal, order_no, task_type, txt(branch),
                created, "Pending", "", now
            ))
            existing.add(key)

        if inserts:
            c.executemany("""
                INSERT INTO pending_tasks(
                    portal,order_no,task_type,branch_code,
                    task_created_date,task_status,team_remarks,last_update
                ) VALUES(?,?,?,?,?,?,?,?)
                ON CONFLICT(portal,order_no,task_type) DO NOTHING
            """, inserts)

        c.commit()

    return len(inserts)


# ============================================================
# DEFERRED DATABASE MAINTENANCE
# ============================================================
def run_deferred_maintenance():
    """
    v15.14: Run heavier data-repair/backfill operations only when explicitly
    needed after source/team updates. They must not block normal app startup.
    """
    repaired_payment = repair_payment_status_consistency()
    repaired_returns = repair_returned_unbilled_no_payment()
    completed = complete_tasks_for_reconciled_orders()
    backfilled = backfill_missing_tasks_from_master()
    invalidate_read_cache()
    return {
        "payment_repairs": int(repaired_payment or 0),
        "return_repairs": int(repaired_returns or 0),
        "tasks_completed": int(completed or 0),
        "tasks_backfilled": int(backfilled or 0),
    }


# ============================================================
# FAST COMPLETION FEEDBACK
# ============================================================
def queue_completion_notice(message, kind="success"):
    st.session_state["_completion_notice"] = {
        "message": txt(message),
        "kind": txt(kind) or "success",
    }

def show_completion_notice():
    notice = st.session_state.pop("_completion_notice", None)
    if not notice:
        return
    message = txt(notice.get("message"))
    kind = txt(notice.get("kind")).lower()
    try:
        st.toast(message, icon="✅" if kind == "success" else "ℹ️")
    except Exception:
        pass
    if kind == "success":
        st.success(message)
    elif kind == "warning":
        st.warning(message)
    else:
        st.info(message)

# ============================================================
# MANAGEMENT MIS / ANALYTICS
# ============================================================
def _normalized_branch(series):
    s = series.fillna("").astype(str).str.strip()
    return s.where(s.ne(""), "Unassigned")


def _exception_category(remark):
    r = txt(remark).lower()
    if not r or r.startswith("reconciled"):
        return "Reconciled / No Exception"
    mapping = [
        ("partial tei", "Partial TEI / Balance Pending"),
        ("tei", "TEI Pending"),
        ("cn", "CN Pending"),
        ("short payment", "Short Payment"),
        ("payment", "Payment Pending"),
        ("billing", "Billing Pending"),
        ("refund", "Refund Review"),
        ("replacement", "Replacement Review"),
        ("mir", "MIR Pending"),
    ]
    for key, label in mapping:
        if key in r:
            return label
    return "Other Review"


def management_mis_frames(master, tasks):
    master = master.copy()
    tasks = tasks.copy()
    if not master.empty:
        master["__branch__"] = _normalized_branch(master.get("branch_code", pd.Series("", index=master.index)))
        master["__portal__"] = master.get("portal", pd.Series("", index=master.index)).fillna("").astype(str).str.strip()
        remark_s = master.get("pending_remarks", pd.Series("", index=master.index)).fillna("").astype(str).str.strip()
        ex_map = {v: _exception_category(v) for v in remark_s.drop_duplicates().tolist()}
        master["__exception__"] = remark_s.map(ex_map).fillna("Other Review")
        master["__reconciled__"] = (remark_s.str.lower().str.startswith("reconciled") | master.get("transaction_status", pd.Series("", index=master.index)).fillna("").astype(str).str.strip().str.lower().eq("reconciled"))
        master["__has_return__"] = pd.to_numeric(master.get("courier_customer_return_qty", 0), errors="coerce").fillna(0).abs().gt(0)
    if not tasks.empty:
        tasks["__branch__"] = _normalized_branch(tasks.get("branch_code", pd.Series("", index=tasks.index)))
        tasks["__portal__"] = tasks.get("portal", pd.Series("", index=tasks.index)).fillna("").astype(str).str.strip()
        tasks["__status__"] = tasks.get("task_status", pd.Series("", index=tasks.index)).fillna("").astype(str).str.strip()
        tasks["__aging__"] = pd.to_numeric(tasks.get("aging_days", 0), errors="coerce").fillna(0).clip(lower=0)
    return master, tasks


def _mis_filter_frames(master, tasks, portals=None, branches=None):
    portals = set(portals or [])
    branches = set(branches or [])
    m, t = master, tasks
    if portals:
        if not m.empty: m = m[m["__portal__"].isin(portals)]
        if not t.empty: t = t[t["__portal__"].isin(portals)]
    if branches:
        if not m.empty: m = m[m["__branch__"].isin(branches)]
        if not t.empty: t = t[t["__branch__"].isin(branches)]
    return m.copy(), t.copy()

# ============================================================
# UI
# ============================================================
st.title("E-Commerce Reconciliation Control Tower")
st.caption("Build: v15.18 — Management MIS + Branch/SLA/Exception Analytics")
show_completion_notice()

# v15.14 lightweight startup diagnostic. This performs only COUNT queries and
# does not scan/repair the full database.
try:
    with db() as c:
        _reco_count = c.execute("SELECT COUNT(*) FROM reconciliation_master").fetchone()[0]
        _task_count = c.execute("SELECT COUNT(*) FROM pending_tasks").fetchone()[0]
    st.caption(
        f"Database loaded: {int(_reco_count or 0):,} reconciliation rows | "
        f"{int(_task_count or 0):,} task rows"
    )
except Exception as _diag_exc:
    st.warning(f"Database diagnostic unavailable: {_diag_exc}")

st.caption(
    "Persistent multi-portal reconciliation. Amazon and Flipkart can be uploaded together "
    "or one by one. The latest source workbook, reconciliation, task and MIR updates are stored in the persistent cloud database and remain "
    "available after refresh, closing the browser, Streamlit restart or Windows shutdown."
)


with st.sidebar:
    st.markdown("### Admin Access")
    sec = load_security()
    if not sec.get("admin_pin_hash"):
        st.info("First-time setup: create an Admin PIN. Team users do not need this PIN to view dashboards or upload assigned task/MIR working sheets.")
        new_pin = st.text_input("Create Admin PIN", type="password", key="create_admin_pin")
        confirm_pin = st.text_input("Confirm Admin PIN", type="password", key="confirm_admin_pin")
        if st.button("Save Admin PIN", use_container_width=True):
            if new_pin != confirm_pin:
                st.error("PIN confirmation does not match.")
            else:
                try:
                    save_admin_pin(new_pin)
                    st.session_state["admin_unlocked"] = True
                    st.success("Admin PIN saved.")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))
    elif not is_admin_unlocked():
        pin = st.text_input("Admin PIN", type="password", key="admin_login_pin")
        if st.button("Unlock Admin", use_container_width=True):
            if pin_hash(pin) == sec.get("admin_pin_hash"):
                st.session_state["admin_unlocked"] = True
                st.success("Admin controls unlocked.")
                st.rerun()
            else:
                st.error("Incorrect Admin PIN.")
    else:
        st.success("Admin controls unlocked")
        if st.button("Lock Admin", use_container_width=True):
            st.session_state["admin_unlocked"] = False
            st.rerun()

workspace = st.radio(
    "Workspace",
    [
        "E-Com Reconciliation Dashboard",
        "Management MIS",
        "Settlement Dashboard",
        "Pending Task Dashboard",
        "All Branch Pending Task Update",
        "MIR 1600",
        "MIR 5400",
        "MIR 5600",
        "MIR 7800",
        "Branch Owner & Email Setup",
        "Upload History",
        "Activity Audit Trail",
    ],
    horizontal=True
)

# ------------------------------------------------------------
# RECONCILIATION DASHBOARD
# ------------------------------------------------------------
if workspace == "E-Com Reconciliation Dashboard":
    # Fast UI mode: source-derived repair work is NOT rerun on every widget
    # interaction. It already runs after a successful source upload, while
    # manual repair buttons remain available for exceptional cases.
    adjustment_autofix = None
    return_type_autofix = None

    if is_admin_unlocked():
        st.subheader("Upload Portal Source Data")
        st.caption(
            "New source uploads refresh only source-derived reconciliation fields. "
            "Existing team updates, MIR/TEI details, tickets, task status, completion dates "
            "and audit history are preserved automatically and reattached by Order ID/task."
        )
        st.caption(
            "Mapped source fields: "
            "Amazon Replacement_Map → Replacement Order Id; "
            "Amazon Reverse → Return Delivery Date; "
            "Amazon engine → Payment Received; "
            "Amazon Sales → Sale/CN Invoice No + Date; "
            "Flipkart Returns → Return Approval Date; "
            "Flipkart engine → Payment Received; "
            "Flipkart ERP Sales Register → Sale/CN Invoice No + Date."
        )

        st.caption(
            "The latest verified Amazon/Flipkart source workbook is stored persistently. "
            "Code deployments and Streamlit restarts do not require the same source to be uploaded again."
        )
        st.write(
            "Upload one or more complete portal workbooks. You may upload Amazon and Flipkart "
            "together, or upload either portal separately on different days."
        )

        files = st.file_uploader(
            "Upload Amazon / Flipkart source workbook(s)",
            type=["xlsx","xlsm"],
            accept_multiple_files=True,
            key="portal_files"
        )

        if files and st.button("Run & Save Reconciliation",type="primary",use_container_width=True):
            upload_batch_started = perf_counter()
            processed = []
            errors = []

            for uploaded in files:
                try:
                    with tempfile.TemporaryDirectory() as tmp:
                        path = Path(tmp) / uploaded.name
                        path.write_bytes(uploaded.getbuffer())
                        portal = detect_portal(path)
                        if not portal:
                            raise ValueError(
                                "Portal could not be auto-detected from workbook sheets."
                            )

                        with st.status(
                            f"Processing {portal}: {uploaded.name}",
                            expanded=True
                        ) as process_status:
                            normalized = process_workbook(
                                path, portal, progress=process_status
                            )
                            process_status.write(
                                f"4/4 Saving {len(normalized):,} reconciled "
                                f"{portal} orders to persistent database…"
                            )

                        integrity = reconciliation_integrity(normalized, portal)
                        expected_dashboard = normalized_dashboard_totals(normalized)

                        coverage = source_field_coverage(normalized)

                        if coverage["payment_rows"] == 0:
                            raise ValueError(
                                f"{portal} Payment Received mapping returned zero rows. "
                                "The workbook was NOT saved. Please verify the source payment sheet."
                            )

                        n = upsert_reconciliation(normalized,uploaded.name)

                        adjustment_sync = None
                        return_type_sync = None

                        # AUTO-SYNC: no manual Repair/Sync button is required after upload.
                        # Every new Amazon workbook immediately refreshes the derived fields
                        # from the same newly uploaded source before the dashboard is shown.
                        flipkart_reimbursement_sync = None

                        if portal == "Amazon":
                            return_type_sync = backfill_amazon_return_type_from_source(path)
                            adjustment_sync = backfill_amazon_adjustment_from_source(path)
                        elif portal == "Flipkart":
                            flipkart_reimbursement_sync = (
                                backfill_flipkart_reimbursement_from_source(path)
                            )

                        # All source-derived mapping is complete at this point.
                        # Explicitly release cached workbook objects before copying the
                        # temporary source on Windows. This prevents WinError 32 file locks.
                        clear_source_cache(path)

                        # Persist the exact uploaded workbook itself in Supabase.
                        # It remains the active source until the next workbook for the
                        # same marketplace is uploaded.
                        save_source_workbook(
                            portal,
                            uploaded.name,
                            uploaded.getvalue()
                        )

                        # Keep a local working copy too (useful on Windows); the cloud
                        # copy above is the permanent source of truth.
                        local_source = SOURCE_DIR / f"{portal.lower()}_latest.xlsx"
                        shutil.copy2(path, local_source)

                        # Verify only after all derived fields have been persisted.
                        db_verified = verify_portal_database_sync(
                            portal, expected_dashboard
                        )
                        verification_warnings = db_verified.get(
                            "_verification_warnings", []
                        )
                        if verification_warnings:
                            st.warning(
                                f"{portal} saved successfully. Aggregate verification "
                                "found non-blocking differences after source backfill: "
                                + " | ".join(verification_warnings)
                            )

                        try:
                            process_status.update(
                                label=f"{portal} reconciliation completed",
                                state="complete",
                                expanded=False
                            )
                        except Exception:
                            pass

                        if portal == "Amazon":
                            st.success(
                                "Amazon uploaded and auto-synced successfully — "
                                f"Return Type updated for "
                                f"{int((return_type_sync or {}).get('updated_orders',0)):,} orders. "
                                "Adjustment has also been refreshed from Payments. "
                                "No Repair/Sync button click is required."
                            )

                        if portal == "Flipkart":
                            st.success(
                                "Flipkart uploaded and auto-synced successfully — "
                                f"Reimbursement ₹"
                                f"{float((flipkart_reimbursement_sync or {}).get('matched_total',0)):,.2f} "
                                f"mapped across "
                                f"{int((flipkart_reimbursement_sync or {}).get('matched_orders',0)):,} orders "
                                "from Flipkart Payments Column P. "
                                "No manual sync is required."
                            )

                        history_after = operational_history_counts(portal)

                        st.info(
                            f"Preserved team history — "
                            f"Tasks {history_after['tasks']:,} | "
                            f"Working {history_after['working']:,} | "
                            f"Completed {history_after['completed']:,} | "
                            f"Tickets {history_after['tickets']:,}. "
                            "MIR/TEI records and audit history are also retained."
                        )

                        log_activity(
                            "SOURCE UPLOAD",
                            portal=portal,
                            filename=uploaded.name,
                            details=(
                                f"{n} rows | Payment={coverage['payment_rows']} | "
                                f"Replacement={coverage['replacement_rows']} | "
                                f"ReplQty={coverage['replacement_qty_rows']} | "
                                f"ReplValue={coverage['replacement_value_rows']} | "
                                f"CourierReturn={coverage['courier_return_rows']} | "
                                f"Invoice={coverage['invoice_rows']} | "
                                f"CN={coverage['cn_rows']} | "
                                f"ReturnDate={coverage['return_date_rows']} | "
                                f"Refund={coverage['refund_rows']} | "
                                f"Reimbursement={coverage['reimbursement_rows']}"
                            )
                        )
                        processed.append(
                            f"{portal}: {n} rows | "
                            f"DB Replacement={db_verified['replacement_orders']} | "
                            f"DB Repl Qty={db_verified['repl_quantity']:,.0f} | "
                            f"DB Courier Return={db_verified['courier_returns']:,.0f} | "
                            f"DB Rece=₹{db_verified['rece_amount']:,.2f} | "
                            f"DB Deferred=₹{db_verified['deferred_amount']:,.2f} | "
                            f"DB Adjustment=₹{db_verified['adjustment']:,.2f}"
                        )

                        # Release the cached 15–30 MB source workbook only after
                        # all portal-derived mapping/backfill work has finished.
                        clear_source_cache(path)

                except Exception as exc:
                    # Ensure temporary uploaded Excel files are never left open in
                    # the in-memory source cache, even when reconciliation fails.
                    try:
                        if "path" in locals():
                            clear_source_cache(path)
                    except Exception:
                        pass
                    errors.append(f"{uploaded.name}: {type(exc).__name__}: {exc}")

            if processed:
                st.success("Saved permanently — " + " | ".join(processed))

                st.success(
                    "Latest source snapshot saved permanently. "
                    "Dashboard now reads the persisted reconciliation together "
                    "with the preserved team/MIR/ticket history."
                )

                cfg = load_m365_config()
                if cfg.get("auto_send"):
                    try:
                        sent = send_pending_emails()
                        st.info(f"Automatic pending-task email sent to {sent} owner(s).")
                    except Exception as exc:
                        st.warning(f"Automatic email skipped: {exc}")

            if errors:
                st.error("\n".join(errors))

            # v15.14: heavy repair/backfill runs only after a successful source
            # update, never during normal app startup.
            if processed and not errors:
                maintenance = run_deferred_maintenance()
                elapsed = perf_counter() - upload_batch_started
                queue_completion_notice(
                    f"Upload complete in {elapsed:,.1f}s. Reconciliation saved successfully and dashboards refreshed. "
                    f"Maintenance: {maintenance['tasks_completed']} task(s) auto-completed, "
                    f"{maintenance['tasks_backfilled']} task(s) backfilled."
                )
                st.rerun()


    else:
        st.info("Dashboard is read-only for team users. Admin PIN is required only for Amazon/Flipkart source uploads.")

    st.divider()
    st.subheader("E-Com Reconciliation Dashboard")
    st.caption(
        "Format follows the attached E-Com Reconciliation sheet. "
        "Summary values recalculate from the current filters."
    )

    try:
        master = load_master()
    except psycopg2.OperationalError:
        # Keep database credentials/server details out of the browser while
        # giving the user a clean recovery path for a transient DB outage.
        st.error(
            "The cloud database is temporarily unavailable, so reconciliation "
            "data could not be loaded. Please retry in a few seconds. If this "
            "continues, check DATABASE_URL / Supabase availability in the app settings."
        )
        if st.button("Retry database connection", type="primary"):
            invalidate_read_cache()
            st.rerun()
        st.stop()
    except Exception:
        st.error(
            "Saved reconciliation data could not be loaded. Please retry. "
            "If the problem continues, check the Streamlit app logs."
        )
        if st.button("Retry loading reconciliation", type="primary"):
            invalidate_read_cache()
            st.rerun()
        st.stop()

    with st.spinner("Loading saved reconciliation and team/MIR working…"):
        display = ecom_process_display(master)

    if display.empty:
        st.info(
            "No reconciliation is stored yet. Upload Amazon / Flipkart source data above."
        )
        st.stop()

    stored_amazon = portal_database_totals("Amazon")
    if stored_amazon["rows"] > 0 and (
        stored_amazon["replacement_orders"] == 0
        or abs(stored_amazon["rece_amount"]) < 0.01
    ):
        st.error(
            "Stored Amazon reconciliation is from an older/incomplete mapping: "
            f"{stored_amazon['rows']:,} orders are present but "
            f"Replacement Orders={stored_amazon['replacement_orders']} and "
            f"Rece Amount=₹{stored_amazon['rece_amount']:,.2f}. "
            "Admin must upload the current Amazon workbook once in this version. "
            "After that the source is saved permanently and can be rebuilt with one click."
        )

    filtered = apply_global_filters(display,"ecom")

    st.markdown("### Reconciliation Summary")

    a1,a2,a3,a4,a5,a6 = st.columns(6)
    a1.metric("Total Orders", f"{filtered['Order No'].nunique():,}")
    a2.metric(
        "Replacement Orders",
        f"{safe_text(filtered, 'Replacement Order Id').str.strip().ne('').sum():,}"
    )
    a3.metric(
        "Repl Quantity",
        f"{safe_numeric(filtered, 'Repl Quantity').sum():,.0f}"
    )
    a4.metric(
        "Replacement Item Price",
        f"₹{safe_numeric(filtered, 'Replacement Item Price').sum():,.2f}"
    )
    a5.metric(
        "Courier & Customer Return Qty",
        f"{safe_numeric(filtered, 'Courier & Customer Return Qty').sum():,.0f}"
    )
    a6.metric(
        "Order Value",
        f"₹{safe_numeric(filtered, 'Order Price').sum():,.2f}"
    )

    b1,b2,b3,b4,b5,b6,b7 = st.columns(7)
    b1.metric(
        "Rece Amount",
        f"₹{safe_numeric(filtered, 'Rece Amount').sum():,.2f}"
    )
    b2.metric(
        "Deferred Amount",
        f"₹{safe_numeric(filtered, 'Deferred Amount').sum():,.2f}"
    )
    b3.metric(
        "Refund",
        f"₹{safe_numeric(filtered, 'Refund').sum():,.2f}"
    )
    b4.metric(
        "Reimbursement",
        f"₹{safe_numeric(filtered, 'Reimbursement').sum():,.2f}"
    )
    adjustment_kpi = safe_numeric(filtered, "Adjustment").sum()
    if len(filtered) == len(display):
        exact_source_adjustment = source_adjustment_total()
        if exact_source_adjustment != 0:
            adjustment_kpi = exact_source_adjustment

    b5.metric(
        "Adjustment",
        f"₹{adjustment_kpi:,.2f}"
    )
    task_status_series = safe_text(filtered, "Task Status").str.strip()
    b6.metric(
        "Pending / Working",
        int(task_status_series.isin(["Pending","Working"]).sum())
    )
    b7.metric(
        "Completed Tasks",
        int((task_status_series=="Completed").sum())
    )

    st.caption(
        f"Replacement IDs: {safe_text(filtered, 'Replacement Order Id').str.strip().ne('').sum():,} | "
        f"Repl Qty: {safe_numeric(filtered, 'Repl Quantity').sum():,.0f} | "
        f"Courier Returns: {safe_numeric(filtered, 'Courier & Customer Return Qty').sum():,.0f} | "
        f"Rece Amount: ₹{safe_numeric(filtered, 'Rece Amount').sum():,.2f} | "
        f"Deferred Amount: ₹{safe_numeric(filtered, 'Deferred Amount').sum():,.2f} | "
        f"Adjustment KPI: ₹{adjustment_kpi:,.2f} | "
        f"Order-wise Adjustment: ₹{safe_numeric(filtered, 'Adjustment').sum():,.2f}"
    )

    st.markdown("### Order-wise Reconciliation")

    st.dataframe(
        filtered[ECOM_VISIBLE_COLUMNS],
        use_container_width=True,
        hide_index=True
    )

    exact = exact_dashboard_export(
        filtered,ECOM_VISIBLE_COLUMNS,ECOM_EXPORT_HEADERS
    )
    mapping = filtered[[
        "Market Place","Branch Code","Order No","Task Type","Task Assign To"
    ]].copy()

    st.download_button(
        "Download Filtered E-Com Reconciliation",
        workbook_bytes(exact,mapping,"E-Com Process Reco"),
        file_name=f"E-Com_Process_Reconciliation_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

elif workspace == "Management MIS":
    st.subheader("Management MIS & Exception Control")
    st.caption("Executive view of reconciliation health, branch performance, task SLA and exceptions from the same persistent Supabase data.")

    master_mis, tasks_mis = management_mis_frames(load_master(), load_tasks())
    if master_mis.empty:
        st.info("No reconciliation data stored yet.")
        st.stop()

    f1, f2, f3 = st.columns([1.1, 1.4, 0.8])
    portal_options = sorted([x for x in master_mis["__portal__"].dropna().unique().tolist() if x])
    branch_options = sorted(master_mis["__branch__"].dropna().unique().tolist())
    selected_portals = f1.multiselect("Marketplace", portal_options, default=portal_options, key="mis_marketplace")
    selected_branches = f2.multiselect("Branch", branch_options, default=[], placeholder="All branches", key="mis_branch")
    sla_days = f3.number_input("Task SLA (Days)", min_value=1, max_value=60, value=5, step=1, key="mis_sla_days")

    mf, tf = _mis_filter_frames(master_mis, tasks_mis, selected_portals, selected_branches)
    open_tasks = tf[tf["__status__"].isin(["Pending", "Working"])].copy() if not tf.empty else tf
    overdue_tasks = open_tasks[open_tasks["__aging__"] > sla_days].copy() if not open_tasks.empty else open_tasks
    completed_tasks = tf[tf["__status__"].eq("Completed")].copy() if not tf.empty else tf

    tab_exec, tab_branch, tab_sla, tab_ex = st.tabs(["Executive MIS", "Branch Performance", "Aging & SLA", "Exception Analytics"])

    with tab_exec:
        total_orders = int(mf["order_no"].nunique())
        reconciled_orders = int(mf.loc[mf["__reconciled__"], "order_no"].nunique()) if total_orders else 0
        reconciliation_rate = reconciled_orders / total_orders * 100 if total_orders else 0
        return_orders = int(mf.loc[mf["__has_return__"], "order_no"].nunique()) if total_orders else 0
        return_rate = return_orders / total_orders * 100 if total_orders else 0
        order_value = pd.to_numeric(mf.get("order_price", 0), errors="coerce").fillna(0).sum()
        received = pd.to_numeric(mf.get("net_pay_received", 0), errors="coerce").fillna(0).sum()
        refunds = pd.to_numeric(mf.get("refund", 0), errors="coerce").fillna(0).sum()
        deferred = pd.to_numeric(mf.get("deferred_amount", 0), errors="coerce").fillna(0).sum()
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total Orders", f"{total_orders:,}")
        c2.metric("Reconciliation Rate", f"{reconciliation_rate:.1f}%")
        c3.metric("Open Tasks", f"{len(open_tasks):,}")
        c4.metric("Overdue > SLA", f"{len(overdue_tasks):,}")
        c5.metric("Return Rate", f"{return_rate:.1f}%")
        c6.metric("Completed Tasks", f"{len(completed_tasks):,}")
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Order Value", f"₹{order_value:,.2f}")
        m2.metric("Received Amount", f"₹{received:,.2f}")
        m3.metric("Refund", f"₹{refunds:,.2f}")
        m4.metric("Deferred Amount", f"₹{deferred:,.2f}")
        st.markdown("#### Marketplace Health")
        market = mf.groupby("__portal__", dropna=False).agg(Orders=("order_no","nunique"), Reconciled=("__reconciled__","sum"), Return_Orders=("__has_return__","sum")).reset_index().rename(columns={"__portal__":"Marketplace"})
        market["Reconciliation %"] = (market["Reconciled"] / market["Orders"].replace(0,pd.NA) * 100).fillna(0).round(1)
        if not tf.empty:
            mt = tf.groupby("__portal__", dropna=False).agg(Total_Tasks=("order_no","size"), Open_Tasks=("__status__", lambda s: s.isin(["Pending","Working"]).sum())).reset_index().rename(columns={"__portal__":"Marketplace"})
            market = market.merge(mt, on="Marketplace", how="left")
        st.dataframe(market.fillna(0), use_container_width=True, hide_index=True)
        exc = mf.loc[~mf["__reconciled__"], "__exception__"].value_counts().rename_axis("Exception").to_frame("Orders")
        if not exc.empty:
            st.markdown("#### Top Open Exceptions")
            st.bar_chart(exc.head(10))

    with tab_branch:
        st.markdown("#### Branch Performance Scorecard")
        branch = mf.groupby("__branch__", dropna=False).agg(Orders=("order_no","nunique"), Reconciled=("__reconciled__","sum"), Return_Orders=("__has_return__","sum"), Order_Value=("order_price","sum"), Received=("net_pay_received","sum")).reset_index().rename(columns={"__branch__":"Branch"})
        branch["Reconciliation %"] = (branch["Reconciled"] / branch["Orders"].replace(0,pd.NA) * 100).fillna(0)
        branch["Return %"] = (branch["Return_Orders"] / branch["Orders"].replace(0,pd.NA) * 100).fillna(0)
        if not tf.empty:
            bt = tf.groupby("__branch__", dropna=False).agg(Total_Tasks=("order_no","size"), Open_Tasks=("__status__", lambda s: s.isin(["Pending","Working"]).sum()), Completed_Tasks=("__status__", lambda s: s.eq("Completed").sum()), Avg_Aging_Days=("__aging__","mean")).reset_index().rename(columns={"__branch__":"Branch"})
            ob = overdue_tasks.groupby("__branch__").size().rename("Overdue_Tasks").reset_index().rename(columns={"__branch__":"Branch"}) if not overdue_tasks.empty else pd.DataFrame(columns=["Branch","Overdue_Tasks"])
            branch = branch.merge(bt.merge(ob,on="Branch",how="left"), on="Branch", how="left")
        for col in ["Total_Tasks","Open_Tasks","Completed_Tasks","Overdue_Tasks","Avg_Aging_Days"]:
            if col not in branch.columns: branch[col] = 0
        branch["Task Completion %"] = (branch["Completed_Tasks"] / branch["Total_Tasks"].replace(0,pd.NA) * 100).fillna(0)
        branch["SLA Compliance %"] = ((branch["Open_Tasks"]-branch["Overdue_Tasks"]).clip(lower=0) / branch["Open_Tasks"].replace(0,pd.NA) * 100).fillna(100)
        cols=["Branch","Orders","Reconciliation %","Return %","Open_Tasks","Overdue_Tasks","Completed_Tasks","Task Completion %","SLA Compliance %","Avg_Aging_Days","Order_Value","Received"]
        branch=branch.sort_values(["Overdue_Tasks","Open_Tasks","Orders"],ascending=[False,False,False])
        st.dataframe(branch[cols].round(2), use_container_width=True, hide_index=True)
        st.markdown("#### Branch Open vs Overdue Tasks")
        st.bar_chart(branch.set_index("Branch")[["Open_Tasks","Overdue_Tasks"]])

    with tab_sla:
        st.markdown(f"#### Open Task SLA — Target: {int(sla_days)} Days")
        if tf.empty:
            st.info("No task data available.")
        else:
            active=tf[tf["__status__"].isin(["Pending","Working"])].copy()
            labels=["0–2 Days","3–5 Days","6–10 Days","11–20 Days","21+ Days"]
            active["Aging Bucket"] = pd.cut(active["__aging__"], bins=[-1,2,5,10,20,10**9], labels=labels)
            aging=active["Aging Bucket"].value_counts(sort=False).reindex(labels,fill_value=0).rename_axis("Aging Bucket").to_frame("Open Tasks")
            st.bar_chart(aging)
            s1,s2,s3,s4=st.columns(4)
            within=max(len(active)-len(overdue_tasks),0)
            compliance=within/len(active)*100 if len(active) else 100
            s1.metric("Open Tasks",f"{len(active):,}")
            s2.metric("Within SLA",f"{within:,}")
            s3.metric("Overdue",f"{len(overdue_tasks):,}")
            s4.metric("SLA Compliance",f"{compliance:.1f}%")
            if not overdue_tasks.empty:
                show=overdue_tasks[[c for c in ["__portal__","__branch__","order_no","task_type","__status__","__aging__","task_assign_to","task_assign_email","team_remarks"] if c in overdue_tasks.columns]].copy().rename(columns={"__portal__":"Marketplace","__branch__":"Branch","order_no":"Order No","task_type":"Task Type","__status__":"Task Status","__aging__":"Aging Days","task_assign_to":"Task Assign To","task_assign_email":"Owner Email","team_remarks":"Team Remarks"})
                st.dataframe(show.sort_values("Aging Days",ascending=False), use_container_width=True, hide_index=True)

    with tab_ex:
        open_ex=mf[~mf["__reconciled__"]].copy()
        st.markdown("#### Exception Mix")
        exs=open_ex.groupby("__exception__",dropna=False).agg(Orders=("order_no","nunique"),Order_Value=("order_price","sum"),Received=("net_pay_received","sum"),Return_Qty=("courier_customer_return_qty","sum")).reset_index().rename(columns={"__exception__":"Exception"}).sort_values("Orders",ascending=False)
        st.dataframe(exs.round(2), use_container_width=True, hide_index=True)
        x1,x2,x3,x4=st.columns(4)
        x1.metric("Open Exception Orders",f"{open_ex['order_no'].nunique():,}")
        x2.metric("Payment-related",f"{open_ex['__exception__'].isin(['Payment Pending','Short Payment']).sum():,}")
        x3.metric("Return/TEI/CN-related",f"{open_ex['__exception__'].isin(['TEI Pending','Partial TEI / Balance Pending','CN Pending','MIR Pending']).sum():,}")
        x4.metric("Billing-related",f"{open_ex['__exception__'].eq('Billing Pending').sum():,}")
        st.markdown("#### High-Value Exception Orders")
        high=open_ex.copy(); high["__value__"]=pd.to_numeric(high.get("order_price",0),errors="coerce").fillna(0)
        cols=["__portal__","__branch__","order_no","order_item","__exception__","pending_remarks","payment_status","__value__","net_pay_received","refund","courier_customer_return_qty","return_completed_date"]
        show=high[[c for c in cols if c in high.columns]].rename(columns={"__portal__":"Marketplace","__branch__":"Branch","order_no":"Order No","order_item":"Order Item","__exception__":"Exception","pending_remarks":"Pending Remarks","payment_status":"Payment Status","__value__":"Order Value","net_pay_received":"Rece Amount","refund":"Refund","courier_customer_return_qty":"Return Qty","return_completed_date":"Return Delivery Date"})
        st.dataframe(show.sort_values("Order Value",ascending=False).head(200), use_container_width=True, hide_index=True)

elif workspace == "Settlement Dashboard":
    st.subheader("Settlement Dashboard")
    st.caption(
        "Format follows the attached Settlement Reco sheet. "
        "Total Deductions is populated only when an explicit deduction field "
        "is present in the uploaded payment source."
    )

    display = settlement_process_display(load_master())

    if display.empty:
        st.info("No reconciliation data stored yet.")
        st.stop()

    # Settlement dashboard intentionally uses only the four requested filters.
    filtered = apply_settlement_date_filters(display)

    c1,c2,c3,c4,c5,c6,c7,c8 = st.columns(8)
    c1.metric("Total Orders",f"{filtered['Order No'].nunique():,}")
    c2.metric("Order Value",f"₹{safe_numeric(filtered, 'Order Price').sum():,.0f}")
    c3.metric("Sale Value",f"₹{safe_numeric(filtered, 'Sale Price').sum():,.0f}")
    c4.metric("Return Value",f"₹{safe_numeric(filtered, 'Return Price').sum():,.0f}")
    c5.metric("Rece Amount",f"₹{safe_numeric(filtered, 'Rece Amount').sum():,.0f}")
    c6.metric("Total Deductions",f"₹{safe_numeric(filtered, 'Total Deductions').sum():,.0f}")
    c7.metric("Reconciled",int((safe_text(filtered, "Transaction Status")=="Reconciled").sum()))
    c8.metric("Pending / Review",int((safe_text(filtered, "Transaction Status")!="Reconciled").sum()))

    st.dataframe(
        filtered[SETTLEMENT_VISIBLE_COLUMNS],
        use_container_width=True,
        hide_index=True
    )

    exact = exact_dashboard_export(
        filtered,SETTLEMENT_VISIBLE_COLUMNS,SETTLEMENT_EXPORT_HEADERS
    )
    mapping_cols = [
        c for c in [
            "Market Place","Branch Code","Order No",
            "Task Type","Task Status","Task Assign To"
        ] if c in filtered.columns
    ]
    mapping = filtered[mapping_cols].copy()

    st.download_button(
        "Download Filtered Settlement Reconciliation",
        workbook_bytes(exact,mapping,"Settlement Reco"),
        file_name=f"Settlement_Process_Reconciliation_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

elif workspace == "Pending Task Dashboard":
    st.subheader("Pending Task Dashboard")

    display = ecom_process_display(load_master())
    if display.empty:
        st.info("No reconciliation/task data stored yet.")
        st.stop()

    tasks_view = display[
        (display["Task Type"].fillna("").astype(str).str.strip() != "")
        & (
            display["Pending Remarks"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            != "reconciled"
        )
    ].copy()
    blank_status = tasks_view["Task Status"].fillna("").astype(str).str.strip() == ""
    tasks_view.loc[blank_status, "Task Status"] = "Pending"


    show_completed = st.checkbox("Include Completed Tasks", value=False)
    if not show_completed:
        tasks_view = tasks_view[tasks_view["Task Status"].fillna("").astype(str) != "Completed"]
    filtered = apply_global_filters(tasks_view,"pending")

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("All Tasks",len(filtered))
    c2.metric("Pending",int((filtered["Task Status"]=="Pending").sum()))
    c3.metric("Working",int((filtered["Task Status"]=="Working").sum()))
    c4.metric("Completed",int((filtered["Task Status"]=="Completed").sum()))

    cols = [
        "Market Place","Branch Code","Order No","Task Type","Pending Remarks",
        "Task Status","Pending Task Created Date",
        "Aging Days of Task Pending/Completed","Task Assign To",
        "Team Remarks","Task Completed Date"
    ]
    st.dataframe(
        filtered[[c for c in cols if c in filtered.columns]],
        use_container_width=True,
        hide_index=True
    )


    if st.download_button(
        "Download Filtered Pending Tasks for Team",
        pending_download(filtered),
        file_name=f"Pending_Tasks_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    ):
        log_activity("PENDING TASK DOWNLOAD",details=f"{len(filtered)} filtered task rows downloaded")

    st.info("Download the filtered Pending Task workbook here. Only MIR No, MIR Date, TEI No, TEI Date, TEI Qty, Ticket Raised If Any, Ticket Raised Date, Team Remarks and Task Completed Date are editable. Upload the completed workbook under the corresponding MIR branch.")

# ------------------------------------------------------------
# ALL BRANCH PENDING TASK WORKING UPDATE
# ------------------------------------------------------------
elif workspace == "All Branch Pending Task Update":
    st.subheader("All Branch Code Pending Task Working Update By Team")
    st.info(
        "Use this workspace when one team member needs to update pending-task working "
        "for multiple branch codes in one file. Branch Code and reconciliation/source "
        "columns remain locked; only approved team-working columns are editable. "
        "Saved changes automatically flow to the E-Com Reconciliation Dashboard, "
        "Settlement Dashboard, Pending Task Dashboard and the individual MIR branch dashboards."
    )

    all_branch_display = ecom_process_display(load_master())
    if all_branch_display.empty:
        st.info("No reconciliation data stored yet.")
    else:
        all_branch_tasks = all_branch_display[
            all_branch_display["Task Type"].fillna("").astype(str).str.strip() != ""
        ].copy()

        blank_status = (
            all_branch_tasks["Task Status"].fillna("").astype(str).str.strip() == ""
        )
        all_branch_tasks.loc[blank_status, "Task Status"] = "Pending"
        all_branch_tasks = all_branch_tasks[
            all_branch_tasks["Task Status"].fillna("").astype(str).str.strip()
            .isin(["Pending", "Working"])
        ].copy()

        filtered_all_branch = apply_global_filters(
            all_branch_tasks, "all_branch_task_update"
        )

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("All Active Tasks", len(filtered_all_branch))
        c2.metric(
            "Branches",
            int(filtered_all_branch["Branch Code"].fillna("").astype(str).str.strip()
                .replace("", pd.NA).dropna().nunique())
        )
        c3.metric(
            "Pending",
            int((safe_text(filtered_all_branch,"Task Status").str.strip()=="Pending").sum())
        )
        c4.metric(
            "Working",
            int((safe_text(filtered_all_branch,"Task Status").str.strip()=="Working").sum())
        )

        all_branch_view = exact_dashboard_export(
            filtered_all_branch, ECOM_VISIBLE_COLUMNS, ECOM_EXPORT_HEADERS
        )
        st.dataframe(all_branch_view, use_container_width=True, hide_index=True)

        st.download_button(
            "Download All Branch Pending Tasks Working Sheet",
            pending_download(filtered_all_branch),
            file_name=f"All_Branch_Pending_Tasks_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()
    st.markdown("### Upload All Branch Working Sheet")
    st.caption(
        "Upload the protected XLSX downloaded above. You may update multiple branch codes "
        "in the same workbook. Editable fields are MIR No, MIR Date, TEI No, TEI Date, TEI Qty, "
        "Ticket Raised If Any, Ticket Raised Date, Team Remarks and Task Completed Date. "
        "Task Status changes automatically to Working/Completed."
    )

    all_branch_team_file = st.file_uploader(
        "Upload All Branch Pending Task Working Sheet",
        type=["xlsx"],
        key="team_update_all_branches",
    )
    if all_branch_team_file and st.button(
        "Apply All Branch Working Updates",
        type="primary",
        use_container_width=True,
        key="apply_team_update_all_branches",
    ):
        try:
            n, unmatched = apply_branch_team_workbook(all_branch_team_file, "ALL")
            log_activity(
                "ALL BRANCH TEAM UPDATE",
                branch_code="ALL",
                filename=all_branch_team_file.name,
                details=f"{n} task record(s) updated across branch codes",
            )
            if n > 0:
                msg = (
                    f"Upload complete — {n} task record(s) updated permanently across branch codes. "
                    "All dashboards have been refreshed."
                )
                notice_kind = "success"
            else:
                msg = (
                    "Upload processed, but no task records were changed. "
                    "Please verify that editable team-working fields contain new values."
                )
                notice_kind = "warning"

            if unmatched:
                st.warning(
                    f"{len(unmatched)} order(s) could not be matched: "
                    + ", ".join(unmatched[:20])
                )

            invalidate_read_cache()
            queue_completion_notice(msg, notice_kind)
            st.rerun()
        except Exception as exc:
            st.error(f"Could not apply all-branch working sheet: {exc}")

# ------------------------------------------------------------
# MIR BRANCH DASHBOARDS
# ------------------------------------------------------------
elif workspace in ["MIR 1600","MIR 5400","MIR 5600","MIR 7800"]:
    branch = workspace.replace("MIR ","")
    st.subheader(f"MIR {branch} — Pending Tasks (Order-wise Reconciliation)")
    st.info(f"Responsible: {MIR_BRANCH_PEOPLE[branch]}")

    st.write(
        "This branch shows only active Pending / Working tasks, using the exact "
        "same Order-wise Reconciliation format as the main E-Com dashboard for "
        "viewing, downloading and team upload. "
        "The team can edit only MIR No, MIR Date, TEI No, TEI Date, TEI Qty, "
        "Ticket Raised If Any, Ticket Raised Date, Team Remarks and "
        "Task Completed Date. All other reconciliation columns remain locked."
    )

    # --------------------------------------------------------
    # A) BRANCH PENDING TASK DOWNLOAD
    # --------------------------------------------------------
    branch_display = ecom_process_display(load_master())

    if branch_display.empty:
        st.info("No reconciliation data stored yet.")
    else:
        branch_tasks = branch_display[
            branch_display["Branch Code"].fillna("").astype(str).str.strip()
            == str(branch)
        ].copy()

        branch_tasks = branch_tasks[
            branch_tasks["Task Type"].fillna("").astype(str).str.strip() != ""
        ].copy()

        blank_status = (
            branch_tasks["Task Status"]
            .fillna("")
            .astype(str)
            .str.strip()
            == ""
        )
        branch_tasks.loc[blank_status, "Task Status"] = "Pending"

        # MIR workspaces show ONLY active pending tasks.
        # Completed tasks remain preserved in history/main reconciliation
        # but are not shown in the MIR team workspace.
        branch_tasks = branch_tasks[
            branch_tasks["Task Status"]
            .fillna("")
            .astype(str)
            .str.strip()
            .isin(["Pending", "Working"])
        ].copy()

        # Same filters as Pending Task Dashboard.
        filtered_branch_tasks = apply_global_filters(
            branch_tasks, f"branch_task_{branch}"
        )

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("All Tasks", len(filtered_branch_tasks))
        c2.metric(
            "Pending",
            int(
                (
                    safe_text(filtered_branch_tasks, "Task Status")
                    .str.strip()
                    == "Pending"
                ).sum()
            ),
        )
        c3.metric(
            "Working",
            int(
                (
                    safe_text(filtered_branch_tasks, "Task Status")
                    .str.strip()
                    == "Working"
                ).sum()
            ),
        )
        c4.metric(
            "Completed",
            int(
                (
                    safe_text(filtered_branch_tasks, "Task Status")
                    .str.strip()
                    == "Completed"
                ).sum()
            ),
        )

        # Exact same visible format as Order-wise Reconciliation.
        branch_orderwise_view = exact_dashboard_export(
            filtered_branch_tasks,
            ECOM_VISIBLE_COLUMNS,
            ECOM_EXPORT_HEADERS,
        )

        st.dataframe(
            branch_orderwise_view,
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            f"Download MIR {branch} Pending Tasks",
            protected_branch_pending_download(filtered_branch_tasks, branch),
            file_name=f"MIR_{branch}_Pending_Orderwise_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()

    # --------------------------------------------------------
    # B) TEAM WORKING SHEET UPLOAD — same logic as old
    #    Branch MIR Task Update, but inside each MIR branch.
    # --------------------------------------------------------
    st.markdown("### Upload MIR Order-wise Working Sheet")
    st.caption(
        "Upload the same protected Order-wise Reconciliation XLSX downloaded above. "
        "Editable columns: MIR No, MIR Date, TEI No, TEI Date, TEI Qty, "
        "Ticket Raised If Any, Ticket Raised Date, Team Remarks and "
        "Task Completed Date. Task Status is updated automatically."
    )

    branch_team_file = st.file_uploader(
        f"Upload MIR {branch} Order-wise Working Sheet",
        type=["xlsx"],
        key=f"team_update_{branch}",
    )

    if branch_team_file and st.button(
        f"Apply MIR {branch} Order-wise Updates",
        type="primary",
        use_container_width=True,
        key=f"apply_team_update_{branch}",
    ):
        try:
            n, unmatched = apply_branch_team_workbook(branch_team_file, branch)

            log_activity(
                "TEAM UPDATE",
                branch_code=branch,
                filename=branch_team_file.name,
                details=f"{n} task record(s) updated",
            )

            if n > 0:
                msg = (
                    f"Upload complete — {n} task record(s) updated permanently for branch {branch}. "
                    "Main reconciliation and dashboards have been refreshed."
                )
                notice_kind = "success"
            else:
                msg = (
                    f"Upload processed for branch {branch}, but no task records were changed. "
                    "Please verify that editable team-working fields contain new values."
                )
                notice_kind = "warning"

            if unmatched:
                st.warning(
                    f"{len(unmatched)} order(s) could not be matched: "
                    + ", ".join(unmatched[:20])
                )

            invalidate_read_cache()
            queue_completion_notice(msg, notice_kind)
            st.rerun()

        except Exception as exc:
            st.error(str(exc))

    st.divider()

    # --------------------------------------------------------
    # C) SAVED MIR / ORDER-WISE DETAIL VIEW
    # The separate legacy MIR sheet uploader has been removed.
    # Team updates now use only the protected Order-wise Reconciliation workbook.

    # C) SAVED MIR / ORDER-WISE DETAIL VIEW
    # --------------------------------------------------------
    st.markdown(f"### MIR {branch} — Current Pending Tasks")

    current_branch = ecom_process_display(load_master())

    if current_branch.empty:
        st.info("No reconciliation data stored yet.")
    else:
        current_branch = current_branch[
            current_branch["Branch Code"].fillna("").astype(str).str.strip()
            == str(branch)
        ].copy()

        # Only task rows.
        current_branch = current_branch[
            current_branch["Task Type"]
            .fillna("")
            .astype(str)
            .str.strip()
            != ""
        ].copy()

        # Blank task status means a newly detected pending task.
        current_blank_status = (
            current_branch["Task Status"]
            .fillna("")
            .astype(str)
            .str.strip()
            == ""
        )
        current_branch.loc[current_blank_status, "Task Status"] = "Pending"

        # Exclude Completed from MIR team workspace.
        current_branch = current_branch[
            current_branch["Task Status"]
            .fillna("")
            .astype(str)
            .str.strip()
            .isin(["Pending", "Working"])
        ].copy()

        filtered_current_branch = apply_global_filters(
            current_branch,
            f"mir_current_{branch}"
        )

        current_orderwise = exact_dashboard_export(
            filtered_current_branch,
            ECOM_VISIBLE_COLUMNS,
            ECOM_EXPORT_HEADERS,
        )

        st.dataframe(
            current_orderwise,
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            f"Download Current MIR {branch} Order-wise Reconciliation",
            protected_branch_pending_download(
                filtered_current_branch,
                branch
            ),
            file_name=(
                f"MIR_{branch}_Current_Orderwise_"
                f"{date.today().isoformat()}.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

elif workspace == "Branch Owner & Email Setup":
    if not is_admin_unlocked():
        st.warning("Admin PIN required to change Branch Owner or Microsoft 365 settings.")
        st.stop()
    else:
        st.subheader("Branch Owner & Email Setup")
        st.caption(
            "Marketplace is not required. The same Branch + Task Type assignment applies "
            "to Amazon and Flipkart."
        )

        owner_template = pd.DataFrame([
            ["1600","CN Pending","Rajan","ss6@glenindia.com"],
            ["1600","CN Pending","Sashikant","ss3@glenindia.com"],
            ["5400","CN Pending","Karthik","bdc1@glenindia.com"],
            ["5400","CN Pending","Sonia","reachus@glenindia.com"],
            ["5600","CN Pending","Jitendra","kolkata_ops@transagile.in"],
            ["5600","CN Pending","Sonia","reachus@glenindia.com"],
            ["7800","CN Pending","Sachin","mdc@glenindia.com"],
            ["7800","CN Pending","Sonia","reachus@glenindia.com"],
            ["All Branch","Payment Pending","Rahul","online5@glenindia.com"],
        ],columns=["Branch","Task Type","Owner Name","Owner Email"])

        st.download_button(
            "Download Branch Owner Template",
            workbook_bytes(owner_template,None,"Branch Owner Master"),
            file_name="Branch_Owner_Master.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        owner_file = st.file_uploader(
            "Upload Branch Owner Master",
            type=["xlsx","csv"],
            key="owner_upload"
        )

        if owner_file and st.button("Save Owner Master",type="primary",use_container_width=True):
            try:
                x = (
                    pd.read_csv(owner_file)
                    if owner_file.name.lower().endswith(".csv")
                    else pd.read_excel(owner_file)
                )

                n = save_owner_rules(x)
                log_activity("OWNER MASTER UPDATE",filename=owner_file.name,details=f"{n} mappings saved")
                st.success(f"{n} owner mapping(s) saved permanently.")
            except Exception as exc:
                st.error(str(exc))

        rules = load_owner_rules()
        if not rules.empty:
            st.markdown("### Saved Branch Owner Mappings")
            st.dataframe(rules,use_container_width=True,hide_index=True)

        st.markdown("### Microsoft 365 Email Setup")
        cfg = load_m365_config()
        c1,c2 = st.columns(2)
        tenant = c1.text_input("Tenant ID",value=cfg.get("tenant_id",""))
        client = c2.text_input("Client ID",value=cfg.get("client_id",""))
        sender = st.text_input("Sender Email",value=cfg.get("sender_email",""))
        auto_send = st.checkbox(
            "Automatically email owners after reconciliation upload",
            value=bool(cfg.get("auto_send",False))
        )

        if st.button("Save Microsoft 365 Settings",use_container_width=True):
            save_m365_config(tenant,client,sender,auto_send)
            st.success("Microsoft 365 settings saved permanently.")

        if st.button("Start Microsoft Sign-In",type="primary",use_container_width=True):
            try:
                save_m365_config(tenant,client,sender,auto_send)
                st.session_state["m365_flow"] = begin_m365_login()
            except Exception as exc:
                st.error(str(exc))

        flow = st.session_state.get("m365_flow")
        if flow:
            st.info(flow.get("message","Follow Microsoft sign-in instructions."))
            st.code(flow.get("user_code",""))
            st.link_button(
                "Open Microsoft Device Login",
                flow.get("verification_uri","https://microsoft.com/devicelogin"),
                use_container_width=True
            )
            if st.button("Complete Microsoft Sign-In",use_container_width=True):
                try:
                    result = complete_m365_login(flow)
                    st.session_state.pop("m365_flow",None)
                    st.success("Microsoft 365 sign-in completed.")
                except Exception as exc:
                    st.error(str(exc))

        if st.button("Send Pending Task Emails Now",use_container_width=True):
            try:
                n = send_pending_emails()
                st.success(f"Email sent to {n} owner(s).")
            except Exception as exc:
                st.error(str(exc))

    # ------------------------------------------------------------
    # UPLOAD HISTORY
    # ------------------------------------------------------------
elif workspace == "Upload History":
    st.subheader("Upload History")
    with db() as c:
        history = pd.read_sql_query("""
            SELECT portal,filename,uploaded_at,rows_processed
            FROM upload_history
            ORDER BY id DESC
        """,c)

    if history.empty:
        st.info("No portal uploads recorded yet.")
    else:
        st.dataframe(history,use_container_width=True,hide_index=True)
        st.download_button(
            "Download Upload History",
            workbook_bytes(history,None,"Upload History"),
            file_name="Upload_History.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


elif workspace == "Activity Audit Trail":
    st.subheader("Activity Audit Trail")
    st.caption("Read-only history of source uploads, task creation, team uploads, MIR uploads and owner-master changes.")
    audit=load_audit()
    if audit.empty:
        st.info("No activity recorded yet.")
    else:
        c1,c2,c3,c4=st.columns(4)
        action=c1.multiselect("Action",sorted([x for x in audit["action_type"].dropna().astype(str).unique() if x]))
        portal=c2.multiselect("Market Place",sorted([x for x in audit["portal"].dropna().astype(str).unique() if x]))
        branch=c3.multiselect("Branch",sorted([x for x in audit["branch_code"].dropna().astype(str).unique() if x]))
        order=c4.text_input("Order ID")
        view=audit.copy()
        if action: view=view[view["action_type"].isin(action)]
        if portal: view=view[view["portal"].isin(portal)]
        if branch: view=view[view["branch_code"].isin(branch)]
        if order: view=view[view["order_no"].astype(str).str.contains(order,case=False,na=False)]
        st.dataframe(view,use_container_width=True,hide_index=True)
        st.download_button(
            "Download Audit Trail",
            workbook_bytes(view,None,"Audit Trail"),
            file_name=f"Reconciliation_Audit_Trail_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()
if using_postgres():
    try:
        source_bits = []
        for item in source_workbook_info():
            size_mb = item["size"] / (1024 * 1024) if item["size"] else 0
            source_bits.append(
                f"{item['portal']}: {item['filename']} ({size_mb:.1f} MB, {item['uploaded_at']})"
            )
        source_text = " | ".join(source_bits) if source_bits else "No source snapshot saved yet"
    except Exception:
        source_text = "Source registry unavailable"
    st.caption(
        "Persistent storage: Supabase PostgreSQL | " + source_text
    )
else:
    st.caption(
        f"Local development storage: {DATA_DIR}. "
        "For Streamlit Cloud permanence, configure DATABASE_URL in Streamlit Secrets."
    )
